# -*- coding: utf-8 -*-
"""
CRM Controllers
Request handling layer for CRM operations
"""
from pydoc import text
from venv import logger

from flask import request, jsonify, g
from typing import Dict, Any
from backend.crm.services.crm_service import CRMService


class CRMController:
    """
    CRM Controller
    Handles HTTP requests and responses for CRM operations
    """
    
    def __init__(self):
        self.crm_service = CRMService()
    
    # ========================================
    # LEAD ENDPOINTS
    # ========================================
    
    def get_leads(self) -> tuple:
        """
        GET /api/crm/leads
        Get all leads for the current tenant
        """
        try:
            tenant_id = g.tenant_id
            
            # Extract query parameters for filtering
            filters = {}
            if request.args.get('stage_id'):
                filters['stage_id'] = int(request.args.get('stage_id'))
            if request.args.get('stage'):
                filters['stage'] = request.args.get('stage')
            if request.args.get('exclude_stage'):
                filters['exclude_stage'] = request.args.get('exclude_stage')
            if request.args.get('status'):
                filters['status'] = request.args.get('status')
            if request.args.get('assigned_to'):
                filters['assigned_to'] = int(request.args.get('assigned_to'))
            # Service filter: electricity=1, water=2
            service_param = request.args.get('service')
            if service_param and isinstance(service_param, str):
                svc = service_param.strip().lower()
                if svc == 'water':
                    filters['service_id'] = 2
                elif svc == 'electricity':
                    filters['service_id'] = 1
            
            result = self.crm_service.get_leads(tenant_id, filters if filters else None)
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def get_lead_detail(self, opportunity_id: int) -> tuple:
        """
        GET /api/crm/leads/<opportunity_id>
        Get details of a specific lead
        """
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_lead_detail(tenant_id, opportunity_id)
            
            if not result.get('success'):
                return jsonify(result), 404
            
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def update_lead_status(self, opportunity_id: int) -> tuple:
        """
        PATCH /api/crm/leads/<opportunity_id>/status
        Update lead status (stage_id) only.
        When stage becomes 'Lost', lead is soft-deleted (deleted_at=NOW()).
        """
        try:
            tenant_id = g.tenant_id
            payload = request.get_json()
            
            if not payload:
                return jsonify({
                    'success': False,
                    'error': 'Invalid request',
                    'message': 'Request body is required'
                }), 400
            
            stage_id = payload.get('stage_id')
            if stage_id is None:
                return jsonify({
                    'success': False,
                    'error': 'Validation error',
                    'message': 'stage_id is required'
                }), 400
            
            try:
                stage_id = int(stage_id)
            except (ValueError, TypeError):
                return jsonify({
                    'success': False,
                    'error': 'Validation error',
                    'message': 'stage_id must be a number'
                }), 400
            
            # Validate that stage_id exists in Stage_Master
            from backend.crm.supabase_client import get_supabase_client
            db = get_supabase_client()
            stage_check = db.execute_query(
                'SELECT "stage_id" FROM "StreemLyne_MT"."Stage_Master" WHERE "stage_id" = %s',
                (stage_id,),
                fetch_one=True
            )
            if not stage_check:
                return jsonify({
                    'success': False,
                    'error': 'Validation error',
                    'message': 'Invalid stage_id'
                }), 400
            
            result = self.crm_service.update_lead_status(tenant_id, opportunity_id, stage_id)
            
            if not result.get('success'):
                return jsonify(result), 404
            
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500

    def get_recycle_bin(self) -> tuple:
        """
        GET /api/crm/leads/recycle-bin
        Get all soft-deleted (Lost) leads for the tenant.
        """
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_recycle_bin(tenant_id)
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def delete_expired_lost_leads(self) -> tuple:
        """
        PATCH /api/crm/leads/cleanup
        Permanently delete Lost leads older than N days.
        Admin operation (controlled by token_required + tenant_from_jwt).
        """
        try:
            tenant_id = g.tenant_id
            payload = request.get_json() or {}
            days = payload.get('days', 30)
            
            try:
                days = int(days)
            except (ValueError, TypeError):
                days = 30
            
            result = self.crm_service.delete_expired_lost_leads(tenant_id, days)
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500

    def create_lead(self) -> tuple:
        """
        POST /api/crm/leads
        Create a new lead — DEPRECATED for manual creation.
        Leads must be created only via the Excel import-confirm endpoint
        (`POST /api/crm/leads/import/confirm`).

        The API will return a 400 validation error for direct lead creation.
        Tenant ID is taken from `g.tenant_id` (middleware enforces presence).
        """
        try:
            tenant_id = g.tenant_id
            payload = request.get_json()

            if not payload:
                return jsonify({
                    'success': False,
                    'error': 'Invalid request',
                    'message': 'Request body is required'
                }), 400

            # Basic validation: either client_id OR client payload/company name must exist
            if not payload.get('client_id') and not (payload.get('client') or payload.get('client_company_name') or payload.get('business_name')):
                return jsonify({
                    'success': False,
                    'error': 'Validation error',
                    'message': 'Provide client_id OR client (object) / business_name in the request body.'
                }), 400

            result = self.crm_service.create_lead(tenant_id, payload)

            # Service returns structured error info
            if not result.get('success'):
                status_code = 400 if result.get('error') and result.get('error').lower().startswith('validation') else 500
                return jsonify(result), status_code

            return jsonify(result), 201
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def update_lead(self, opportunity_id: int) -> tuple:
        """
        PUT /api/crm/leads/<opportunity_id>
        Update an existing lead
        """
        try:
            tenant_id = g.tenant_id
            lead_data = request.get_json()
            
            if not lead_data:
                return jsonify({
                    'success': False,
                    'error': 'Invalid request',
                    'message': 'Request body is required'
                }), 400
            
            result = self.crm_service.update_lead(tenant_id, opportunity_id, lead_data)
            
            if not result.get('success'):
                return jsonify(result), 404
            
            return jsonify(result), 200
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def assign_leads(self) -> tuple:
        """
        PATCH /api/crm/leads/assign
        Bulk assign leads to an employee.
        """
        try:
            from flask import request, jsonify, g
            from backend.db import SessionLocal
            from sqlalchemy import text
            import logging
            
            logger = logging.getLogger(__name__)

            user = getattr(request, 'current_user', None)
            if not user:
                return jsonify({'success': False, 'error': 'Authentication required'}), 401

            tenant_id = str(g.tenant_id)
            payload = request.get_json()
            if not payload:
                return jsonify({'success': False, 'error': 'Request body required'}), 400

            lead_ids = payload.get('lead_ids')
            employee_id = payload.get('employee_id') or payload.get('assigned_to_id')
            assignment_notes = payload.get('assignment_notes', '')
            
            if not lead_ids or not isinstance(lead_ids, list) or len(lead_ids) == 0:
                return jsonify({'success': False, 'error': 'lead_ids must be a non-empty list'}), 400
            if employee_id is None:
                return jsonify({'success': False, 'error': 'employee_id or assigned_to_id required'}), 400
            try:
                employee_id = int(employee_id)
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': 'employee_id must be a number'}), 400

            session = SessionLocal()
            
            try:
                # ✅ Get employee name for response
                employee_name = "Unassigned"
                if employee_id:
                    emp_result = session.execute(text("""
                        SELECT employee_name FROM "StreemLyne_MT"."Employee_Master"
                        WHERE employee_id = :id AND tenant_id = :tid
                    """), {'id': employee_id, 'tid': tenant_id}).fetchone()
                    if emp_result:
                        employee_name = emp_result[0]
                
                # Determine is_allocated flag
                current_user_employee_id = getattr(user, 'employee_id', None)
                is_allocated = (current_user_employee_id is None) or (employee_id != current_user_employee_id)
                
                # Perform the assignment
                for lead_id in lead_ids:
                    session.execute(text("""
                        UPDATE "StreemLyne_MT"."Opportunity_Details"
                        SET opportunity_owner_employee_id = :emp_id,
                            is_allocated = :is_allocated,
                            is_draft = FALSE
                        WHERE opportunity_id = :id AND tenant_id = :tid
                    """), {
                        'emp_id': employee_id, 
                        'id': lead_id, 
                        'tid': tenant_id,
                        'is_allocated': is_allocated
                    })
                    
                    # Log assignment in history
                    if assignment_notes:
                        from datetime import datetime
                        try:
                            lead_row = session.execute(text("""
                                SELECT client_id FROM "StreemLyne_MT"."Opportunity_Details"
                                WHERE opportunity_id = :id AND tenant_id = :tid
                                LIMIT 1
                            """), {'id': lead_id, 'tid': tenant_id}).fetchone()
                            
                            client_id = lead_row[0] if lead_row else None
                            
                            if client_id:
                                session.execute(text("""
                                    INSERT INTO "StreemLyne_MT"."Client_Interactions"
                                        (client_id, contact_date, contact_method, notes, next_steps, created_at)
                                    VALUES (:cid, CURRENT_DATE, 1, :notes, 'Assignment', :now)
                                """), {
                                    'cid': client_id,
                                    'notes': f"[Assignment] Assigned to {employee_name}: {assignment_notes}",
                                    'now': datetime.utcnow()
                                })
                        except Exception as e:
                            logger.error(f'❌ Error logging assignment history: {e}')
                
                session.commit()
                
                return jsonify({
                    'success': True,
                    'assigned_count': len(lead_ids),
                    'message': f"Assigned {len(lead_ids)} lead(s) successfully",
                    'employee_name': employee_name,  # ✅ ADDED
                    'employee_id': employee_id,      # ✅ ADDED
                }), 200
                
            except Exception as e:
                session.rollback()
                logger.error(f"❌ Error in assign_leads: {e}")
                import traceback
                traceback.print_exc()
                raise
            finally:
                session.close()
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def delete_lead(self, opportunity_id: int) -> tuple:
        """
        DELETE /api/crm/leads/<opportunity_id>
        Delete a lead
        """
        try:
            tenant_id = g.tenant_id
            
            result = self.crm_service.delete_lead(tenant_id, opportunity_id)
            
            if not result.get('success'):
                return jsonify(result), 404
            
            return jsonify(result), 200
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500

    def bulk_delete_leads(self):
        try:
            from flask import request, jsonify, g
            from backend.db import SessionLocal
            from backend.models import Opportunity_Details, Client_Master
            from sqlalchemy import or_, func, String, cast
            import logging

            logger = logging.getLogger(__name__)

            data = request.get_json()
            if not data or 'opportunity_ids' not in data:
                return jsonify({'success': False, 'error': 'opportunity_ids is required'}), 400

            lead_ids = data.get('opportunity_ids', [])
            if not isinstance(lead_ids, list) or len(lead_ids) == 0:
                return jsonify({'success': False, 'error': 'opportunity_ids must be a non-empty list'}), 400

            tenant_id = g.get('tenant_id')
            if not tenant_id:
                return jsonify({'success': False, 'error': 'Tenant ID not found'}), 400

            logger.warning(f'🗑️ bulk_delete_leads: tenant={tenant_id}, ids={lead_ids}')

            session = SessionLocal()
            try:
                # Cast both sides to VARCHAR to avoid type mismatch
                leads = (
                    session.query(Opportunity_Details)
                    .filter(
                        or_(
                            Opportunity_Details.opportunity_id.in_(lead_ids),
                            Opportunity_Details.tenant_lead_id.in_(lead_ids),
                        )
                    )
                    .filter(
                        cast(Opportunity_Details.tenant_id, String) == str(tenant_id)
                    )
                    .all()
                )

                # Fallback: imported leads may only have tenant_id on client or
                # stored as a different type — try int cast as well
                if not leads:
                    logger.warning(f'First query found nothing, trying int cast fallback')
                    try:
                        leads = (
                            session.query(Opportunity_Details)
                            .filter(
                                or_(
                                    Opportunity_Details.opportunity_id.in_(lead_ids),
                                    Opportunity_Details.tenant_lead_id.in_(lead_ids),
                                )
                            )
                            .filter(Opportunity_Details.tenant_id == int(tenant_id))
                            .all()
                        )
                    except (ValueError, TypeError):
                        leads = []

                logger.warning(f'Resolved {len(leads)} leads to delete out of {len(lead_ids)} requested')

                deleted_count = 0
                for lead in leads:
                    if lead.client_id:
                        from backend.models import Project_Details
                        other_projects = (
                            session.query(Project_Details)
                            .filter(Project_Details.client_id == lead.client_id)
                            .count()
                        )
                        if other_projects == 0:
                            client = session.query(Client_Master).filter(
                                Client_Master.client_id == lead.client_id
                            ).first()
                            if client:
                                session.delete(client)

                    session.delete(lead)
                    deleted_count += 1

                session.commit()

                logger.warning(f'✅ Deleted {deleted_count} leads')
                return jsonify({
                    'success': True,
                    'deleted': deleted_count,
                    'total_requested': len(lead_ids),
                    'errors': [],
                    'message': f'{deleted_count} lead(s) deleted successfully.',
                }), 200

            except Exception as e:
                session.rollback()
                logger.exception(f'❌ bulk_delete_leads DB error: {e}')
                return jsonify({'success': False, 'error': str(e)}), 500
            finally:
                session.close()

        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).exception(f'❌ bulk_delete_leads outer error: {e}')
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def import_leads(self) -> tuple:
            """
            POST /api/crm/leads/import
            Bulk import leads from Excel/CSV file
            """
            try:
                tenant_id = g.tenant_id
                
                # Check if file is present
                if 'file' not in request.files:
                    return jsonify({
                        'success': False,
                        'error': 'No file provided',
                        'message': 'Please upload a file'
                    }), 400
                
                file = request.files['file']
                
                if file.filename == '':
                    return jsonify({
                        'success': False,
                        'error': 'No file selected',
                        'message': 'Please select a file to upload'
                    }), 400
                
                # Validate file extension
                allowed_extensions = {'.xlsx', '.xls', '.csv'}
                file_ext = '.' + file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                
                if file_ext not in allowed_extensions:
                    return jsonify({
                        'success': False,
                        'error': 'Invalid file type',
                        'message': 'Only .xlsx, .xls, and .csv files are allowed'
                    }), 400
                
                # Process the file
                result = self.crm_service.import_leads_from_file(tenant_id, file, file_ext)
                
                if not result.get('success'):
                    return jsonify(result), 400
                
                return jsonify(result), 200
                
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': 'Internal server error',
                    'message': str(e)
                }), 500
    
    def download_leads_template(self) -> tuple:
        """
        GET /api/crm/leads/import/template
        Download Excel template for bulk lead import
        """
        try:
            from flask import send_file
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leads Template"
            
            # Define headers
            headers = [
                'Business Name',
                'Contact Person',
                'Tel Number',
                'Email',
                'MPAN/MPR',
                'Start Date',
                'End Date'
            ]
            
            # Write headers with styling
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Add example row
            example_data = [
                'ABC Energy Ltd',
                'John Smith',
                '01234567890',
                'john.smith@abcenergy.com',
                '1234567890123',
                '2025-01-01',
                '2026-01-01'
            ]
            
            for col, value in enumerate(example_data, start=1):
                ws.cell(row=2, column=col, value=value)
            
            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='leads_import_template.xlsx'
            )
            
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500

    def get_priced(self) -> tuple:
        """
        GET /api/crm/priced - Get all priced leads and renewals
        """
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_priced(tenant_id)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    # ========================================
    # PROJECT ENDPOINTS
    # ========================================
    
    def get_projects(self) -> tuple:
        """
        GET /api/crm/projects
        Get all projects for the current tenant
        """
        try:
            tenant_id = g.tenant_id
            
            # Extract query parameters for filtering
            filters = {}
            if request.args.get('status'):
                filters['status'] = request.args.get('status')
            if request.args.get('project_manager_id'):
                filters['project_manager_id'] = int(request.args.get('project_manager_id'))
            
            result = self.crm_service.get_projects(tenant_id, filters if filters else None)
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def get_project_detail(self, project_id: int) -> tuple:
        """
        GET /api/crm/projects/<project_id>
        Get details of a specific project
        """
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_project_detail(tenant_id, project_id)
            
            if not result.get('success'):
                return jsonify(result), 404
            
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    # ========================================
    # DEAL ENDPOINTS
    # ========================================
    
    def get_deals(self) -> tuple:
        """
        GET /api/crm/deals
        Get all deals/contracts for the current tenant
        """
        try:
            tenant_id = g.tenant_id
            
            # Extract query parameters for filtering
            filters = {}
            if request.args.get('status'):
                filters['status'] = request.args.get('status')
            if request.args.get('contract_owner_id'):
                filters['contract_owner_id'] = int(request.args.get('contract_owner_id'))
            
            result = self.crm_service.get_deals(tenant_id, filters if filters else None)
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def get_deal_detail(self, contract_id: int) -> tuple:
        """
        GET /api/crm/deals/<contract_id>
        Get details of a specific deal
        """
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_deal_detail(tenant_id, contract_id)
            
            if not result.get('success'):
                return jsonify(result), 404
            
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    # ========================================
    # USER ENDPOINTS
    # ========================================
    
    def get_users(self) -> tuple:
        """
        GET /api/crm/users
        Get all users for the current tenant
        """
        try:
            tenant_id = g.tenant_id
            active_only = request.args.get('active_only', 'true').lower() == 'true'
            
            result = self.crm_service.get_users(tenant_id, active_only)
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500

    def get_employees(self) -> tuple:
        """
        GET /api/crm/employees
        Get all employees for the current tenant (for assignment dropdowns)
        """
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_employees(tenant_id)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    # ========================================
    # SUPPORTING DATA ENDPOINTS
    # ========================================
    
    def get_roles(self) -> tuple:
        """GET /api/crm/roles - Get all roles"""
        try:
            tenant_id = g.get('tenant_id')
            result = self.crm_service.get_roles(tenant_id)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def get_stages(self) -> tuple:
        """GET /api/crm/stages - Get all pipeline stages"""
        try:
            pipeline_type = request.args.get('pipeline_type')
            result = self.crm_service.get_stages(pipeline_type)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def get_services(self) -> tuple:
        """GET /api/crm/services - Get all services"""
        try:
            tenant_id = g.get('tenant_id')
            result = self.crm_service.get_services(tenant_id)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def get_suppliers(self) -> tuple:
        """GET /api/crm/suppliers - Get all suppliers"""
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_suppliers(tenant_id)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def get_interactions(self) -> tuple:
        """GET /api/crm/interactions - Get all client interactions"""
        try:
            tenant_id = g.tenant_id
            
            # Extract filters
            filters = {}
            if request.args.get('client_id'):
                filters['client_id'] = int(request.args.get('client_id'))
            if request.args.get('interaction_type'):
                filters['interaction_type'] = request.args.get('interaction_type')
            if request.args.get('user_id'):
                filters['user_id'] = int(request.args.get('user_id'))
            
            result = self.crm_service.get_interactions(tenant_id, filters if filters else None)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def get_leads_table(self) -> tuple:
        """
        GET /api/crm/leads/table
        Get leads table for CRM UI (flat rows with 14 columns from joined tables).
        """
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_leads_table(tenant_id)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500

    def preview_lead_import(self) -> tuple:
        """
        POST /api/crm/leads/import/preview
        Controller entrypoint for lead import preview. Expects multipart/form-data
        with a file field named 'file'. Delegates parsing/validation to the
        CRMService and returns the preview response (no DB writes).
        """
        try:
            tenant_id = g.tenant_id

            if 'file' not in request.files:
                return jsonify({'success': False, 'error': 'No file provided', 'message': "Include a file under the 'file' form field."}), 400

            file_storage = request.files.get('file')
            result = self.crm_service.preview_lead_import(tenant_id, file_storage)

            # Service returns { success: bool, ... }
            status = 200 if result.get('success') else 400
            return jsonify(result), status
        except Exception as e:
            logger.exception("preview_lead_import controller error: %s", e)
            return jsonify({'success': False, 'error': 'Internal server error', 'message': str(e)}), 500

    def import_leads_confirm(self) -> tuple:
        """
        POST /api/crm/leads/import/confirm
        Accepts JSON payload: an array of rows (validated by preview). Delegates
        insertion to CRMService.confirm_lead_import which performs tenant-scoped
        inserts and returns a summary (partial success allowed).
        """
        try:
            tenant_id = g.tenant_id
            payload = request.get_json()

            if not payload or not isinstance(payload, list):
                return jsonify({'success': False, 'error': 'Invalid request', 'message': 'Expected a JSON array of rows.'}), 400

            created_by = getattr(request.current_user, 'id', None)

            service_param = request.args.get('service')
            service_value = service_param.strip().lower() if isinstance(service_param, str) else None
            if service_value == 'water':
                service_id = 2
            else:
                service_id = 1

            result = self.crm_service.confirm_lead_import(tenant_id, payload, created_by, service_id)

            status = 200 if result.get('success') else 400
            return jsonify(result), status
        except Exception as e:
            logger.exception('import_leads_confirm controller error: %s', e)
            return jsonify({'success': False, 'error': 'Internal server error', 'message': str(e)}), 500
        
    def import_leads_preview(self) -> tuple:
        """
        POST /api/crm/leads/import/preview
        Accepts multipart/form-data with a file (CSV or XLSX) and returns a
        preview of validation results. DOES NOT write to the database.
        """
        try:
            tenant_id = g.tenant_id

            if 'file' not in request.files:
                return jsonify({'success': False, 'error': 'No file provided'}), 400

            file = request.files.get('file')
            result = self.crm_service.preview_lead_import(tenant_id, file)

            # Service returns structured response
            status = 200 if result.get('success') else 400
            return jsonify(result), status
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    def get_leads_by_customer_type(self) -> tuple:
        """
        GET /api/crm/leads/customer-type?type=NEW|EXISTING
        Get leads filtered by customer type
        """
        try:
            tenant_id = g.tenant_id
            customer_type_param = request.args.get('type', None)
            
            # Extract query parameters for filtering
            filters = {}
            if request.args.get('stage_id'):
                filters['stage_id'] = int(request.args.get('stage_id'))
            if request.args.get('lead_status'):
                filters['lead_status'] = request.args.get('lead_status')
            if request.args.get('assigned_employee_id'):
                filters['assigned_employee_id'] = int(request.args.get('assigned_employee_id'))
            
            result = self.crm_service.get_leads_by_customer_type(
                tenant_id, 
                customer_type_param, 
                filters if filters else None
            )
            return jsonify(result), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    def create_client(self) -> tuple:
        """
        POST /api/crm/clients
        Create a new client in Client_Master and automatically create one
        Opportunity_Details record (lead) for that client.
        """
        try:
            tenant_id = g.tenant_id
            data = request.get_json()
            if not data:
                return jsonify({
                    'success': False,
                    'error': 'Invalid request',
                    'message': 'Request body is required'
                }), 400
            company = data.get('client_company_name') or data.get('business_name')
            if not company:
                return jsonify({
                    'success': False,
                    'error': 'Validation error',
                    'message': 'client_company_name or business_name is required'
                }), 400
            result = self.crm_service.create_client(tenant_id, data)
            if not result.get('success'):
                return jsonify(result), 400
            return jsonify(result), 201
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500

    def create_call_summary(self, client_id: int) -> tuple:
        """
        POST /api/crm/clients/<client_id>/call-summary
        Create a call summary/interaction record
        """
        try:
            tenant_id = g.tenant_id
            call_data = request.get_json()
            
            if not call_data:
                return jsonify({
                    'success': False,
                    'error': 'Invalid request',
                    'message': 'Request body is required'
                }), 400
            
            result = self.crm_service.create_call_summary(tenant_id, client_id, call_data)
            
            if not result.get('success'):
                return jsonify(result), 400
            
            return jsonify(result), 201
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Internal server error',
                'message': str(e)
            }), 500
    
    # ========================================
    # DASHBOARD
    # ========================================
    
    def get_dashboard(self) -> tuple:
        """
        GET /api/crm/dashboard
        Get CRM dashboard summary
        """
        try:
            tenant_id = g.tenant_id
            result = self.crm_service.get_dashboard_summary(tenant_id)
            return jsonify(result), 200
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

def upload_document(self) -> tuple:
    """
    POST /api/crm/documents/upload
    Upload a new document (admin)
    """
    try:
        tenant_id = g.tenant_id
        
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided'
            }), 400
        
        file = request.files['file']
        
        document_data = {
            'document_name': request.form.get('document_name', ''),
            'document_description': request.form.get('document_description', ''),
            'category': request.form.get('category', 'OTHER'),
            'is_template': request.form.get('is_template', 'true').lower() == 'true'
        }
        
        result = self.document_service.upload_document(
            tenant_id, 
            file, 
            document_data,
            uploaded_by_client=False
        )
        
        if not result.get('success'):
            return jsonify(result), 400
        
        return jsonify(result), 201
    except Exception as e:
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'message': str(e)
        }), 500

def client_upload_document(self, client_id: int) -> tuple:
    """
    POST /api/crm/clients/<client_id>/upload
    Upload a document for a specific client (client upload)
    """
    try:
        tenant_id = g.tenant_id
        
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided'
            }), 400
        
        file = request.files['file']
        
        document_data = {
            'document_name': request.form.get('document_name', file.filename),
            'document_description': request.form.get('document_description', ''),
            'category': request.form.get('category', 'CLIENT_UPLOAD'),
            'is_template': False
        }
        
        result = self.document_service.upload_document(
            tenant_id, 
            file, 
            document_data,
            uploaded_by_client=True,
            client_id=client_id
        )
        
        if not result.get('success'):
            return jsonify(result), 400
        
        return jsonify(result), 201
    except Exception as e:
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'message': str(e)
        }), 500