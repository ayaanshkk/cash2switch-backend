"""
Bulk Import Route for Leads (CRM)
Mirrors import_routes.py energy-customers logic but maps to CRM leads tables.
Add this to import_routes.py (or register as a separate blueprint).
"""

from flask import request, jsonify, current_app, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import os
from datetime import datetime
from sqlalchemy import text
import tempfile
import logging

from ..models import (
    Client_Master, Project_Details, Energy_Contract_Master,
    Opportunity_Details, Supplier_Master, Employee_Master, Stage_Master
)
from .auth_helpers import token_required
from ..db import SessionLocal

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
# Column mapping: canonical field → possible Excel header names
# ─────────────────────────────────────────────────────────────
LEADS_COLUMN_MAP = {
    'client_name':    ['client name', 'business name', 'company name', 'name'],
    'trading_name':   ['trading name', 'trading', 'business', 'company'],
    'main_contact':   ['main contact', 'contact person', 'contact', 'full name'],
    'position':       ['position', 'role', 'title', 'job title'],
    'tel_no':         ['tel no', 'phone', 'telephone', 'tel', 'tel number', 'phone number'],
    'mobile_no':      ['mobile no', 'mobile', 'cell', 'mobile number'],
    'email':          ['email', 'e-mail', 'email address'],
    'address_line_1': ['address line 1', 'address 1', 'street', 'address'],
    'address_line_2': ['address line 2', 'address 2'],
    'town':           ['town', 'city'],
    'county':         ['county', 'region'],
    'postcode':       ['postcode', 'post code', 'zip'],
    'mpan_top':       ['mpan top', 'mpan', 'mpan core', 'mpan/mpr', 'mpr'],
    'mpan_bottom':    ['mpan bottom', 'mpan llf'],
    'supplier':       ['supplier', 'supplier name', 'current supplier'],
    'old_supplier':   ['old supplier', 'previous supplier'],
    'start_date':     ['start date', 'contract start'],
    'contract_end':   ['contract end', 'end date', 'expiry', 'contract expiry'],
    'annual_usage':   ['annual usage', 'usage', 'kwh', 'annual kwh'],
    'unit_rate':      ['unit rate', 'rate', 'rate 1', 'rate1'],
    'stand_charge':   ['stand charge', 'standing charge'],
    'payment_type':   ['payment type'],
    'site_name':      ['site name', 'site'],
    'data_source':    ['data source', 'source'],
}


def _resolve_columns(df_columns):
    """Map normalised df column names → canonical field names."""
    normalised = [c.strip().lower() for c in df_columns]
    actual = {}
    for field, aliases in LEADS_COLUMN_MAP.items():
        for col in normalised:
            if col in aliases:
                # store the original (un-normalised) column name
                original_idx = normalised.index(col)
                actual[field] = df_columns[original_idx]
                break
    return actual


def _safe_str(value):
    if pd.isna(value) or value is None or str(value).strip() in ('', 'nan'):
        return ''
    s = str(value).strip()
    if s.endswith('.0') and s[:-2].replace('.', '', 1).isdigit():
        s = s[:-2]
    return s


def _parse_date(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s or s.lower() == 'nan':
        return None
    fmts = [
        '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%d-%m-%Y',
        '%d.%m.%Y', '%d %b %Y', '%d %B %Y', '%Y-%m-%d', '%m/%d/%Y',
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(val):
    if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() in ('', 'nan'):
        return None
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, AttributeError):
        return None


def _get_or_create_supplier(supplier_name, suppliers_dict, session):
    if not supplier_name:
        return 1
    key = supplier_name.lower().strip()
    if key in suppliers_dict:
        return suppliers_dict[key]
    try:
        new_s = Supplier_Master(
            supplier_company_name=supplier_name,
            supplier_contact_name='Auto-imported',
            supplier_provisions=3,
            created_at=datetime.utcnow()
        )
        session.add(new_s)
        session.flush()
        suppliers_dict[key] = new_s.supplier_id
        return new_s.supplier_id
    except Exception:
        return 1


def _get_tenant_id(user):
    if hasattr(user, 'tenant_id') and user.tenant_id is not None:
        return user.tenant_id
    session = SessionLocal()
    try:
        emp = session.query(Employee_Master).filter_by(employee_id=user.employee_id).first()
        return emp.tenant_id if emp else None
    finally:
        session.close()


# ─────────────────────────────────────────────────────────────
# POST /import/leads  –  main import endpoint
# ─────────────────────────────────────────────────────────────
# Register this on import_bp  e.g.  @import_bp.route('/leads', methods=['POST', 'OPTIONS'])

def import_leads_handler():
    """
    Bulk-import leads from Excel/CSV.

    Mirrors /import/energy-customers exactly:
    - Reads file, normalises columns
    - Creates Client_Master + Opportunity_Details + Project_Details + Energy_Contract_Master
    - Deduplicates on MPAN (cross-tenant aware)
    - Auto-archives older contracts when a newer one is detected
    - Returns the same JSON shape as energy-customers import
    """
    print("\n\n🔥🔥🔥 LEADS IMPORT CALLED 🔥🔥🔥\n\n")

    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session = SessionLocal()
    sql_logger = logging.getLogger('sqlalchemy.engine')
    original_level = sql_logger.level
    sql_logger.setLevel(logging.WARNING)

    try:
        # ── Auth / tenant ──────────────────────────────────────────────
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        filename = secure_filename(file.filename)
        if not ('.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls', 'csv'}):
            return jsonify({'error': 'Invalid file type. Please upload .xlsx, .xls, or .csv'}), 400

        tenant_id = _get_tenant_id(request.current_user)
        if not tenant_id:
            return jsonify({'error': 'Tenant not found for user'}), 400

        employee_id = request.current_user.employee_id

        assigned_employee_id = request.form.get('assigned_employee_id', type=int)
        opportunity_owner_id = assigned_employee_id if assigned_employee_id else employee_id

        assigned_employee_name = None
        if assigned_employee_id:
            ae = session.query(Employee_Master).filter_by(
                employee_id=assigned_employee_id, tenant_id=tenant_id
            ).first()
            if ae:
                assigned_employee_name = ae.employee_name
            else:
                return jsonify({'error': f'Invalid employee ID: {assigned_employee_id}'}), 400

        # ── Service / stage ────────────────────────────────────────────
        service_param = request.args.get('service', 'electricity')
        service_id_map = {'electricity': 1, 'utilities': 1, 'water': 2, 'gas': 3}
        import_service_id = service_id_map.get(service_param.strip().lower(), 1)

        # Default stage (first stage in Stage_Master)
        default_stage = session.query(Stage_Master).order_by(Stage_Master.stage_id).first()
        default_stage_id = default_stage.stage_id if default_stage else 1

        # ── Read file ──────────────────────────────────────────────────
        file_ext = filename.rsplit('.', 1)[1].lower()
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}') as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name
            if file_ext == 'csv':
                df = pd.read_csv(tmp_path, encoding='utf-8-sig', dtype=str)
            else:
                try:
                    df = pd.read_excel(tmp_path, engine='openpyxl', dtype=str)
                except Exception:
                    df = pd.read_excel(tmp_path, engine='xlrd', dtype=str)
            os.unlink(tmp_path)
        except Exception as e:
            return jsonify({'error': f'Failed to read file: {str(e)}'}), 400

        # Normalise column headers for matching, keep originals for data access
        original_cols = list(df.columns)
        df.columns = df.columns.str.strip().str.lower().str.replace('_', ' ').str.replace(r'\s+', ' ', regex=True)
        actual_columns = _resolve_columns(list(df.columns))
        # Restore original column names in df so _safe_str works
        df.columns = original_cols
        # Re-resolve against originals (headers are now original-cased)
        actual_columns_orig = {}
        norm_to_orig = {c.strip().lower().replace('_', ' '): c for c in original_cols}
        for field, aliases in LEADS_COLUMN_MAP.items():
            for alias in aliases:
                norm_alias = alias.strip().lower()
                if norm_alias in norm_to_orig:
                    actual_columns_orig[field] = norm_to_orig[norm_alias]
                    break

        # ── Pre-load suppliers ─────────────────────────────────────────
        suppliers_dict = {
            s.supplier_company_name.lower().strip(): s.supplier_id
            for s in session.query(Supplier_Master).all()
        }

        # ── Pre-load existing MPANs (cross-tenant) ─────────────────────
        existing_mpans = {}
        for contract, ct_tenant_id, emp_id, emp_name, co_name, is_arch in session.query(
            Energy_Contract_Master,
            Client_Master.tenant_id,
            Client_Master.assigned_employee_id,
            Employee_Master.employee_name,
            Client_Master.client_company_name,
            Client_Master.is_archived,
        ).join(Project_Details, Energy_Contract_Master.project_id == Project_Details.project_id
        ).join(Client_Master, Project_Details.client_id == Client_Master.client_id
        ).outerjoin(Employee_Master, Client_Master.assigned_employee_id == Employee_Master.employee_id
        ).all():
            if contract.mpan_number:
                key = contract.mpan_number.strip().lower()
                existing_mpans.setdefault(key, []).append({
                    'contract': contract,
                    'tenant_id': ct_tenant_id,
                    'assigned_to_id': emp_id,
                    'assigned_to_name': emp_name or 'Unassigned',
                    'company_name': co_name,
                    'is_archived': is_arch,
                })

        # ── Process rows ───────────────────────────────────────────────
        total_rows = len(df)
        success_count = error_count = duplicate_count = 0
        errors = []
        duplicate_details = []
        cross_tenant_duplicates = []
        BATCH_SIZE = 50

        def gcol(field):
            """Get column name for a canonical field."""
            return actual_columns_orig.get(field, '')

        for index, row in df.iterrows():
            try:
                client_name   = _safe_str(row.get(gcol('client_name'), ''))
                trading_name  = _safe_str(row.get(gcol('trading_name'), ''))
                main_contact  = _safe_str(row.get(gcol('main_contact'), ''))
                position      = _safe_str(row.get(gcol('position'), ''))
                tel_no        = _safe_str(row.get(gcol('tel_no'), ''))
                mobile_no     = _safe_str(row.get(gcol('mobile_no'), ''))
                email         = _safe_str(row.get(gcol('email'), ''))
                site_name     = _safe_str(row.get(gcol('site_name'), ''))
                addr1         = _safe_str(row.get(gcol('address_line_1'), ''))
                addr2         = _safe_str(row.get(gcol('address_line_2'), ''))
                town          = _safe_str(row.get(gcol('town'), ''))
                county        = _safe_str(row.get(gcol('county'), ''))
                postcode      = _safe_str(row.get(gcol('postcode'), ''))
                mpan_top      = _safe_str(row.get(gcol('mpan_top'), ''))
                mpan_bottom   = _safe_str(row.get(gcol('mpan_bottom'), ''))
                supplier_name = _safe_str(row.get(gcol('supplier'), ''))
                old_sup_name  = _safe_str(row.get(gcol('old_supplier'), ''))
                payment_type  = _safe_str(row.get(gcol('payment_type'), ''))
                data_source   = _safe_str(row.get(gcol('data_source'), ''))
                start_date    = _parse_date(row.get(gcol('start_date'), ''))
                end_date      = _parse_date(row.get(gcol('contract_end'), ''))
                annual_usage  = _parse_number(row.get(gcol('annual_usage'), ''))
                unit_rate     = _parse_number(row.get(gcol('unit_rate'), ''))
                stand_charge  = _parse_number(row.get(gcol('stand_charge'), ''))

                address_parts = [p for p in [addr1, addr2, town, county] if p]
                address       = ', '.join(address_parts)
                site_address  = site_name or address
                business_name = trading_name or client_name
                contact_person = main_contact or client_name

                # Skip blank rows
                if not business_name and not tel_no and not email and not mpan_top and not contact_person:
                    continue

                supplier_id     = _get_or_create_supplier(supplier_name, suppliers_dict, session) if supplier_name else 1
                old_supplier_id = _get_or_create_supplier(old_sup_name, suppliers_dict, session) if old_sup_name else None

                # ── MPAN duplicate logic (same as energy-customers) ────
                if mpan_top:
                    mpan_key = mpan_top.strip().lower()
                    existing_records = existing_mpans.get(mpan_key)

                    if existing_records:
                        duplicate_count += 1
                        cross_rec = next((r for r in existing_records if r['tenant_id'] != tenant_id), None)
                        same_rec  = next((r for r in existing_records if r['tenant_id'] == tenant_id), None)

                        if cross_rec:
                            cross_tenant_duplicates.append({
                                'row': index + 2,
                                'mpan': mpan_top,
                                'new_company': business_name,
                                'existing_company': cross_rec['company_name'],
                                'existing_tenant_id': cross_rec['tenant_id'],
                                'assigned_to': cross_rec['assigned_to_name'],
                                'is_archived': cross_rec['is_archived'],
                            })
                            continue

                        if same_rec:
                            existing_contract = same_rec['contract']
                            existing_end = existing_contract.contract_end_date
                            new_end = end_date

                            if existing_end and new_end and existing_end == new_end:
                                action = 'Exact duplicate - skipped'
                            elif existing_end and new_end and new_end < existing_end:
                                action = 'Older record - created as archived'
                            elif existing_end and new_end and new_end > existing_end:
                                action = 'Newer record - archived existing'
                            else:
                                action = 'Updated existing record'

                            duplicate_details.append({
                                'row': index + 2, 'mpan': mpan_top,
                                'company': business_name,
                                'assigned_to': same_rec['assigned_to_name'],
                                'action': action,
                            })

                            if same_rec['is_archived']:
                                continue
                            if not new_end:
                                continue

                            if existing_end and new_end < existing_end:
                                # Create as archived
                                try:
                                    arc_client = Client_Master(
                                        tenant_id=tenant_id,
                                        assigned_employee_id=opportunity_owner_id,
                                        client_company_name=business_name or '',
                                        client_contact_name=contact_person or '',
                                        address=address or '',
                                        post_code=postcode or '',
                                        client_phone=tel_no or '',
                                        client_mobile=mobile_no or None,
                                        client_email=email or '',
                                        client_website='',
                                        default_currency_id=1,
                                        created_at=datetime.utcnow(),
                                        is_archived=True,
                                        archived_at=datetime.utcnow(),
                                        archived_reason=f"Historical record (ended {new_end}) - superseded by existing contract ending {existing_end}",
                                    )
                                    session.add(arc_client)
                                    session.flush()

                                    arc_opp = Opportunity_Details(
                                        client_id=arc_client.client_id,
                                        opportunity_title=business_name or '',
                                        opportunity_description='Imported lead (archived)',
                                        opportunity_date=datetime.utcnow().date(),
                                        opportunity_owner_employee_id=opportunity_owner_id,
                                        stage_id=default_stage_id,
                                        opportunity_value=0,
                                        currency_id=1,
                                        created_at=datetime.utcnow(),
                                    )
                                    session.add(arc_opp)
                                    session.flush()

                                    arc_proj = Project_Details(
                                        client_id=arc_client.client_id,
                                        opportunity_id=arc_opp.opportunity_id,
                                        project_title=business_name or '',
                                        project_description='Imported lead site',
                                        start_date=start_date or datetime.utcnow().date(),
                                        end_date=end_date,
                                        employee_id=employee_id,
                                        created_at=datetime.utcnow(),
                                        updated_at=datetime.utcnow(),
                                        address=site_address or address or '',
                                        Misc_Col2=int(annual_usage) if annual_usage else None,
                                        site_name=site_name or None,
                                        town=town or None,
                                        county=county or None,
                                    )
                                    session.add(arc_proj)
                                    session.flush()

                                    arc_contract = Energy_Contract_Master(
                                        project_id=arc_proj.project_id,
                                        employee_id=employee_id,
                                        supplier_id=supplier_id,
                                        old_supplier_id=old_supplier_id,
                                        contract_start_date=start_date or datetime.utcnow().date(),
                                        contract_end_date=end_date,
                                        terms_of_sale='',
                                        service_id=import_service_id,
                                        unit_rate=unit_rate or 0.0,
                                        currency_id=1,
                                        created_at=datetime.utcnow(),
                                        updated_at=datetime.utcnow(),
                                        mpan_number=mpan_top or '',
                                        mpan_bottom=mpan_bottom or '',
                                        standing_charge=stand_charge,
                                        rate_1=unit_rate,
                                        payment_type=payment_type or None,
                                    )
                                    session.add(arc_contract)
                                    session.flush()

                                    existing_mpans.setdefault(mpan_key, []).append({
                                        'contract': arc_contract,
                                        'tenant_id': tenant_id,
                                        'assigned_to_id': opportunity_owner_id,
                                        'assigned_to_name': assigned_employee_name or 'Unassigned',
                                        'company_name': business_name,
                                        'is_archived': True,
                                    })
                                    session.commit()
                                    success_count += 1
                                    continue
                                except Exception as ae:
                                    session.rollback()
                                    error_count += 1
                                    errors.append(f"Row {index + 2}: Archive creation failed - {str(ae)}")
                                    continue

                            if existing_end and new_end == existing_end:
                                continue

                            if existing_end and new_end > existing_end:
                                proj = session.query(Project_Details).filter_by(
                                    project_id=existing_contract.project_id
                                ).first()
                                if proj:
                                    client_to_arch = session.query(Client_Master).filter_by(
                                        client_id=proj.client_id
                                    ).first()
                                    if client_to_arch:
                                        client_to_arch.is_archived = True
                                        client_to_arch.archived_at = datetime.utcnow()
                                        client_to_arch.archived_reason = f"Superseded by newer contract (ending {new_end})"
                                        session.flush()
                                        for r in existing_mpans[mpan_key]:
                                            if r['contract'].energy_contract_master_id == existing_contract.energy_contract_master_id:
                                                r['is_archived'] = True
                                                break

                # ── Create new lead ────────────────────────────────────
                try:
                    new_client = Client_Master(
                        tenant_id=tenant_id,
                        assigned_employee_id=opportunity_owner_id,
                        client_company_name=business_name or '',
                        client_contact_name=contact_person or '',
                        address=address or '',
                        post_code=postcode or '',
                        client_phone=tel_no or '',
                        client_mobile=mobile_no or None,
                        client_email=email or '',
                        client_website='',
                        default_currency_id=1,
                        created_at=datetime.utcnow(),
                        position=position or None,
                        is_archived=False,
                    )
                    session.add(new_client)
                    session.flush()

                    new_opp = Opportunity_Details(
                        client_id=new_client.client_id,
                        opportunity_title=business_name or '',
                        opportunity_description='Imported lead',
                        opportunity_date=datetime.utcnow().date(),
                        opportunity_owner_employee_id=opportunity_owner_id,
                        stage_id=default_stage_id,
                        opportunity_value=0,
                        currency_id=1,
                        created_at=datetime.utcnow(),
                        # Mark as [IMPORTED LEADS] placeholder so energy routes ignore it
                        Misc_Col1=None,
                    )
                    session.add(new_opp)
                    session.flush()

                    new_proj = Project_Details(
                        client_id=new_client.client_id,
                        opportunity_id=new_opp.opportunity_id,
                        project_title=business_name or '',
                        project_description='Imported lead site',
                        start_date=start_date or datetime.utcnow().date(),
                        end_date=end_date,
                        employee_id=employee_id,
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow(),
                        address=site_address or address or '',
                        Misc_Col2=int(annual_usage) if annual_usage else None,
                        site_name=site_name or None,
                        town=town or None,
                        county=county or None,
                    )
                    session.add(new_proj)
                    session.flush()

                    new_contract = Energy_Contract_Master(
                        project_id=new_proj.project_id,
                        employee_id=employee_id,
                        supplier_id=supplier_id,
                        old_supplier_id=old_supplier_id,
                        contract_start_date=start_date or datetime.utcnow().date(),
                        contract_end_date=end_date,
                        terms_of_sale='',
                        service_id=import_service_id,
                        unit_rate=unit_rate or 0.0,
                        currency_id=1,
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow(),
                        mpan_number=mpan_top or '',
                        mpan_bottom=mpan_bottom or '',
                        standing_charge=stand_charge,
                        rate_1=unit_rate,
                        payment_type=payment_type or None,
                    )
                    session.add(new_contract)
                    session.flush()

                    if mpan_top:
                        mpan_key = mpan_top.strip().lower()
                        existing_mpans.setdefault(mpan_key, []).append({
                            'contract': new_contract,
                            'tenant_id': tenant_id,
                            'assigned_to_id': opportunity_owner_id,
                            'assigned_to_name': assigned_employee_name or 'Unassigned',
                            'company_name': business_name,
                            'is_archived': False,
                        })

                    success_count += 1

                    if (success_count + duplicate_count) % BATCH_SIZE == 0:
                        session.commit()

                except Exception as row_err:
                    session.rollback()
                    error_count += 1
                    err_str = str(row_err)
                    if 'duplicate key' in err_str or 'UniqueViolation' in err_str:
                        errors.append(f"Duplicate record (row {index + 2})")
                    elif 'IntegrityError' in err_str or 'not-null' in err_str:
                        errors.append(f"Missing required field (row {index + 2})")
                    else:
                        errors.append(err_str.split('\n')[0][:150])
                    continue

            except Exception as outer_err:
                session.rollback()
                error_count += 1
                errors.append(str(outer_err).split('\n')[0][:150])
                continue

        # ── Final commit ───────────────────────────────────────────────
        try:
            session.commit()
        except Exception:
            session.rollback()

        # ── Duplicate report ───────────────────────────────────────────
        duplicate_report = []
        if duplicate_details:
            duplicate_report.append("📋 SAME-TENANT DUPLICATES:")
            for d in duplicate_details:
                duplicate_report.append(
                    f"  Row {d['row']}: {d['company']} (MPAN: {d['mpan']}) - "
                    f"Assigned to: {d['assigned_to']} - {d['action']}"
                )
        if cross_tenant_duplicates:
            duplicate_report.append("⚠️ CROSS-TENANT DUPLICATES (SKIPPED):")
            for d in cross_tenant_duplicates:
                duplicate_report.append(
                    f"  Row {d['row']}: {d['new_company']} (MPAN: {d['mpan']}) - "
                    f"Exists in another account"
                )

        return jsonify({
            'success': True,
            'message': 'Import completed',
            'total_rows': total_rows,
            'successful': success_count,
            'duplicates': duplicate_count,
            'same_tenant_duplicates': len(duplicate_details),
            'cross_tenant_duplicates': len(cross_tenant_duplicates),
            'failed': error_count,
            'errors': errors[:50],
            'duplicate_report': duplicate_report,
            'assigned_to': assigned_employee_name,
            'assigned_employee_id': assigned_employee_id,
        }), 200

    except Exception as e:
        session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Import failed: {str(e)}'}), 500
    finally:
        sql_logger.setLevel(original_level)
        session.close()


# ─────────────────────────────────────────────────────────────
# GET /import/leads/template  –  download Excel template
# ─────────────────────────────────────────────────────────────

def download_leads_template_handler():
    """Download Excel template matching the leads import format."""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Import Template"

        headers = [
            "Client Name", "Trading Name", "Main Contact", "Position",
            "Tel No", "Mobile No", "Email",
            "Address Line 1", "Address Line 2", "Town", "County", "Postcode",
            "MPAN Top", "MPAN Bottom",
            "Supplier", "Old Supplier",
            "Start Date", "Contract End",
            "Annual Usage", "Unit Rate", "Stand Charge",
            "Payment Type", "Site Name", "Data Source",
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        example = [
            "ABC Limited", "ABC Trading", "John Smith", "Director",
            "07700900000", "07700900001", "john@abc.com",
            "123 Main St", "Unit 5", "London", "Greater London", "SW1A 1AA",
            "1100012314490", "04031N12",
            "British Gas", "EDF Energy",
            "01/01/2024", "31/12/2024",
            "25000", "35.00", "45.13",
            "DD", "Main Site", "Renewals",
        ]
        for col, v in enumerate(example, 1):
            ws.cell(row=2, column=col, value=v)

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='leads_import_template.xlsx',
        )
    except Exception as e:
        current_app.logger.exception(f"❌ Leads template download failed: {e}")
        return jsonify({'error': 'Failed to generate template'}), 500