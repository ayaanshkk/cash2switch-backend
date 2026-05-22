# -*- coding: utf-8 -*-
"""
CRM Routes Blueprint
Defines API endpoints for CRM module
"""
from asyncio.log import logger

from backend import db
from flask import Blueprint, request, g, jsonify, current_app
from functools import wraps
from datetime import datetime, timedelta
import re
from sqlalchemy import text, bindparam, func, String
from backend.db import SessionLocal
from backend.models import (
    Opportunity_Details, Client_Master, Stage_Master, 
    Employee_Master, Supplier_Master, Client_Interactions
)
from backend.crm.controllers.crm_controller import CRMController
from backend.crm.middleware.tenant_middleware import require_tenant
from backend.crm.utils.role_helpers import is_crm_leads_admin_role
from .auth_helpers import token_required
from backend.crm.utils.display_order_helpers import recalculate_display_order
from ..dummy_local_dashboard_data import (
    dummy_leads_by_stage,
    dummy_leads_list,
    dummy_leads_period_breakdown,
    dummy_leads_salesperson_breakdown,
    dummy_leads_stage_breakdown,
    dummy_leads_stats,
    dummy_leads_staff_performance,
    dummy_leads_supplier_breakdown,
    local_demo_dashboard_enabled,
)

def _iso(v):
    """Convert datetime/date objects to ISO format strings"""
    if v is None:
        return None
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    return v

def _serial(v):
    """Serialize values for JSON response"""
    if v is None:
        return None
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    try:
        from decimal import Decimal
        if isinstance(v, Decimal):
            return float(v)
    except ImportError:
        pass
    return v

def _safe_float(v):
    try:
        if v is None:
            return 0.0
        return float(v)
    except Exception:
        try:
            txt = str(v).replace(",", "").strip()
            return float(txt) if txt else 0.0
        except Exception:
            return 0.0

def _lead_stage_bucket(stage_name: str) -> str:
    s = (stage_name or "").strip().lower()
    if s in ("converted", "already renewed", "renewed", "renewed directly"):
        return "converted"
    if s in ("lost", "lost cot", "invalid number", "meter de-energised"):
        return "lost"
    if s in ("callback", "not answered", "broker in place", "email only", "complaint", "incorrect supplier", "priced", "end date changed"):
        return "in_progress"
    return "not_contacted"

# Lightweight helper: attach tenant_id from decoded JWT to `g` (no new auth logic)
def tenant_from_jwt(f):
    """Set g.tenant_id from request.current_user.tenant_id (returns 401 if missing).

    This is a thin wiring decorator that relies on the existing `token_required`
    to populate `request.current_user`. It does NOT perform authentication itself.
    """
    @wraps(f)
    def _wrap(*args, **kwargs):
        current_user = getattr(request, 'current_user', None)
        if not current_user or getattr(current_user, 'tenant_id', None) is None:
            return jsonify({
                'error': 'Missing tenant in token',
                'message': 'Authenticated token must include tenant_id'
            }), 401
        # Propagate tenant_id to Flask `g` for downstream code that expects it
        g.tenant_id = getattr(current_user, 'tenant_id')
        return f(*args, **kwargs)
    return _wrap

# Create blueprint
crm_bp = Blueprint('crm', __name__, url_prefix='/api/crm')

# Initialize controller
crm_controller = CRMController()

OFFSHORE_ROLE_ID = 5


def _staff_period_bounds(period: str):
    now = datetime.now()
    key = (period or "daily").strip().lower()
    if key == "weekly":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=7)
        multiplier = 7
    elif key == "monthly":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1)
        else:
            end = start.replace(month=start.month + 1)
        multiplier = 30
    else:
        key = "daily"
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
        multiplier = 1
    return key, start, end, multiplier


def _resolve_opportunity_ts_expr(db) -> str:
    rows = db.execute_query(
        '''
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'StreemLyne_MT'
              AND table_name = 'Opportunity_Details'
        '''
    )
    columns = {str((r or {}).get('column_name') or '').strip().lower() for r in (rows or [])}
    if 'updated_at' in columns:
        return 'od."updated_at"'
    if 'modified_at' in columns:
        return 'od."modified_at"'
    if 'created_at' in columns:
        return 'od."created_at"'
    return 'NOW()'


def _safe_float(v):
    try:
        if v is None:
            return 0.0
        return float(v)
    except Exception:
        try:
            txt = str(v).replace(",", "").strip()
            return float(txt) if txt else 0.0
        except Exception:
            return 0.0


def _lead_stage_bucket(stage_name: str) -> str:
    s = (stage_name or "").strip().lower()
    if s in ("converted", "already renewed", "renewed", "renewed directly"):
        return "converted"
    if s in ("lost", "lost cot", "invalid number", "meter de-energised"):
        return "lost"
    if s in ("callback", "not answered", "broker in place", "email only", "complaint", "incorrect supplier", "priced", "end date changed"):
        return "in_progress"
    return "not_contacted"

# ========================================
# LEAD ROUTES
# ========================================

@crm_bp.route('/leads', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads():
    try:
        tenant_id = g.tenant_id
        current_user = request.current_user
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1
        exclude_stage = request.args.get('exclude_stage', '')

        user_id = getattr(current_user, 'id', None) or getattr(current_user, 'user_id', None)
        employee_id = getattr(current_user, 'employee_id', None)
        role_name = getattr(current_user, 'role', None)
        normalized_role = str(role_name).strip().lower() if role_name else None
        admin_user = is_crm_leads_admin_role(role_name)

        current_app.logger.warning(
            'crm.get_leads start tenant=%s user_id=%s employee_id=%s is_admin=%s role=%s service=%s exclude_stage=%s',
            tenant_id, user_id, employee_id, admin_user, normalized_role, service_param, exclude_stage
        )

        if local_demo_dashboard_enabled():
            scoped_employee_id = request.args.get('employee_id', type=int) if admin_user else employee_id
            rows = dummy_leads_list(scoped_employee_id)
            if exclude_stage:
                rows = [r for r in rows if (r.get('stage_name') or '').strip().lower() != exclude_stage.strip().lower()]
            return jsonify(rows), 200

        if not admin_user and not employee_id:
            current_app.logger.warning(
                'crm.get_leads empty_result reason=no_employee_id tenant=%s user_id=%s',
                tenant_id, user_id
            )
            return jsonify([]), 200

        session = SessionLocal()
        try:
            query = (
                session.query(
                    Opportunity_Details,
                    Stage_Master.stage_name,
                    Employee_Master.employee_name.label('assigned_to_name'),
                    func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name'),
                    Supplier_Master.supplier_company_name.label('supplier_name')
                )
                .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
                .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
                .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
                .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
                .filter(
                    (Opportunity_Details.tenant_id == tenant_id) |
                    ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
                )
                .filter(Opportunity_Details.service_id == service_id)
                .filter(Opportunity_Details.opportunity_owner_employee_id.isnot(None))
                .filter((Opportunity_Details.is_draft == False) | (Opportunity_Details.is_draft.is_(None)))
                .filter(
                    (Client_Master.is_deleted.is_(None)) |
                    (Client_Master.is_deleted == False)
                )
            )

            if not admin_user:
                query = query.filter(
                    Opportunity_Details.opportunity_owner_employee_id == employee_id,
                    (Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None))
                )

            if exclude_stage:
                query = query.filter(
                    (Stage_Master.stage_name.is_(None)) |
                    (func.lower(Stage_Master.stage_name) != exclude_stage.lower())
                )

            query = query.order_by(Opportunity_Details.created_at.desc())
            rows = query.all()

            results = []
            for row in rows:
                od = row[0]
                results.append({
                    'opportunity_id': od.opportunity_id,
                    'tenant_lead_id': od.tenant_lead_id,
                    'business_name': row.business_name,
                    'contact_person': od.contact_person,
                    'tel_number': str(od.tel_number).replace('.0', '') if od.tel_number else None,
                    'mobile_no': od.mobile_no,
                    'email': od.email,
                    'mpan_mpr': od.mpan_mpr,
                    'mpan_bottom': od.mpan_bottom,
                    'start_date': _iso(od.start_date),
                    'end_date': _iso(od.end_date),
                    'service_id': od.service_id,
                    'stage_id': od.stage_id,
                    'stage_name': row.stage_name,
                    'opportunity_owner_employee_id': od.opportunity_owner_employee_id,
                    'assigned_to_name': row.assigned_to_name,
                    'created_at': _iso(od.created_at),
                    'supplier_id': od.supplier_id,
                    'supplier_name': row.supplier_name,
                    'annual_usage': od.annual_usage,
                    'stand_charge': od.stand_charge,
                    'rate_1': od.rate_1,
                    'net_notch': od.net_notch,
                    'payment_type': od.payment_type,
                    'postcode': od.postcode,
                })

            current_app.logger.warning(
                'crm.get_leads result tenant=%s user_id=%s employee_id=%s is_admin=%s returned=%s first_ids=%s',
                tenant_id, user_id, employee_id, admin_user, len(results),
                [r.get('tenant_lead_id') or r.get('opportunity_id') for r in results[:5]]
            )

            # ✅ Build team stats for admin
            if admin_user:
                stats_rows = (
                    session.query(
                        Employee_Master.employee_id,
                        Employee_Master.employee_name,
                        func.count(Opportunity_Details.opportunity_id).label('count')
                    )
                    .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
                    .join(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
                    .filter(
                        (Opportunity_Details.tenant_id == tenant_id) |
                        ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
                    )
                    .filter(Opportunity_Details.service_id == service_id)
                    .filter(Opportunity_Details.opportunity_owner_employee_id.isnot(None))
                    .filter((Opportunity_Details.is_draft == False) | (Opportunity_Details.is_draft.is_(None)))
                    .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
                    .filter(
                        (Client_Master.is_deleted.is_(None)) |
                        (Client_Master.is_deleted == False)
                    )
                    .group_by(Employee_Master.employee_id, Employee_Master.employee_name)
                    .having(func.count(Opportunity_Details.opportunity_id) > 0)
                    .order_by(Employee_Master.employee_name.asc())
                    .all()
                )

                team_stats = [
                    {
                        'employee_id': r.employee_id,
                        'employee_name': r.employee_name,
                        'count': int(r.count or 0),
                    }
                    for r in stats_rows
                ]

                return jsonify({
                    'data': results,
                    'team_stats': team_stats,
                    'total': len(results)
                }), 200

            # Non-admin: return plain array (frontend handles both formats)
            return jsonify(results), 200

        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@crm_bp.route('/leads/<int:opportunity_id>', methods=['GET'])
@token_required
@tenant_from_jwt
def get_lead_detail(opportunity_id):
    """
    GET /api/crm/leads/<id>
    Uses opportunity_id (primary key) as the primary lookup.
    Pass ?use_display_id=true to use tenant_lead_id instead.
    """
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        use_display_id = request.args.get('use_display_id', 'false').lower() == 'true'
 
        # Build base query
        base_query = (
            session.query(
                Opportunity_Details,
                Stage_Master.stage_name,
                Employee_Master.employee_name.label('assigned_to_name'),
                func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name'),
                Supplier_Master.supplier_company_name.label('supplier_name')
            )
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
            .filter(Opportunity_Details.tenant_id == tenant_id)
        )
 
        # Choose filter based on query parameter
        if use_display_id:
            row = base_query.filter(Opportunity_Details.tenant_lead_id == opportunity_id).first()
        else:
            # DEFAULT: Use opportunity_id (the primary key)
            row = base_query.filter(Opportunity_Details.opportunity_id == opportunity_id).first()
 
        if not row:
            return jsonify({'error': 'Lead not found'}), 404
 
        od = row[0]
        result = {
            **{k: _serial(getattr(od, k)) for k in od.__table__.columns.keys()},
            'stage_name': row.stage_name,
            'assigned_to_name': row.assigned_to_name,
            'business_name': row.business_name,
            'supplier_name': row.supplier_name,
        }
        
        return jsonify(result), 200
 
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


def _resolve_lead_client_id(db, tenant_id, opportunity_id):
    """Resolve lead by tenant_lead_id or opportunity_id; return (real_opportunity_id, client_id) or (None, None)."""
    lead = db.execute_query(
        '''
        SELECT opportunity_id, client_id
        FROM "StreemLyne_MT"."Opportunity_Details"
        WHERE tenant_id = %s
          AND ("tenant_lead_id" = %s OR opportunity_id = %s)
        LIMIT 1
        ''',
        (tenant_id, opportunity_id, opportunity_id),
        fetch_one=True,
    )
    if not lead:
        return None, None
    return lead.get('opportunity_id'), lead.get('client_id')


@crm_bp.route('/leads/<int:opportunity_id>/history', methods=['GET'])
@token_required
@tenant_from_jwt
def get_lead_history(opportunity_id):
    """
    GET /api/crm/leads/<id>/history
    Returns Client_Interactions for the lead's client.
    Uses opportunity_id (primary key) as the primary lookup.
    Pass ?use_display_id=true to use tenant_lead_id instead.
    Response: { "interactions": [ { interaction_id, interaction_type, notes, ... } ] }
    """
    session = SessionLocal()
    try:
        print(f"\n🔍 get_lead_history called for opportunity_id={opportunity_id}")
        
        tenant_id = g.tenant_id
        use_display_id = request.args.get('use_display_id', 'false').lower() == 'true'
        print(f"   tenant_id: {tenant_id}, use_display_id: {use_display_id}")
        
        # Resolve lead
        query = (
            session.query(Opportunity_Details.opportunity_id, Opportunity_Details.client_id)
            .filter(Opportunity_Details.tenant_id == tenant_id)
        )
        
        if use_display_id:
            query = query.filter(Opportunity_Details.tenant_lead_id == opportunity_id)
        else:
            # DEFAULT: Use opportunity_id (the primary key)
            query = query.filter(Opportunity_Details.opportunity_id == opportunity_id)
        
        lead = query.first()
 
        if not lead:
            return jsonify({'error': 'Lead not found'}), 404
 
        client_id = lead.client_id
        if not client_id:
            return jsonify({'interactions': []}), 200
 
        print(f"   ✅ Fetching interactions for client_id={client_id}")
 
        # Get interactions
        rows = (
            session.query(Client_Interactions)
            .filter(Client_Interactions.client_id == client_id)
            .order_by(Client_Interactions.created_at.desc().nullslast(), Client_Interactions.interaction_id.desc())
            .all()
        )
        
        print(f"   Found {len(rows)} interactions")
 
        interactions = []
        for ci in rows:
            notes = ci.notes or ''
            m = re.match(r'^\[([^\]]+)\]', notes.strip())
            interaction_type = (ci.next_steps or '').strip() or (m.group(1) if m else 'Note')
            
            interactions.append({
                'interaction_id': ci.interaction_id,
                'interaction_type': interaction_type,
                'contact_date': _serial(ci.contact_date),
                'reminder_date': _serial(ci.reminder_date),
                'notes': notes,
                'created_at': _serial(ci.created_at),
            })
 
        print(f"   ✅ Returning {len(interactions)} interactions")
        return jsonify({'interactions': interactions}), 200
 
    except Exception as e:
        print(f"   ❌ Exception in get_lead_history: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/<int:opportunity_id>/history/<int:interaction_id>', methods=['DELETE'])
@token_required
@tenant_from_jwt
def delete_lead_history(opportunity_id, interaction_id):
    """Same fix applied - use opportunity_id by default"""
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        use_display_id = request.args.get('use_display_id', 'false').lower() == 'true'
 
        # Resolve lead
        query = (
            session.query(Opportunity_Details.opportunity_id, Opportunity_Details.client_id)
            .filter(Opportunity_Details.tenant_id == tenant_id)
        )
        
        if use_display_id:
            query = query.filter(Opportunity_Details.tenant_lead_id == opportunity_id)
        else:
            query = query.filter(Opportunity_Details.opportunity_id == opportunity_id)
        
        lead = query.first()
 
        if not lead:
            return jsonify({'error': 'Lead not found'}), 404
 
        client_id = lead.client_id
        if not client_id:
            return jsonify({'error': 'No client for this lead'}), 400
 
        # Check interaction exists
        interaction = (
            session.query(Client_Interactions)
            .filter(Client_Interactions.interaction_id == interaction_id)
            .filter(Client_Interactions.client_id == client_id)
            .first()
        )
 
        if not interaction:
            return jsonify({'error': 'Interaction not found'}), 404
 
        session.delete(interaction)
        session.commit()
 
        return jsonify({'success': True}), 200
 
    except Exception as e:
        session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads', methods=['POST'])
@token_required
@tenant_from_jwt
def create_lead():
    """
    Create a new lead

    Request Body:
        - opportunity_name: Lead/opportunity name (required)
        - client_name: Client name (required)
        - stage_id: Stage identifier (optional)
        - status: Lead status (optional)
        - estimated_value: Estimated value (optional)
        - assigned_to: Assigned user ID (optional)

    Authentication:
        - JWT (token must include `tenant_id`)

    Returns:
        201: Lead created successfully
        400: Invalid request data
        500: Internal server error
    """
    return crm_controller.create_lead()


@crm_bp.route('/leads/<int:opportunity_id>', methods=['PUT', 'PATCH'])
@token_required
@tenant_from_jwt
def update_lead(opportunity_id):
    """
    PUT  /api/crm/leads/<opportunity_id>  — full update via controller
    PATCH /api/crm/leads/<opportunity_id> — partial field update (any fields)
    """
    if request.method == 'PATCH':
        session = SessionLocal()
        try:
            tenant_id = g.tenant_id
            data = request.get_json() or {}
            
            current_app.logger.info(f"🔧 PATCH /api/crm/leads/{opportunity_id}")
            current_app.logger.info(f"   Data: {data}")
 
            ALLOWED_PATCH_FIELDS = {
                'stage_id', 'status',
                'business_name', 'contact_person', 'tel_number', 'mobile_no',
                'email', 'position', 'company_number', 'date_of_birth',
                'opportunity_owner_employee_id',
                'mpan_mpr', 'mpan_bottom', 'supplier_id',
                'annual_usage', 'start_date', 'end_date', 'payment_type',
                'term_sold', 'net_notch', 'comms_paid', 'aggregator',
                'site_name', 'month_sold',
                'house_name', 'house_number', 'door_number', 'address',
                'town', 'county', 'postcode',
                'stand_charge', 'rate_1', 'rate_2', 'rate_3',
                'night_charge', 'eve_weekend_charge',
                'other_charges_1', 'other_charges_2', 'other_charges_3',
                'bank_name', 'bank_account_number', 'bank_sort_code',
                'charity_ltd_company_number', 'partner_details',
                'meter_ref', 'uplift', 'comments', 'document_details',
            }
 
            update_fields = {
                k: v for k, v in data.items()
                if k in ALLOWED_PATCH_FIELDS
            }
 
            if not update_fields:
                return jsonify({'error': 'No valid fields provided'}), 400
 
            # Resolve opportunity_id
            lead = (
                session.query(Opportunity_Details.opportunity_id, Opportunity_Details.client_id)
                .filter(Opportunity_Details.tenant_id == tenant_id)
                .filter(
                    (Opportunity_Details.tenant_lead_id == opportunity_id) |
                    (Opportunity_Details.opportunity_id == opportunity_id)
                )
                .first()
            )
 
            if not lead:
                return jsonify({'error': 'Lead not found'}), 404
 
            real_id = lead.opportunity_id
            client_id = lead.client_id
            
            current_app.logger.info(f"   Resolved: real_id={real_id}, client_id={client_id}")
 
            # ✅ CAPTURE OLD VALUES BEFORE UPDATE
            old_lead = session.query(Opportunity_Details).filter(
                Opportunity_Details.opportunity_id == real_id,
                Opportunity_Details.tenant_id == tenant_id
            ).first()
 
            if not old_lead:
                return jsonify({'error': 'Lead not found'}), 404
 
            # Store old values for comparison
            old_values = {}
            tracked_fields = [
                'net_notch', 'aggregator', 'rate_1', 'rate_2', 'rate_3',
                'stand_charge', 'annual_usage', 'comms_paid', 'term_sold',
                'payment_type', 'supplier_id', 'start_date', 'end_date',
                'mpan_mpr', 'mpan_bottom', 'business_name', 'contact_person',
                'tel_number', 'mobile_no', 'email', 'address', 'postcode',
            ]
            
            for field in tracked_fields:
                if hasattr(old_lead, field):
                    old_values[field] = getattr(old_lead, field)
 
            current_app.logger.info(f"   Old values captured: {len(old_values)} fields")
 
            # Update the lead
            session.query(Opportunity_Details).filter(
                Opportunity_Details.opportunity_id == real_id,
                Opportunity_Details.tenant_id == tenant_id
            ).update(update_fields, synchronize_session=False)
            
            session.flush()
            current_app.logger.info(f"   ✅ Lead updated in database")
 
            # ✅ ENSURE CLIENT EXISTS
            if not client_id:
                current_app.logger.warning(f'   ⚠️ Lead {real_id} has no client_id, creating one...')
                
                client = Client_Master(
                    tenant_id=int(tenant_id),
                    assigned_employee_id=old_lead.opportunity_owner_employee_id,
                    client_company_name=old_lead.business_name or old_lead.opportunity_title or '[IMPORTED LEADS]',
                    client_contact_name=old_lead.contact_person or '',
                    client_phone=old_lead.tel_number or '',
                    client_mobile=old_lead.mobile_no or '',
                    client_email=old_lead.email or '',
                    default_currency_id=1,
                    created_at=datetime.utcnow()
                )
                session.add(client)
                session.flush()
                
                client_id = client.client_id
                old_lead.client_id = client_id
                session.flush()
                current_app.logger.info(f'   ✅ Created client_id={client_id} for lead {real_id}')
 
            # ✅ COMPREHENSIVE FIELD CHANGE TRACKING
            if client_id:
                field_labels = {
                    'net_notch': 'Net Notch',
                    'aggregator': 'Aggregator',
                    'rate_1': 'Rate 1',
                    'rate_2': 'Rate 2',
                    'rate_3': 'Rate 3',
                    'stand_charge': 'Standing Charge',
                    'annual_usage': 'Annual Usage',
                    'comms_paid': 'Comms Paid',
                    'term_sold': 'Term Sold',
                    'payment_type': 'Payment Type',
                    'supplier_id': 'Supplier',
                    'start_date': 'Start Date',
                    'end_date': 'End Date',
                    'mpan_mpr': 'MPAN/MPR',
                    'mpan_bottom': 'MPAN Bottom',
                    'business_name': 'Business Name',
                    'contact_person': 'Contact Person',
                    'tel_number': 'Phone',
                    'mobile_no': 'Mobile',
                    'email': 'Email',
                    'address': 'Address',
                    'postcode': 'Postcode',
                }
                
                changes = []
                end_date_changed = False
                new_end_date = None
                
                # Compare each field
                for field, label in field_labels.items():
                    if field in update_fields:
                        old_val = old_values.get(field)
                        new_val = update_fields[field]
                        
                        # Handle None/empty values
                        def format_value(val):
                            if val is None or val == '':
                                return 'None'
                            return val
                        
                        old_display = format_value(old_val)
                        new_display = format_value(new_val)
                        
                        # Convert to string for comparison
                        old_str = str(old_display)
                        new_str = str(new_display)
                        
                        # Only log if changed
                        if old_str != new_str:
                            # ✅ Track end_date changes specially for calendar
                            if field == 'end_date':
                                end_date_changed = True
                                new_end_date = new_val
                            
                            try:
                                # Special handling for supplier_id - show supplier name
                                if field == 'supplier_id':
                                    old_supplier_name = 'None'
                                    new_supplier_name = 'None'
                                    
                                    if old_val:
                                        old_sup = session.query(Supplier_Master).filter_by(supplier_id=old_val).first()
                                        if old_sup:
                                            old_supplier_name = old_sup.supplier_company_name
                                    
                                    if new_val:
                                        new_sup = session.query(Supplier_Master).filter_by(supplier_id=new_val).first()
                                        if new_sup:
                                            new_supplier_name = new_sup.supplier_company_name
                                    
                                    old_display = old_supplier_name
                                    new_display = new_supplier_name
                                
                                # Special handling for dates
                                elif field in ['start_date', 'end_date']:
                                    if hasattr(old_val, 'isoformat'):
                                        old_display = old_val.strftime('%d/%m/%Y')
                                    if isinstance(new_val, str):
                                        try:
                                            new_date = datetime.strptime(new_val, '%Y-%m-%d')
                                            new_display = new_date.strftime('%d/%m/%Y')
                                        except:
                                            new_display = new_val
                                    elif hasattr(new_val, 'strftime'):
                                        new_display = new_val.strftime('%d/%m/%Y')
                                
                                changes.append(f"{label}: {old_display} → {new_display}")
                            
                            except Exception as field_error:
                                current_app.logger.warning(f"   ⚠️ Error formatting {field}: {field_error}")
                                changes.append(f"{label} updated")
                
                # ✅ Log all changes in one interaction
                if changes:
                    change_summary = " | ".join(changes)
                    current_app.logger.info(f"   📝 Logging changes: {change_summary}")
                    
                    try:
                        # ✅ CRITICAL: If end_date changed, set reminder_date for calendar
                        reminder_date = None
                        if end_date_changed and new_end_date:
                            if isinstance(new_end_date, str):
                                reminder_date = datetime.strptime(new_end_date, '%Y-%m-%d').date()
                            elif hasattr(new_end_date, 'date'):
                                reminder_date = new_end_date.date() if callable(new_end_date.date) else new_end_date
                            else:
                                reminder_date = new_end_date
                            
                            current_app.logger.info(f"   📅 Setting calendar reminder for end_date: {reminder_date}")
                        
                        interaction = Client_Interactions(
                            client_id=client_id,
                            contact_date=datetime.utcnow().date(),
                            contact_method=1,
                            notes=f"[Field Update] {change_summary}",
                            next_steps='Field Update',
                            reminder_date=reminder_date,  # ✅ THIS PUTS IT IN CALENDAR
                            created_at=datetime.utcnow()
                        )
                        session.add(interaction)
                        session.flush()
                        current_app.logger.info(f"   ✅ History logged successfully (interaction_id={interaction.interaction_id})")
                    except Exception as log_error:
                        current_app.logger.error(f"   ❌ Error logging history: {log_error}")
                        pass
                else:
                    current_app.logger.info(f"   ℹ️  No changes detected to log")
            else:
                current_app.logger.warning(f"   ⚠️ No client_id - skipping history log")
 
            session.commit()
            current_app.logger.info(f"   ✅ Transaction committed")
 
            # Fetch updated record
            updated = (
                session.query(
                    Opportunity_Details,
                    Stage_Master.stage_name,
                    Employee_Master.employee_name.label('assigned_to_name'),
                    func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name'),
                    Supplier_Master.supplier_company_name.label('supplier_name')
                )
                .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
                .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
                .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
                .filter(Opportunity_Details.opportunity_id == real_id)
                .filter(Opportunity_Details.tenant_id == tenant_id)
                .first()
            )
 
            if updated:
                od = updated[0]
                result = {
                    **{k: _serial(getattr(od, k)) for k in od.__table__.columns.keys()},
                    'stage_name': updated.stage_name,
                    'assigned_to_name': updated.assigned_to_name,
                    'business_name': updated.business_name,
                    'supplier_name': updated.supplier_name,
                }
                current_app.logger.info(f"   ✅ PATCH completed successfully")
                return jsonify(result), 200
 
            return jsonify({'success': True}), 200
 
        except Exception as e:
            session.rollback()
            current_app.logger.exception(f"❌ Error in PATCH /api/crm/leads/{opportunity_id}: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()
 
    # PUT — full update via controller
    return crm_controller.update_lead(opportunity_id)

@crm_bp.route('/leads/<int:opportunity_id>/status', methods=['PATCH'])
@token_required
@tenant_from_jwt
def update_lead_status(opportunity_id):
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        data = request.get_json() or {}

        stage_id = data.get('stage_id')
        status = data.get('status')

        current_app.logger.info(f'🔧 PATCH /api/crm/leads/{opportunity_id}/status — data: {data}')

        if not stage_id and not status:
            return jsonify({'error': 'Either stage_id or status is required'}), 400

        # ✅ Always resolve stage from DB
        if status and not stage_id:
            stage = session.query(Stage_Master).filter(Stage_Master.stage_name == status).first()
            if not stage:
                return jsonify({'error': f'Stage not found: {status}'}), 400
            stage_id = stage.stage_id
        else:
            stage = session.query(Stage_Master).filter(Stage_Master.stage_id == stage_id).first()
            if not stage:
                return jsonify({'error': f'Stage not found for id: {stage_id}'}), 400
            status = stage.stage_name

        # Resolve opportunity
        lead = (
            session.query(Opportunity_Details)
            .filter(Opportunity_Details.tenant_id == tenant_id)
            .filter(
                (Opportunity_Details.opportunity_id == opportunity_id) |
                (Opportunity_Details.tenant_lead_id == opportunity_id)
            )
            .first()
        )

        if not lead:
            return jsonify({'error': 'Lead not found'}), 404

        real_id = lead.opportunity_id
        lead_client_id = lead.client_id

        # ✅ Capture old stage for history
        old_stage_name = None
        if lead.stage_id:
            old_stage_row = session.query(Stage_Master).filter_by(stage_id=lead.stage_id).first()
            old_stage_name = old_stage_row.stage_name if old_stage_row else None

        # ✅ CRITICAL: Create client if missing
        if not lead_client_id:
            current_app.logger.warning(f'Lead {real_id} has no client_id, creating one...')
            client = Client_Master(
                tenant_id=int(tenant_id),
                assigned_employee_id=lead.opportunity_owner_employee_id,
                client_company_name=lead.business_name or lead.opportunity_title or '[IMPORTED LEADS]',
                client_contact_name=lead.contact_person or '',
                client_phone=lead.tel_number or '',
                client_mobile=lead.mobile_no or '',
                client_email=lead.email or '',
                default_currency_id=1,
                created_at=datetime.utcnow()
            )
            session.add(client)
            session.flush()
            lead_client_id = client.client_id
            lead.client_id = lead_client_id
            session.flush()

        # ✅ Update stage
        lead.stage_id = stage_id
        session.flush()
        current_app.logger.info(f'✅ Updated lead {real_id}: {old_stage_name} → {status}')

        # ✅ Write history BEFORE any early returns
        transition = f"Status: {old_stage_name or 'None'} → {status}"
        notes = data.get('notes', '')
        formatted_notes = f"[{status}] {transition}" + (f" | {notes}" if notes else "")

        interaction = Client_Interactions(
            client_id=lead_client_id,
            contact_date=datetime.utcnow().date(),
            contact_method=1,
            notes=formatted_notes,
            next_steps=status,
            created_at=datetime.utcnow()
        )
        session.add(interaction)
        session.flush()

        CLEANSING_STATUSES = {'Invalid Number', 'Incorrect Supplier'}
        RECYCLE_BIN_STATUSES = {'Lost', 'Lost COT', 'Meter De-energised', 'Complaint'}

        if status in CLEANSING_STATUSES:
            client = session.query(Client_Master).filter(
                Client_Master.client_id == lead_client_id,
                Client_Master.tenant_id == tenant_id
            ).first()
            if client:
                client.is_deleted = True
                client.deleted_at = datetime.utcnow()
                client.deleted_reason = status
                if hasattr(client, 'is_cleansing'):
                    client.is_cleansing = True
                session.flush()

            session.commit()
            return jsonify({
                'success': True,
                'message': f'Status updated to {status}',
                'stage_id': stage_id,
                'stage_name': status,
                'moved_to_cleansing': True,
                'moved_to_recycle_bin': False,
                'interaction_id': interaction.interaction_id,
            }), 200

        elif status in RECYCLE_BIN_STATUSES:
            client = session.query(Client_Master).filter(
                Client_Master.client_id == lead_client_id,
                Client_Master.tenant_id == tenant_id
            ).first()
            if client:
                client.is_deleted = True
                client.deleted_at = datetime.utcnow()
                client.deleted_reason = status
                if hasattr(client, 'is_cleansing'):
                    client.is_cleansing = False
                session.flush()

            session.commit()
            return jsonify({
                'success': True,
                'message': f'Status updated to {status}',
                'stage_id': stage_id,
                'stage_name': status,
                'moved_to_cleansing': False,
                'moved_to_recycle_bin': True,
                'interaction_id': interaction.interaction_id,
            }), 200

        session.commit()
        return jsonify({
            'success': True,
            'message': f'Status updated to {status}',
            'stage_id': stage_id,
            'stage_name': status,
            'moved_to_cleansing': False,
            'moved_to_recycle_bin': False,
            'interaction_id': interaction.interaction_id,
        }), 200

    except Exception as e:
        session.rollback()
        current_app.logger.exception(f'❌ Error in update_lead_status: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/assign', methods=['PATCH'])
@token_required
@tenant_from_jwt
def assign_leads():
    """
    PATCH /api/crm/leads/assign
    Bulk assign leads to an employee. Available to all users.
    Request body: { lead_ids: [...], employee_id: N }
    """
    return crm_controller.assign_leads()


@crm_bp.route('/leads/<int:opportunity_id>', methods=['DELETE'])
@token_required
@tenant_from_jwt
def delete_lead(opportunity_id):
    """
    Delete a lead

    Path Parameters:
        - opportunity_id: Opportunity identifier

    Authentication:
        - JWT (token must include `tenant_id`)

    Returns:
        200: Lead deleted successfully
        404: Lead not found
        500: Internal server error
    """
    return crm_controller.delete_lead(opportunity_id)


@crm_bp.route('/leads/drafts', methods=['DELETE', 'OPTIONS'])
@token_required
@tenant_from_jwt
def delete_draft_leads():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        data = request.get_json(silent=True) or {}
        raw_ids = data.get('lead_ids') or []

        try:
            lead_ids = sorted({int(lead_id) for lead_id in raw_ids})
        except (TypeError, ValueError):
            return jsonify({'error': 'lead_ids must be a list of numbers'}), 400

        if not lead_ids:
            return jsonify({'error': 'lead_ids are required'}), 400

        # Delete draft leads (is_draft=TRUE and no employee assigned)
        deleted_count = (
            session.query(Opportunity_Details)
            .filter(Opportunity_Details.tenant_id == tenant_id)
            .filter(Opportunity_Details.opportunity_id.in_(lead_ids))
            .filter(Opportunity_Details.opportunity_owner_employee_id.is_(None))
            .filter(Opportunity_Details.is_draft == True)
            .delete(synchronize_session=False)
        )

        session.commit()

        skipped_ids = [lead_id for lead_id in lead_ids if lead_id not in range(deleted_count)]

        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'deleted_ids': lead_ids[:deleted_count],
            'skipped_ids': skipped_ids,
            'message': f'Deleted {deleted_count} draft leads'
        }), 200

    except Exception as e:
        session.rollback()
        current_app.logger.exception("Error deleting draft leads")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/drafts', methods=['GET'])
@token_required
@tenant_from_jwt
def get_draft_leads():
    """
    GET /api/crm/leads/drafts
    Get all unassigned draft leads (is_draft=TRUE, opportunity_owner_employee_id IS NULL)
    """
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1

        rows = (
            session.query(
                Opportunity_Details,
                Stage_Master.stage_name,
                Supplier_Master.supplier_company_name.label('supplier_name'),
                func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name')
            )
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter(Opportunity_Details.is_draft == True)
            .filter(Opportunity_Details.opportunity_owner_employee_id.is_(None))
            .order_by(Opportunity_Details.created_at.desc())
            .all()
        )

        results = []
        for row in rows:
            od = row[0]
            results.append({
                'opportunity_id': od.opportunity_id,
                'tenant_lead_id': od.tenant_lead_id,
                'business_name': row.business_name,
                'contact_person': od.contact_person,
                'tel_number': str(od.tel_number).replace('.0', '') if od.tel_number else None,
                'mobile_no': od.mobile_no,
                'email': od.email,
                'mpan_mpr': od.mpan_mpr,
                'mpan_bottom': od.mpan_bottom,
                'start_date': _iso(od.start_date),
                'end_date': _iso(od.end_date),
                'service_id': od.service_id,
                'stage_id': od.stage_id,
                'stage_name': row.stage_name,
                'created_at': _iso(od.created_at),
                'supplier_id': od.supplier_id,
                'supplier_name': row.supplier_name,
                'annual_usage': od.annual_usage,
                'stand_charge': od.stand_charge,
                'rate_1': od.rate_1,
                'net_notch': od.net_notch,
                'payment_type': od.payment_type,
                'postcode': od.postcode,
            })

        return jsonify(results), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/search-all', methods=['GET'])
@token_required
@tenant_from_jwt
def search_all_leads():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        q = request.args.get('q', '').strip()
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1

        if not q or len(q) < 2:
            return jsonify([]), 200

        like_q = f'%{q.lower()}%'

        rows = (
            session.query(
                Opportunity_Details,
                Stage_Master.stage_name,
                Employee_Master.employee_name.label('assigned_to_name'),
                func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name'),
                Supplier_Master.supplier_company_name.label('supplier_name')
            )
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter(Opportunity_Details.opportunity_owner_employee_id.isnot(None))
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
            .filter(
                (func.lower(func.coalesce(Opportunity_Details.business_name, Client_Master.client_company_name, Opportunity_Details.opportunity_title, '')).like(like_q)) |
                (func.lower(func.coalesce(Opportunity_Details.contact_person, '')).like(like_q)) |
                (func.lower(func.coalesce(func.cast(Opportunity_Details.tel_number, String), '')).like(like_q)) |
                (func.lower(func.coalesce(Opportunity_Details.email, '')).like(like_q)) |
                (func.lower(func.coalesce(Opportunity_Details.mpan_mpr, '')).like(like_q))
            )
            .order_by(Opportunity_Details.created_at.desc())
            .all()
        )

        results = []
        for row in rows:
            od = row[0]
            results.append({
                'opportunity_id': od.opportunity_id,
                'tenant_lead_id': od.tenant_lead_id,
                'business_name': row.business_name,
                'contact_person': od.contact_person,
                'tel_number': str(od.tel_number).replace('.0', '') if od.tel_number else None,
                'mobile_no': od.mobile_no,
                'email': od.email,
                'mpan_mpr': od.mpan_mpr,
                'mpan_bottom': od.mpan_bottom,
                'start_date': _iso(od.start_date),
                'end_date': _iso(od.end_date),
                'service_id': od.service_id,
                'stage_id': od.stage_id,
                'stage_name': row.stage_name,
                'opportunity_owner_employee_id': od.opportunity_owner_employee_id,
                'assigned_to_name': row.assigned_to_name,
                'created_at': _iso(od.created_at),
                'supplier_id': od.supplier_id,
                'supplier_name': row.supplier_name,
                'annual_usage': od.annual_usage,
                'stand_charge': od.stand_charge,
                'rate_1': od.rate_1,
                'net_notch': od.net_notch,
                'payment_type': od.payment_type,
                'postcode': od.postcode,
            })

        return jsonify(results), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/stats', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_stats():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1

        current_user = request.current_user
        role_name = getattr(current_user, 'role', None)
        admin_user = is_crm_leads_admin_role(role_name)
        my_emp_id = getattr(current_user, 'employee_id', None)
        
        # ✅ For non-admins: always use their own employee_id
        # ✅ For admins: allow ?employee_id=X to view specific user, otherwise show all
        if not admin_user:
            employee_id = my_emp_id
        else:
            requested_employee_id = request.args.get('employee_id', type=int)
            employee_id = requested_employee_id  # Can be None (all) or specific employee

        if local_demo_dashboard_enabled():
            return jsonify(dummy_leads_stats(employee_id)), 200

        # Build query
        query = (
            session.query(
                Opportunity_Details.opportunity_id,
                Opportunity_Details.created_at,
                Opportunity_Details.end_date,
                Opportunity_Details.annual_usage,
                func.coalesce(Stage_Master.stage_name, 'Unknown').label('stage_name')
            )
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
        )

        # ✅ Always filter by employee_id for non-admins
        if employee_id:
            query = query.filter(Opportunity_Details.opportunity_owner_employee_id == employee_id)

        rows = query.all()

        today = datetime.utcnow().date()
        last_30 = today - timedelta(days=30)

        converted_leads = 0
        in_progress = 0
        lost_leads = 0
        new_leads = 0
        leads_30_60 = 0
        leads_61_90 = 0
        leads_91_180 = 0
        not_due = 0
        total_annual_usage = 0.0
        recent_30d = 0
        stage_breakdown = {}

        for r in rows:
            stage_name = r.stage_name or 'Unknown'
            bucket = _lead_stage_bucket(stage_name)
            if bucket == "converted":
                converted_leads += 1
            elif bucket == "in_progress":
                in_progress += 1
            elif bucket == "lost":
                lost_leads += 1
            else:
                new_leads += 1

            stage_breakdown[stage_name] = stage_breakdown.get(stage_name, 0) + 1
            total_annual_usage += _safe_float(r.annual_usage)

            if r.created_at and hasattr(r.created_at, 'date') and r.created_at.date() >= last_30:
                recent_30d += 1

            if r.end_date:
                d = r.end_date if not hasattr(r.end_date, 'date') else r.end_date.date()
                days = (d - today).days
                if 30 <= days <= 60:
                    leads_30_60 += 1
                elif 61 <= days <= 90:
                    leads_61_90 += 1
                elif 91 <= days <= 180:
                    leads_91_180 += 1
                elif days >= 365:
                    not_due += 1

        total_leads = len(rows)
        active_leads = max(0, total_leads - lost_leads)
        conversion_rate = round((converted_leads / total_leads) * 100, 1) if total_leads else 0

        return jsonify({
            'total_leads': total_leads,
            'active_leads': active_leads,
            'converted_leads': converted_leads,
            'new_leads': new_leads,
            'in_progress': in_progress,
            'lost_leads': lost_leads,
            'conversion_rate': conversion_rate,
            'total_value': 0,
            'recent_leads_30d': recent_30d,
            'allocated_leads': 0,
            'unallocated_leads': total_leads,
            'stage_breakdown': stage_breakdown,
            'leads_30_60_days': leads_30_60,
            'leads_61_90_days': leads_61_90,
            'leads_91_180_days': leads_91_180,
            'not_due_leads': not_due,
            'total_annual_usage': total_annual_usage,
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads/stage-breakdown', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_stage_breakdown():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1
        current_user = request.current_user
        admin_user = is_crm_leads_admin_role(getattr(current_user, 'role', None))
        employee_id = request.args.get('employee_id', type=int) if admin_user else getattr(current_user, 'employee_id', None)

        if local_demo_dashboard_enabled():
            return jsonify(dummy_leads_stage_breakdown(employee_id)), 200

        query = (
            session.query(
                func.coalesce(Stage_Master.stage_name, 'Unknown').label('stage_name'),
                func.count(Opportunity_Details.opportunity_id).label('count')
            )
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
        )

        if employee_id:
            query = query.filter(Opportunity_Details.opportunity_owner_employee_id == employee_id)

        query = query.group_by(func.coalesce(Stage_Master.stage_name, 'Unknown'))
        query = query.order_by(func.count(Opportunity_Details.opportunity_id).desc())

        rows = query.all()

        return jsonify([
            {
                'stage_id': i + 1,
                'stage_name': r.stage_name or 'Unknown',
                'count': int(r.count or 0),
                'total_value': 0
            }
            for i, r in enumerate(rows)
        ]), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads/supplier-breakdown', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_supplier_breakdown():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1
        current_user = request.current_user
        admin_user = is_crm_leads_admin_role(getattr(current_user, 'role', None))
        employee_id = request.args.get('employee_id', type=int) if admin_user else getattr(current_user, 'employee_id', None)

        if local_demo_dashboard_enabled():
            return jsonify(dummy_leads_supplier_breakdown(employee_id)), 200

        query = (
            session.query(
                func.coalesce(Supplier_Master.supplier_company_name, 'Unknown').label('supplier_name'),
                func.count(Opportunity_Details.opportunity_id).label('lead_count')
            )
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
        )

        if employee_id:
            query = query.filter(Opportunity_Details.opportunity_owner_employee_id == employee_id)

        query = query.group_by(func.coalesce(Supplier_Master.supplier_company_name, 'Unknown'))
        query = query.order_by(func.count(Opportunity_Details.opportunity_id).desc())

        rows = query.all()

        return jsonify([
            {
                'supplier_name': r.supplier_name or 'Unknown',
                'lead_count': int(r.lead_count or 0),
                'total_value': 0
            }
            for r in rows
        ]), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads/salesperson-breakdown', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_salesperson_breakdown():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1
        current_user = request.current_user
        admin_user = is_crm_leads_admin_role(getattr(current_user, 'role', None))
        
        if not admin_user:
            return jsonify([]), 200
            
        if local_demo_dashboard_enabled():
            return jsonify(dummy_leads_salesperson_breakdown()), 200

        rows = (
            session.query(
                Employee_Master.employee_id,
                Employee_Master.employee_name,
                func.coalesce(Stage_Master.stage_name, 'Unknown').label('stage_name'),
                func.count(Opportunity_Details.opportunity_id).label('cnt')
            )
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .join(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
            .group_by(Employee_Master.employee_id, Employee_Master.employee_name, func.coalesce(Stage_Master.stage_name, 'Unknown'))
            .order_by(Employee_Master.employee_name.asc())
            .all()
        )

        grouped = {}
        for r in rows:
            eid = r.employee_id
            if eid not in grouped:
                grouped[eid] = {
                    'employee_id': eid,
                    'employee_name': r.employee_name or 'Unknown',
                    'total_leads': 0,
                    'converted_count': 0,
                    'in_progress_count': 0,
                    'not_contacted_count': 0,
                    'lost_count': 0,
                    'conversion_rate': 0,
                    'total_value': 0,
                }
            c = int(r.cnt or 0)
            grouped[eid]['total_leads'] += c
            bucket = _lead_stage_bucket(r.stage_name or '')
            if bucket == "converted":
                grouped[eid]['converted_count'] += c
            elif bucket == "in_progress":
                grouped[eid]['in_progress_count'] += c
            elif bucket == "lost":
                grouped[eid]['lost_count'] += c
            else:
                grouped[eid]['not_contacted_count'] += c

        out = []
        for v in grouped.values():
            total = v['total_leads'] or 0
            v['conversion_rate'] = round((v['converted_count'] / total) * 100, 1) if total else 0
            out.append(v)
        
        out.sort(key=lambda x: x['total_leads'], reverse=True)
        return jsonify(out), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads/by-stage', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_by_stage():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        stage = (request.args.get('stage') or '').strip().lower()
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1
        current_user = request.current_user
        admin_user = is_crm_leads_admin_role(getattr(current_user, 'role', None))
        employee_id = request.args.get('employee_id', type=int) if admin_user else getattr(current_user, 'employee_id', None)

        if local_demo_dashboard_enabled():
            return jsonify(dummy_leads_by_stage(stage, employee_id)), 200

        query = (
            session.query(
                Opportunity_Details.opportunity_id,
                func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name'),
                Opportunity_Details.contact_person,
                Opportunity_Details.tel_number,
                Opportunity_Details.email,
                func.coalesce(Stage_Master.stage_name, 'Unknown').label('stage_name'),
                func.coalesce(Employee_Master.employee_name, 'Unassigned').label('assigned_to_name'),
                Opportunity_Details.opportunity_owner_employee_id.label('assigned_to_id'),
                Opportunity_Details.created_at,
                Opportunity_Details.annual_usage,
                case(
                    [(Opportunity_Details.service_id == 2, 'water')],
                    else_='utilities'
                ).label('service_name'),
                Opportunity_Details.end_date
            )
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
        )

        if employee_id:
            query = query.filter(Opportunity_Details.opportunity_owner_employee_id == employee_id)

        # Stage filtering
        if stage == 'in_progress':
            query = query.filter(
                func.lower(func.coalesce(Stage_Master.stage_name, '')).in_([
                    'callback', 'not answered', 'broker in place', 'email only',
                    'complaint', 'incorrect supplier', 'priced', 'end date changed'
                ])
            )
        elif stage == 'lost':
            query = query.filter(
                func.lower(func.coalesce(Stage_Master.stage_name, '')).in_([
                    'lost', 'lost cot', 'invalid number', 'meter de-energised'
                ])
            )
        elif stage:
            query = query.filter(func.lower(func.coalesce(Stage_Master.stage_name, '')) == stage)

        query = query.order_by(Opportunity_Details.created_at.desc())
        rows = query.all()

        today = datetime.utcnow().date()
        leads = []
        for r in rows:
            end_date = r.end_date
            end_d = end_date.date() if hasattr(end_date, 'date') else end_date
            days = (end_d - today).days if end_d else None
            
            leads.append({
                'opportunity_id': r.opportunity_id,
                'business_name': r.business_name,
                'contact_person': r.contact_person,
                'tel_number': str(r.tel_number or ''),
                'email': r.email,
                'stage_name': r.stage_name or 'Unknown',
                'opportunity_value': 0,
                'assigned_to_name': r.assigned_to_name or 'Unassigned',
                'assigned_to_id': r.assigned_to_id,
                'created_at': r.created_at.isoformat() if r.created_at and hasattr(r.created_at, 'isoformat') else None,
                'annual_usage': _safe_float(r.annual_usage),
                'service_name': r.service_name or 'utilities',
                'end_date': end_d.isoformat() if end_d and hasattr(end_d, 'isoformat') else None,
                'days_until_due': days,
            })

        return jsonify({'leads': leads}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads/period-breakdown', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_period_breakdown():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        period = (request.args.get('period') or '').strip().lower()
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1
        current_user = request.current_user
        admin_user = is_crm_leads_admin_role(getattr(current_user, 'role', None))
        employee_id = request.args.get('employee_id', type=int) if admin_user else getattr(current_user, 'employee_id', None)

        if local_demo_dashboard_enabled():
            return jsonify(dummy_leads_period_breakdown(period, employee_id)), 200

        query = (
            session.query(
                Opportunity_Details.opportunity_id,
                func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name'),
                Opportunity_Details.contact_person,
                Opportunity_Details.tel_number,
                Opportunity_Details.email,
                func.coalesce(Stage_Master.stage_name, 'Unknown').label('stage_name'),
                func.coalesce(Employee_Master.employee_name, 'Unassigned').label('assigned_to_name'),
                Opportunity_Details.opportunity_owner_employee_id.label('assigned_to_id'),
                Opportunity_Details.created_at,
                Opportunity_Details.annual_usage,
                Opportunity_Details.end_date,
                case(
                    [(Opportunity_Details.service_id == 2, 'water')],
                    else_='utilities'
                ).label('service_name')
            )
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
        )

        if employee_id:
            query = query.filter(Opportunity_Details.opportunity_owner_employee_id == employee_id)

        query = query.order_by(Opportunity_Details.created_at.desc())
        rows = query.all()

        today = datetime.utcnow().date()
        leads = []
        
        for r in rows:
            end_date = r.end_date
            end_d = end_date.date() if hasattr(end_date, 'date') else end_date
            days = (end_d - today).days if end_d else None

            include = False
            if period == '30-60':
                include = days is not None and 30 <= days <= 60
            elif period == '61-90':
                include = days is not None and 61 <= days <= 90
            elif period == '91-180':
                include = days is not None and 91 <= days <= 180
            elif period == 'not-due':
                include = days is not None and days >= 365
            else:
                include = True

            if not include:
                continue

            leads.append({
                'opportunity_id': r.opportunity_id,
                'business_name': r.business_name,
                'contact_person': r.contact_person,
                'tel_number': str(r.tel_number or ''),
                'email': r.email,
                'stage_name': r.stage_name or 'Unknown',
                'opportunity_value': 0,
                'assigned_to_name': r.assigned_to_name or 'Unassigned',
                'assigned_to_id': r.assigned_to_id,
                'created_at': r.created_at.isoformat() if r.created_at and hasattr(r.created_at, 'isoformat') else None,
                'annual_usage': _safe_float(r.annual_usage),
                'service_name': r.service_name or 'utilities',
                'end_date': end_d.isoformat() if end_d and hasattr(end_d, 'isoformat') else None,
                'days_until_due': days,
            })

        return jsonify({'period': period, 'leads': leads}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads/performance', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_performance():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1

        current_user = request.current_user
        role_name = getattr(current_user, 'role', None)
        admin_user = is_crm_leads_admin_role(role_name)
        my_emp_id = getattr(current_user, 'employee_id', None)

        # ✅ Admin sees ALL leads across tenant (no employee filter)
        # ✅ Non-admin sees only their own
        if admin_user:
            requested_employee_id = request.args.get('employee_id', type=int)
            employee_id = requested_employee_id  # None = all tenant, or specific employee
        else:
            employee_id = my_emp_id  # always scoped to themselves

        current_app.logger.warning(
            f'📊 get_leads_performance: tenant={tenant_id}, my_emp_id={my_emp_id}, '
            f'is_admin={admin_user}, filtering_by_employee_id={employee_id}'
        )

        query = (
            session.query(Stage_Master.stage_name)
            .select_from(Opportunity_Details)
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
            # ✅ Exclude soft-deleted leads
            .filter(
                (Client_Master.is_deleted.is_(None)) |
                (Client_Master.is_deleted == False)
            )
        )

        # ✅ Only filter by employee if non-admin OR admin specified ?employee_id=X
        if employee_id:
            query = query.filter(
                Opportunity_Details.opportunity_owner_employee_id == employee_id
            )
        elif not admin_user:
            # Non-admin with no employee_id - return zeros
            return jsonify({
                'converted_count': 0,
                'renewed_count': 0,
                'renewed_directly_count': 0,
                'end_date_changed_count': 0,
                'priced_count': 0,
                'contacted_count': 0,
                'not_contacted_count': 0,
                'lost_count': 0,
                'success_rate': 0,
                'total_customers': 0,
            }), 200

        rows = query.all()

        converted_count = 0
        renewed_count = 0
        in_progress_count = 0
        not_contacted_count = 0
        lost_count = 0
        renewed_directly_count = 0
        end_date_changed_count = 0
        priced_count = 0

        for r in rows:
            stage = (r.stage_name or '').lower()
            if stage == 'converted':
                converted_count += 1
            elif stage in ['already renewed', 'renewed']:
                renewed_count += 1
            elif stage == 'renewed directly':
                renewed_directly_count += 1
            elif stage == 'end date changed':
                end_date_changed_count += 1
            elif stage == 'priced':
                priced_count += 1
            elif stage in ['callback', 'not answered', 'broker in place', 'email only',
                           'complaint', 'incorrect supplier']:
                in_progress_count += 1
            elif stage in ['lost', 'lost cot', 'invalid number', 'meter de-energised']:
                lost_count += 1
            else:
                not_contacted_count += 1

        total = len(rows)
        success_rate = round(
            ((converted_count + renewed_count + renewed_directly_count) / total * 100), 1
        ) if total > 0 else 0

        current_app.logger.warning(
            f'✅ Performance for employee_id={employee_id}: total={total}, '
            f'not_contacted={not_contacted_count}, converted={converted_count}'
        )

        return jsonify({
            'converted_count': converted_count,
            'renewed_count': renewed_count,
            'renewed_directly_count': renewed_directly_count,
            'end_date_changed_count': end_date_changed_count,
            'priced_count': priced_count,
            'contacted_count': in_progress_count,
            'not_contacted_count': not_contacted_count,
            'lost_count': lost_count,
            'success_rate': success_rate,
            'total_customers': total,
        }), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/staff-performance', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_staff_performance():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1
        period, start_dt, end_dt, goal_multiplier = _staff_period_bounds(request.args.get('period', 'daily'))

        current_user = request.current_user
        role_name = getattr(current_user, 'role', None)
        admin_user = is_crm_leads_admin_role(role_name)
        my_emp_id = getattr(current_user, 'employee_id', None)
        only_employee_id = request.args.get('employee_id', type=int)

        if not admin_user:
            if my_emp_id is None:
                return jsonify([]), 200
            only_employee_id = my_emp_id

        if local_demo_dashboard_enabled():
            return jsonify(dummy_leads_staff_performance(period, only_employee_id)), 200

        # Use raw SQL with text() because of complex subquery
        ts_column = 'od."updated_at"'  # or resolve dynamically
        
        sql = text(f"""
            SELECT
                em."employee_id",
                em."employee_name",
                COALESCE(sm."stage_name", 'Unknown') AS stage_name,
                COUNT(od."opportunity_id")::bigint AS cnt,
                CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM "StreemLyne_MT"."User_Master" um
                        INNER JOIN "StreemLyne_MT"."User_Role_Mapping" urm
                            ON urm."user_id" = um."user_id"
                        WHERE um."employee_id" = em."employee_id"
                            AND urm."role_id" = :offshore_role_id
                    ) THEN 1 ELSE 0
                END AS is_offshore
            FROM "StreemLyne_MT"."Opportunity_Details" od
            LEFT JOIN "StreemLyne_MT"."Client_Master" cm
                ON od."client_id" = cm."client_id"
            JOIN "StreemLyne_MT"."Employee_Master" em
                ON od."opportunity_owner_employee_id" = em."employee_id"
            LEFT JOIN "StreemLyne_MT"."Stage_Master" sm
                ON od."stage_id" = sm."stage_id"
            WHERE (od."tenant_id" = :tenant_id OR (od."client_id" IS NOT NULL AND cm."tenant_id" = :tenant_id))
                AND em."tenant_id" = :tenant_id
                AND od."service_id" = :service_id
                AND od."opportunity_owner_employee_id" IS NOT NULL
                AND {ts_column} >= :start_dt
                AND {ts_column} < :end_dt
                AND (od."is_allocated" = FALSE OR od."is_allocated" IS NULL)
                {' AND od."opportunity_owner_employee_id" = :employee_id' if only_employee_id else ''}
            GROUP BY em."employee_id", em."employee_name", COALESCE(sm."stage_name", 'Unknown'), is_offshore
            ORDER BY em."employee_name" ASC
        """)

        params = {
            'tenant_id': tenant_id,
            'service_id': service_id,
            'start_dt': start_dt,
            'end_dt': end_dt,
            'offshore_role_id': OFFSHORE_ROLE_ID,
        }
        if only_employee_id:
            params['employee_id'] = only_employee_id

        rows = session.execute(sql, params).mappings().all()

        employees = {}
        for r in rows:
            eid = r['employee_id']
            if eid is None:
                continue

            if eid not in employees:
                employees[eid] = {
                    'employee_id': eid,
                    'employee_name': r['employee_name'] or 'Unknown',
                    'role_id': OFFSHORE_ROLE_ID if int(r['is_offshore'] or 0) == 1 else None,
                    'converted_count': 0,
                    'in_progress_count': 0,
                    'not_contacted_count': 0,
                    'lost_count': 0,
                    'renewed_count': 0,
                    'renewed_directly_count': 0,
                    'end_date_changed_count': 0,
                    'priced_count': 0,
                    'total_contacts': 0,
                }

            stage = str(r['stage_name'] or '').strip().lower()
            count = int(r['cnt'] or 0)
            employees[eid]['total_contacts'] += count

            if stage == 'converted':
                employees[eid]['converted_count'] += count
            elif stage in ('already renewed', 'renewed'):
                employees[eid]['converted_count'] += count
                employees[eid]['renewed_count'] += count
            elif stage == 'renewed directly':
                employees[eid]['converted_count'] += count
                employees[eid]['renewed_directly_count'] += count
            elif stage == 'end date changed':
                employees[eid]['end_date_changed_count'] += count
            elif stage == 'priced':
                employees[eid]['priced_count'] += count
                employees[eid]['in_progress_count'] += count
            elif stage in ('callback', 'not answered', 'broker in place', 'email only', 'complaint', 'incorrect supplier'):
                employees[eid]['in_progress_count'] += count
            elif stage in ('lost', 'lost cot', 'invalid number', 'meter de-energised'):
                employees[eid]['lost_count'] += count
            else:
                employees[eid]['not_contacted_count'] += count

        output = []
        for emp in employees.values():
            total = emp['total_contacts'] or 0
            converted = emp['converted_count'] or 0
            conversion_rate = round((converted / total) * 100) if total > 0 else 0
            daily_target = 180 if emp.get('role_id') == OFFSHORE_ROLE_ID else 100
            goal_target = daily_target * goal_multiplier
            goal_achieved = total
            goal_progress_pct = round((goal_achieved / goal_target) * 100, 1) if goal_target > 0 else 0

            output.append({
                'employee_id': emp['employee_id'],
                'employee_name': emp['employee_name'],
                'role_id': emp.get('role_id'),
                'total_contacts': total,
                'converted_count': converted,
                'renewed_count': converted,
                'in_progress_count': emp['in_progress_count'],
                'not_contacted_count': emp['not_contacted_count'],
                'lost_count': emp['lost_count'],
                'renewed_directly_count': emp['renewed_directly_count'],
                'end_date_changed_count': emp['end_date_changed_count'],
                'priced_count': emp['priced_count'],
                'conversion_rate': conversion_rate,
                'goal_target': goal_target,
                'goal_achieved': goal_achieved,
                'goal_progress_pct': min(100, max(0, goal_progress_pct)),
                'goal_hit': goal_achieved >= goal_target,
                'period': period,
            })

        output.sort(key=lambda x: x['employee_name'].lower())
        return jsonify(output), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads/stats-by-employee', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_stats_by_employee():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        current_user = request.current_user
        role_name = getattr(current_user, 'role', None)

        # ✅ Only platform admins can see team overview
        if not is_crm_leads_admin_role(role_name):
            return jsonify({'stats': []}), 200

        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1

        rows = (
            session.query(
                Employee_Master.employee_id,
                Employee_Master.employee_name,
                func.count(Opportunity_Details.opportunity_id).label('count')
            )
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .join(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter(Opportunity_Details.opportunity_owner_employee_id.isnot(None))
            .filter((Opportunity_Details.is_draft == False) | (Opportunity_Details.is_draft.is_(None)))
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
            # ✅ Exclude soft-deleted
            .filter(
                (Client_Master.is_deleted.is_(None)) |
                (Client_Master.is_deleted == False)
            )
            .group_by(Employee_Master.employee_id, Employee_Master.employee_name)
            .having(func.count(Opportunity_Details.opportunity_id) > 0)
            .order_by(Employee_Master.employee_name.asc())
            .all()
        )

        stats = [
            {
                'employee_id': r.employee_id,
                'employee_name': r.employee_name,
                'count': int(r.count or 0),
            }
            for r in rows
        ]

        return jsonify({'stats': stats}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/stats-by-employee-detailed', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_stats_by_employee_detailed():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1

        current_user = request.current_user
        role_name = getattr(current_user, 'role', None)
        admin_user = is_crm_leads_admin_role(role_name)
        my_emp_id = getattr(current_user, 'employee_id', None)

        only_employee_id = request.args.get('employee_id', type=int)
        if not admin_user:
            if my_emp_id is None:
                return jsonify({'employees': []}), 200
            only_employee_id = my_emp_id

        query = (
            session.query(
                Employee_Master.employee_id,
                Employee_Master.employee_name,
                func.coalesce(Stage_Master.stage_name, 'Unknown').label('stage_name'),
                func.count(Opportunity_Details.opportunity_id).label('cnt')
            )
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .join(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter(Opportunity_Details.opportunity_owner_employee_id.isnot(None))
            .filter((Opportunity_Details.is_allocated == False) | (Opportunity_Details.is_allocated.is_(None)))
        )

        if only_employee_id:
            query = query.filter(Opportunity_Details.opportunity_owner_employee_id == only_employee_id)

        query = query.group_by(
            Employee_Master.employee_id,
            Employee_Master.employee_name,
            func.coalesce(Stage_Master.stage_name, 'Unknown')
        )
        query = query.having(func.count(Opportunity_Details.opportunity_id) > 0)
        query = query.order_by(Employee_Master.employee_name.asc(), func.count(Opportunity_Details.opportunity_id).desc())

        rows = query.all()

        grouped = {}
        for r in rows:
            eid = r.employee_id
            if eid is None:
                continue
            if eid not in grouped:
                grouped[eid] = {
                    'employee_id': eid,
                    'employee_name': r.employee_name or '—',
                    'total': 0,
                    'by_stage': [],
                }
            c = int(r.cnt or 0)
            grouped[eid]['by_stage'].append({
                'stage_name': r.stage_name or 'Unknown',
                'count': c,
            })
            grouped[eid]['total'] += c

        for v in grouped.values():
            v['by_stage'].sort(key=lambda x: -x['count'])

        employees = sorted(grouped.values(), key=lambda x: -x['total'])

        return jsonify({'employees': employees}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/bulk-delete', methods=['POST'])
@token_required
@tenant_from_jwt
def bulk_delete_leads():
    return crm_controller.bulk_delete_leads()


@crm_bp.route('/leads/table', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_table():
    """
    Get leads table for CRM UI (flat rows: id, name, business_name, contact_person,
    tel_number, mpan_mpr, supplier, annual_usage, start_date, end_date, status,
    assigned_to, callback_parameter, call_summary).

    Authentication:
        - JWT (token must include `tenant_id`)

    Returns:
        200: { success, data, count }
        500: Internal server error
    """
    return crm_controller.get_leads_table()


@crm_bp.route('/leads/import/preview', methods=['POST'])
@token_required
@tenant_from_jwt
def import_leads_preview():
    """
    POST /api/crm/leads/import/preview
    Accepts multipart/form-data with an Excel (.xlsx) or CSV (.csv) file and
    returns a validation preview (no DB writes).

    Notes:
      - Tenant is derived from the authenticated JWT (request.current_user.tenant_id)
      - Frontend must NOT send `X-Tenant-ID`

    Request:
      - file: file to import

    Returns:
      200: preview JSON (see API docs)
      400: invalid request / unsupported file
    """
    return crm_controller.import_leads_preview()


@crm_bp.route('/leads/import/confirm', methods=['POST'])
@token_required
@tenant_from_jwt
def import_leads_confirm():
    """
    POST /api/crm/leads/import/confirm
    Accepts JSON array (validated rows from preview) and inserts Opportunity_Details
    where possible. Partial success allowed; duplicates/skipped rows reported.

    Notes:
      - Tenant is derived from the authenticated JWT (request.current_user.tenant_id)
      - Frontend must NOT send `X-Tenant-ID`

    Request body: [ { row_number, data: {...}, is_valid, errors }, ... ]

    Returns:
      200: { success, inserted, skipped, errors }
      400: invalid request
    """
    return crm_controller.import_leads_confirm()


@crm_bp.route('/leads/recycle-bin', methods=['GET'])
@token_required
@tenant_from_jwt
def get_recycle_bin():
    """
    Get all soft-deleted (Lost) leads for the tenant.

    Query Parameters:
        - None

    Authentication:
        - JWT (token must include `tenant_id`)

    Returns:
        200: { success, data, count } - List of deleted leads with deleted_at timestamp
        500: Internal server error
    """
    return crm_controller.get_recycle_bin()


@crm_bp.route('/leads/cleanup', methods=['PATCH'])
@token_required
@tenant_from_jwt
def delete_expired_lost_leads():
    """
    Permanently delete Lost leads older than N days.
    Admin operation (controlled by token_required + tenant_from_jwt).

    Request Body (optional):
        { "days": 30 }  # Default: 30 days. Records with deleted_at < NOW() - INTERVAL will be permanently removed.

    Authentication:
        - JWT (token must include `tenant_id`)

    Returns:
        200: { success, deleted_count, message }
        500: Internal server error
    """
    return crm_controller.delete_expired_lost_leads()


@crm_bp.route('/leads/import', methods=['POST'])
@token_required
@tenant_from_jwt
def import_leads():
    """
    POST /api/crm/leads/import
    Single-step import: accepts file, validates, and imports in one request.
    """
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id

        # Map service query param to service_id
        service_param = request.args.get('service', 'electricity')
        service_value = service_param.strip().lower() if isinstance(service_param, str) else 'electricity'
        service_id = 2 if service_value == 'water' else 1

        # Check if file is provided
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No file provided',
                'total_rows': 0,
                'successful': 0,
                'failed': 1,
                'errors': ['No file uploaded']
            }), 400

        file = request.files.get('file')

        # IMPORTANT: Before importing, delete existing draft leads for this tenant and service
        # This allows re-uploading the same file without duplicates
        try:
            deleted_drafts = (
                session.query(Opportunity_Details)
                .filter(Opportunity_Details.tenant_id == tenant_id)
                .filter(Opportunity_Details.service_id == service_id)
                .filter(Opportunity_Details.is_draft == True)
                .filter(Opportunity_Details.opportunity_owner_employee_id.is_(None))
                .delete(synchronize_session=False)
            )
            session.commit()
            current_app.logger.info(f'Deleted {deleted_drafts} existing draft leads before import')
        except Exception as cleanup_err:
            current_app.logger.warning(f'Could not clean up drafts before import: {cleanup_err}')
            session.rollback()

        # Step 1: Validate and preview
        preview_result = crm_controller.crm_service.preview_lead_import(tenant_id, file)

        if not preview_result.get('success'):
            return jsonify({
                'success': False,
                'message': preview_result.get('message', 'Validation failed'),
                'total_rows': preview_result.get('total_rows', 0),
                'successful': 0,
                'failed': preview_result.get('total_rows', 1),
                'errors': preview_result.get('errors', ['Validation failed'])
            }), 400

        # If no valid rows, return early
        valid_rows = preview_result.get('valid_rows', 0)
        if valid_rows == 0:
            return jsonify({
                'success': False,
                'message': 'No valid rows to import',
                'total_rows': preview_result.get('total_rows', 0),
                'successful': 0,
                'failed': preview_result.get('invalid_rows', 0),
                'errors': preview_result.get('errors', ['No valid data found'])
            }), 400

        # Step 2: Import the validated rows
        all_rows = preview_result.get('rows', [])
        validated_data = [row['data'] for row in all_rows if row.get('is_valid', False)]

        created_by = getattr(request.current_user, 'id', None)

        confirm_result = crm_controller.crm_service.confirm_lead_import(tenant_id, validated_data, created_by, service_id)

        # Check if confirm returned an error
        if 'success' in confirm_result and not confirm_result['success']:
            return jsonify({
                'success': False,
                'message': confirm_result.get('message', 'Import failed'),
                'total_rows': preview_result.get('total_rows', 0),
                'successful': 0,
                'failed': preview_result.get('total_rows', 0),
                'errors': [confirm_result.get('error', 'Import failed')]
            }), 400

        # Format response
        inserted = confirm_result.get('inserted', 0)
        skipped = confirm_result.get('skipped', 0)

        return jsonify({
            'success': inserted > 0 or skipped == preview_result.get('total_rows', 0),
            'message': f"Successfully imported {inserted} lead(s)" if inserted > 0 else "No new leads imported",
            'total_rows': preview_result.get('total_rows', 0),
            'successful': inserted,
            'failed': skipped,
            'errors': confirm_result.get('errors', [])
        }), 200

    except Exception as e:
        session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'total_rows': 0,
            'successful': 0,
            'failed': 1,
            'errors': [str(e)]
        }), 500
    finally:
        session.close()


@crm_bp.route('/leads/import/template', methods=['GET'])
def download_leads_template():
    """
    GET /api/crm/leads/import/template
    Downloads an Excel template for lead imports.
    No authentication required for template download.

    Returns:
      200: Excel file with headers and example data
    """
    try:
        from flask import send_file
        from openpyxl import Workbook
        from io import BytesIO

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Import"

        # Headers (match the expected column names from validation)
        headers = [
            'Business Name', 'Contact Person', 'Tel Number', 'Email',
            'MPAN_MPR', 'Start Date', 'End Date', 'Annual Usage',
            'Address', 'Site Address'
        ]
        ws.append(headers)

        # Example data
        ws.append([
            'Acme Corp', 'John Doe', '0207123456', 'john@acme.com',
            '1234567890123', '2024-01-01', '2025-01-01', '50000',
            '123 Main St, London', '456 Business Park, London'
        ])
        ws.append([
            'Tech Solutions Ltd', 'Jane Smith', '0207987654', 'jane@techsolutions.co.uk',
            '9876543210987', '2024-06-01', '2025-06-01', '75000',
            '789 Tech Ave, Manchester', '789 Tech Ave, Manchester'
        ])

        # Style headers
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='leads_import_template.xlsx'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@crm_bp.route('/leads/customer-type', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_by_customer_type():
    """
    Get leads filtered by customer type (NEW/EXISTING)

    Query Parameters:
        - type: 'NEW' or 'EXISTING' (optional, returns all if not specified)
        - stage_id: Filter by stage
        - lead_status: Filter by lead status
        - assigned_employee_id: Filter by assigned employee

    Notes:
        - Tenant is derived from the authenticated JWT (request.current_user.tenant_id)
        - Do not send `X-Tenant-ID` for Leads endpoints

    Returns:
        200: List of leads with customer_type classification
        500: Internal server error
    """
    return crm_controller.get_leads_by_customer_type()

@crm_bp.route('/clients', methods=['POST'])
@require_tenant
def create_client():
    """
    Create a new client (Client_Master). Automatically inserts one record
    in Opportunity_Details so the client appears as a lead.

    Request Body:
        - client_company_name or business_name (required)
        - client_contact_name, client_phone, client_email, address, etc. (optional)

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        201: { success, data: { client, opportunity }, message }
        400: Validation error or missing body
        500: Internal server error
    """
    return crm_controller.create_client()


@crm_bp.route('/clients/<int:client_id>/call-summary', methods=['POST'])
@require_tenant
def create_call_summary(client_id):
    """
    Create a call summary/interaction record for a client

    Path Parameters:
        - client_id: Client identifier

    Request Body:
        - call_status: Call status (Phone, Email, Meeting, Other)
        - call_result: Result of the call
        - remarks: Additional remarks/notes
        - next_follow_up_date: Next follow-up date (YYYY-MM-DD)

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        201: Call summary created successfully
        400: Invalid request data
        500: Internal server error
    """
    return crm_controller.create_call_summary(client_id)

@crm_bp.route('/clients/<int:client_id>/upload', methods=['POST'])
@require_tenant
def client_upload_document(client_id):
    """
    POST /api/crm/clients/<client_id>/upload
    Upload a document for a specific client

    Path Parameters:
        - client_id: Client identifier

    Form Data:
        - file: Document file (required)
        - document_name: Name of the document (optional)
        - document_description: Description (optional)
        - category: Document category (default: CLIENT_UPLOAD)

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        201: Document uploaded successfully
        400: Invalid file or data
        500: Internal server error
    """
    return crm_controller.client_upload_document(client_id)


# ========================================
# PROJECT ROUTES
# ========================================

@crm_bp.route('/projects', methods=['GET'])
@require_tenant
def get_projects():
    """
    Get all projects for the current tenant

    Query Parameters:
        - status: Filter by project status
        - project_manager_id: Filter by project manager

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        200: List of projects with statistics
        500: Internal server error
    """
    return crm_controller.get_projects()


@crm_bp.route('/projects/<int:project_id>', methods=['GET'])
@require_tenant
def get_project_detail(project_id):
    """
    Get details of a specific project

    Path Parameters:
        - project_id: Project identifier

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        200: Project details
        404: Project not found
        500: Internal server error
    """
    return crm_controller.get_project_detail(project_id)


# ========================================
# DEAL/CONTRACT ROUTES
# ========================================

@crm_bp.route('/deals', methods=['GET'])
@require_tenant
def get_deals():
    """
    Get all deals/contracts for the current tenant

    Query Parameters:
        - status: Filter by contract status
        - contract_owner_id: Filter by owner

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        200: List of deals with statistics
        500: Internal server error
    """
    return crm_controller.get_deals()


@crm_bp.route('/deals/<int:contract_id>', methods=['GET'])
@require_tenant
def get_deal_detail(contract_id):
    """
    Get details of a specific deal

    Path Parameters:
        - contract_id: Contract identifier

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        200: Deal details
        404: Deal not found
        500: Internal server error
    """
    return crm_controller.get_deal_detail(contract_id)


# ========================================
# USER ROUTES
# ========================================

@crm_bp.route('/users', methods=['GET'])
@require_tenant
def get_users():
    """
    Get all users for the current tenant

    Query Parameters:
        - active_only: Filter active users only (default: true)

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        200: List of users
        500: Internal server error
    """
    return crm_controller.get_users()


@crm_bp.route('/employees', methods=['GET'])
@token_required
@tenant_from_jwt
def get_employees():
    """
    GET /api/crm/employees
    Get all employees for assignment dropdowns
    """
    return crm_controller.get_employees()


# ========================================
# SUPPORTING DATA ROUTES
# ========================================

@crm_bp.route('/roles', methods=['GET'])
def get_roles():
    """
    Get all roles (system + tenant-specific)

    Returns:
        200: List of roles
        500: Internal server error
    """
    return crm_controller.get_roles()


@crm_bp.route('/stages', methods=['GET'], strict_slashes=False)
@token_required
@tenant_from_jwt
def get_stages():
    """
    Get all pipeline stages

    Query Parameters:
        - pipeline_type: Filter by pipeline type (lead, sales, training)

    Returns:
        200: List of stages with stage_id and stage_name
        500: Internal server error
    """
    return crm_controller.get_stages()


@crm_bp.route('/services', methods=['GET'])
def get_services():
    """
    Get all services

    Returns:
        200: List of services
        500: Internal server error
    """
    return crm_controller.get_services()


@crm_bp.route('/suppliers', methods=['GET'])
@require_tenant
def get_suppliers():
    """
    Get all suppliers for the current tenant

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        200: List of suppliers
        500: Internal server error
    """
    return crm_controller.get_suppliers()


@crm_bp.route('/interactions', methods=['GET'])
@require_tenant
def get_interactions():
    """
    Get all client interactions for the current tenant

    Query Parameters:
        - client_id: Filter by client
        - interaction_type: Filter by interaction type
        - user_id: Filter by user

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        200: List of interactions
        500: Internal server error
    """
    return crm_controller.get_interactions()


# ========================================
# DASHBOARD ROUTE
# ========================================

@crm_bp.route('/dashboard', methods=['GET'])
@require_tenant
def get_dashboard():
    """
    Get CRM dashboard summary with key metrics

    Headers:
        - X-Tenant-ID: Tenant identifier (required)

    Returns:
        200: Dashboard metrics (leads, projects, deals statistics)
        500: Internal server error
    """
    return crm_controller.get_dashboard()


# ========================================
# HEALTH CHECK
# ========================================

@crm_bp.route('/health', methods=['GET'])
def health_check():
    """
    Health check endpoint for CRM module

    Returns:
        200: CRM module is operational
    """
    return {
        'success': True,
        'module': 'CRM',
        'status': 'operational',
        'message': 'StreemLyne CRM module is running'
    }, 200


@crm_bp.route('/debug/tenant/<int:tenant_id>', methods=['GET'])
def debug_tenant_lookup(tenant_id):
    """Debug endpoint to test tenant lookup directly (NO middleware)"""
    try:
        from backend.crm.repositories.tenant_repository import TenantRepository
        repo = TenantRepository()
        tenant = repo.get_tenant_by_id(tenant_id)

        return {
            'success': True if tenant else False,
            'tenant_id_requested': tenant_id,
            'tenant_found': tenant is not None,
            'tenant_data': tenant,
            'message': 'Direct lookup (no middleware)'
        }, 200 if tenant else 404
    except Exception as e:
        import traceback
        return {
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }, 500

@crm_bp.route('/priced', methods=['GET'])
@token_required
@tenant_from_jwt
def get_priced():
    """
    GET /api/crm/priced
    Get all priced leads and renewals

    Returns leads where stage_id = 8 (Priced stage)
    Returns renewals where Misc_Col1 = 'priced'

    Authentication:
        - JWT (token must include `tenant_id`)

    Returns:
        200: {
            success: true,
            leads: [...],
            renewals: [...],
            total_leads: int,
            total_renewals: int,
            total: int
        }
        500: Internal server error
    """
    return crm_controller.get_priced()


def _priced_stage_id(session, stage_name: str):
    row = session.execute(text("""
        SELECT stage_id
        FROM "StreemLyne_MT"."Stage_Master"
        WHERE LOWER(stage_name) = LOWER(:stage_name)
        LIMIT 1
    """), {'stage_name': stage_name}).mappings().first()
    return row['stage_id'] if row else None


def _priced_lead(session, opportunity_id: int, tenant_id):
    return session.execute(text("""
        SELECT od.*, COALESCE(od.business_name, od.opportunity_title) AS resolved_business_name
        FROM "StreemLyne_MT"."Opportunity_Details" od
        LEFT JOIN "StreemLyne_MT"."Stage_Master" sm ON od.stage_id = sm.stage_id
        LEFT JOIN "StreemLyne_MT"."Client_Master" cm ON od.client_id = cm.client_id
        WHERE od.tenant_id = :tenant_id
          AND (od.opportunity_id = :id OR od.tenant_lead_id = :id)
          AND LOWER(COALESCE(sm.stage_name, '')) = 'priced'
          AND (cm.is_deleted IS NULL OR cm.is_deleted = FALSE)
        ORDER BY CASE WHEN od.opportunity_id = :id THEN 0 ELSE 1 END
        LIMIT 1
    """), {'tenant_id': tenant_id, 'id': opportunity_id}).mappings().first()


def _insert_priced_interaction(session, client_id: int, status: str, notes: str):
    session.execute(text("""
        INSERT INTO "StreemLyne_MT"."Client_Interactions"
            (client_id, contact_date, contact_method, notes, next_steps, created_at)
        VALUES (:client_id, CURRENT_DATE, 1, :notes, :status, :created_at)
    """), {
        'client_id': client_id,
        'notes': notes,
        'status': status,
        'created_at': datetime.utcnow(),
    })


@crm_bp.route('/priced/leads/<int:opportunity_id>/accept', methods=['POST', 'OPTIONS'])
@token_required
@tenant_from_jwt
def accept_priced_lead(opportunity_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        lead = _priced_lead(session, opportunity_id, tenant_id)
        if not lead:
            return jsonify({'error': 'Priced lead not found'}), 404

        real_id = lead['opportunity_id']
        assigned_employee_id = lead.get('opportunity_owner_employee_id')
        client_id = lead.get('client_id')
        if not client_id:
            client = Client_Master(
                tenant_id=int(tenant_id),
                assigned_employee_id=assigned_employee_id,
                client_company_name=lead.get('resolved_business_name') or '[IMPORTED LEADS]',
                client_contact_name=lead.get('contact_person') or '',
                client_phone=lead.get('tel_number') or '',
                client_email=lead.get('email') or '',
                default_currency_id=1,
                created_at=datetime.utcnow()
            )
            session.add(client)
            session.flush()
            client_id = client.client_id

        business_name = lead.get('resolved_business_name') or 'Unknown'
        now = datetime.utcnow()

        session.execute(text("""
            UPDATE "StreemLyne_MT"."Client_Master"
            SET assigned_employee_id = :assigned_employee_id,
                client_company_name = COALESCE(NULLIF(:business_name, ''), client_company_name),
                client_contact_name = COALESCE(NULLIF(:contact_person, ''), client_contact_name),
                client_phone = COALESCE(NULLIF(:tel_number, ''), client_phone),
                client_mobile = COALESCE(NULLIF(:mobile_no, ''), client_mobile),
                client_email = COALESCE(NULLIF(:email, ''), client_email),
                address = COALESCE(NULLIF(:address, ''), address),
                post_code = COALESCE(NULLIF(:postcode, ''), post_code),
                is_deleted = FALSE,
                deleted_at = NULL,
                deleted_reason = NULL,
                is_archived = FALSE,
                is_allocated = FALSE
            WHERE client_id = :client_id AND tenant_id = :tenant_id
        """), {
            'assigned_employee_id': assigned_employee_id,
            'business_name': business_name,
            'contact_person': lead.get('contact_person') or '',
            'tel_number': lead.get('tel_number') or '',
            'mobile_no': lead.get('mobile_no') or '',
            'email': lead.get('email') or '',
            'address': lead.get('address') or '',
            'postcode': lead.get('postcode') or '',
            'client_id': client_id,
            'tenant_id': tenant_id,
        })

        project = session.execute(text("""
            SELECT project_id
            FROM "StreemLyne_MT"."Project_Details"
            WHERE opportunity_id = :opportunity_id OR client_id = :client_id
            ORDER BY CASE WHEN opportunity_id = :opportunity_id THEN 0 ELSE 1 END
            LIMIT 1
        """), {'opportunity_id': real_id, 'client_id': client_id}).mappings().first()

        if project:
            project_id = project['project_id']
            session.execute(text("""
                UPDATE "StreemLyne_MT"."Project_Details"
                SET client_id = :client_id,
                    opportunity_id = :opportunity_id,
                    project_title = COALESCE(NULLIF(:business_name, ''), project_title),
                    start_date = COALESCE(:start_date, start_date),
                    end_date = COALESCE(:end_date, end_date),
                    address = COALESCE(NULLIF(:address, ''), address),
                    "Misc_Col2" = COALESCE(:annual_usage, "Misc_Col2"),
                    assigned_employee_id = :assigned_employee_id,
                    status = NULL,
                    updated_at = :now
                WHERE project_id = :project_id
            """), {
                'client_id': client_id,
                'opportunity_id': real_id,
                'business_name': business_name,
                'start_date': lead.get('start_date'),
                'end_date': lead.get('end_date'),
                'address': lead.get('address') or '',
                'annual_usage': lead.get('annual_usage'),
                'assigned_employee_id': assigned_employee_id,
                'now': now,
                'project_id': project_id,
            })
        else:
            project_id = session.execute(text("""
                INSERT INTO "StreemLyne_MT"."Project_Details"
                    (client_id, opportunity_id, project_title, project_description, start_date, end_date,
                     employee_id, created_at, updated_at, address, "Misc_Col2", assigned_employee_id, status)
                VALUES
                    (:client_id, :opportunity_id, :business_name, 'Converted priced lead',
                     :start_date, :end_date, :employee_id, :now, :now, :address, :annual_usage,
                     :assigned_employee_id, NULL)
                RETURNING project_id
            """), {
                'client_id': client_id,
                'opportunity_id': real_id,
                'business_name': business_name,
                'start_date': lead.get('start_date'),
                'end_date': lead.get('end_date'),
                'employee_id': getattr(request.current_user, 'employee_id', None),
                'now': now,
                'address': lead.get('address') or '',
                'annual_usage': lead.get('annual_usage'),
                'assigned_employee_id': assigned_employee_id,
            }).scalar()

        contract_params = {
            'project_id': project_id,
            'employee_id': getattr(request.current_user, 'employee_id', None),
            'supplier_id': lead.get('supplier_id'),
            'contract_start_date': lead.get('start_date'),
            'contract_end_date': lead.get('end_date'),
            'service_id': lead.get('service_id') or 1,
            'unit_rate': lead.get('rate_1') or 0,
            'currency_id': lead.get('currency_id') or 1,
            'now': now,
            'mpan_number': lead.get('mpan_mpr') or '',
            'mpan_bottom': lead.get('mpan_bottom') or '',
            'standing_charge': str(lead.get('stand_charge')) if lead.get('stand_charge') is not None else None,
            'rate_1': lead.get('rate_1'),
            'rate_2': lead.get('rate_2'),
            'rate_3': lead.get('rate_3'),
            'net_notch': lead.get('net_notch'),
            'payment_type': lead.get('payment_type'),
        }

        contract = session.execute(text("""
            SELECT energy_contract_master_id
            FROM "StreemLyne_MT"."Energy_Contract_Master"
            WHERE project_id = :project_id
            LIMIT 1
        """), {'project_id': project_id}).mappings().first()

        if contract:
            session.execute(text("""
                UPDATE "StreemLyne_MT"."Energy_Contract_Master"
                SET supplier_id = COALESCE(:supplier_id, supplier_id),
                    contract_start_date = COALESCE(:contract_start_date, contract_start_date),
                    contract_end_date = COALESCE(:contract_end_date, contract_end_date),
                    service_id = COALESCE(:service_id, service_id),
                    unit_rate = COALESCE(:unit_rate, unit_rate),
                    currency_id = COALESCE(:currency_id, currency_id),
                    updated_at = :now,
                    mpan_number = COALESCE(NULLIF(:mpan_number, ''), mpan_number),
                    mpan_bottom = COALESCE(NULLIF(:mpan_bottom, ''), mpan_bottom),
                    standing_charge = COALESCE(:standing_charge, standing_charge),
                    rate_1 = COALESCE(:rate_1, rate_1),
                    rate_2 = COALESCE(:rate_2, rate_2),
                    rate_3 = COALESCE(:rate_3, rate_3),
                    net_notch = COALESCE(:net_notch, net_notch),
                    payment_type = COALESCE(:payment_type, payment_type)
                WHERE energy_contract_master_id = :contract_id
            """), {**contract_params, 'contract_id': contract['energy_contract_master_id']})
        else:
            session.execute(text("""
                INSERT INTO "StreemLyne_MT"."Energy_Contract_Master"
                    (project_id, employee_id, supplier_id, contract_start_date, contract_end_date,
                     terms_of_sale, service_id, unit_rate, currency_id, created_at, updated_at,
                     mpan_number, mpan_bottom, standing_charge, rate_1, rate_2, rate_3, net_notch, payment_type)
                VALUES
                    (:project_id, :employee_id, :supplier_id, :contract_start_date, :contract_end_date,
                     '', :service_id, :unit_rate, :currency_id, :now, :now,
                     :mpan_number, :mpan_bottom, :standing_charge, :rate_1, :rate_2, :rate_3, :net_notch, :payment_type)
            """), contract_params)

        converted_stage_id = _priced_stage_id(session, 'Converted') or _priced_stage_id(session, 'Won')
        if converted_stage_id:
            session.execute(text("""
                UPDATE "StreemLyne_MT"."Opportunity_Details"
                SET stage_id = :stage_id, client_id = :client_id
                WHERE opportunity_id = :opportunity_id AND tenant_id = :tenant_id
            """), {
                'stage_id': converted_stage_id,
                'client_id': client_id,
                'opportunity_id': real_id,
                'tenant_id': tenant_id,
            })

        _insert_priced_interaction(session, client_id, 'Priced Accepted', '[Priced Accepted] Customer onboarded from priced leads')
        session.commit()
        return jsonify({'success': True, 'message': 'Lead moved to renewals', 'client_id': client_id}), 200
    except Exception as e:
        session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/priced/leads/<int:opportunity_id>/reject', methods=['POST', 'OPTIONS'])
@token_required
@tenant_from_jwt
def reject_priced_lead(opportunity_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        lead = _priced_lead(session, opportunity_id, tenant_id)
        if not lead:
            return jsonify({'error': 'Priced lead not found'}), 404

        real_id = lead['opportunity_id']
        client_id = lead.get('client_id')
        if not client_id:
            client = Client_Master(
                tenant_id=int(tenant_id),
                assigned_employee_id=lead.get('opportunity_owner_employee_id'),
                client_company_name=lead.get('resolved_business_name') or '[IMPORTED LEADS]',
                client_contact_name=lead.get('contact_person') or '',
                client_phone=lead.get('tel_number') or '',
                client_email=lead.get('email') or '',
                default_currency_id=1,
                created_at=datetime.utcnow()
            )
            session.add(client)
            session.flush()
            client_id = client.client_id

        lost_stage_id = _priced_stage_id(session, 'Lost')
        session.execute(text("""
            UPDATE "StreemLyne_MT"."Client_Master"
            SET is_deleted = TRUE, deleted_at = :now, deleted_reason = 'Lost'
            WHERE client_id = :client_id AND tenant_id = :tenant_id
        """), {'now': datetime.utcnow(), 'client_id': client_id, 'tenant_id': tenant_id})

        if lost_stage_id:
            session.execute(text("""
                UPDATE "StreemLyne_MT"."Opportunity_Details"
                SET stage_id = :stage_id, client_id = :client_id
                WHERE opportunity_id = :opportunity_id AND tenant_id = :tenant_id
            """), {
                'stage_id': lost_stage_id,
                'client_id': client_id,
                'opportunity_id': real_id,
                'tenant_id': tenant_id,
            })

        _insert_priced_interaction(session, client_id, 'Lost', '[Lost] Rejected from priced leads')
        session.commit()
        return jsonify({'success': True, 'message': 'Lead moved to lost'}), 200
    except Exception as e:
        session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/priced/renewals/<int:client_id>/accept', methods=['POST', 'OPTIONS'])
@token_required
@tenant_from_jwt
def accept_priced_renewal(client_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        row = session.execute(text("""
            SELECT cm.client_id, pd.project_id, pd.assigned_employee_id
            FROM "StreemLyne_MT"."Client_Master" cm
            JOIN "StreemLyne_MT"."Project_Details" pd ON cm.client_id = pd.client_id
            WHERE cm.client_id = :client_id
              AND cm.tenant_id = :tenant_id
              AND LOWER(COALESCE(pd.status, '')) = 'priced'
            LIMIT 1
        """), {'client_id': client_id, 'tenant_id': tenant_id}).mappings().first()
        if not row:
            return jsonify({'error': 'Priced renewal not found'}), 404

        session.execute(text("""
            UPDATE "StreemLyne_MT"."Client_Master"
            SET is_deleted = FALSE,
                deleted_at = NULL,
                deleted_reason = NULL,
                is_archived = FALSE,
                is_allocated = FALSE,
                assigned_employee_id = :assigned_employee_id
            WHERE client_id = :client_id AND tenant_id = :tenant_id
        """), {
            'assigned_employee_id': row['assigned_employee_id'],
            'client_id': client_id,
            'tenant_id': tenant_id,
        })
        session.execute(text("""
            UPDATE "StreemLyne_MT"."Project_Details"
            SET status = NULL, updated_at = :now
            WHERE project_id = :project_id
        """), {'now': datetime.utcnow(), 'project_id': row['project_id']})
        _insert_priced_interaction(session, client_id, 'Priced Accepted', '[Priced Accepted] Customer onboarded from priced renewals')
        session.commit()
        return jsonify({'success': True, 'message': 'Renewal moved to renewals'}), 200
    except Exception as e:
        session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/priced/renewals/<int:client_id>/reject', methods=['POST', 'OPTIONS'])
@token_required
@tenant_from_jwt
def reject_priced_renewal(client_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        row = session.execute(text("""
            SELECT cm.client_id, pd.project_id
            FROM "StreemLyne_MT"."Client_Master" cm
            JOIN "StreemLyne_MT"."Project_Details" pd ON cm.client_id = pd.client_id
            WHERE cm.client_id = :client_id
              AND cm.tenant_id = :tenant_id
              AND LOWER(COALESCE(pd.status, '')) = 'priced'
            LIMIT 1
        """), {'client_id': client_id, 'tenant_id': tenant_id}).mappings().first()
        if not row:
            return jsonify({'error': 'Priced renewal not found'}), 404

        session.execute(text("""
            UPDATE "StreemLyne_MT"."Client_Master"
            SET is_deleted = TRUE, deleted_at = :now, deleted_reason = 'Lost'
            WHERE client_id = :client_id AND tenant_id = :tenant_id
        """), {'now': datetime.utcnow(), 'client_id': client_id, 'tenant_id': tenant_id})
        session.execute(text("""
            UPDATE "StreemLyne_MT"."Project_Details"
            SET status = 'Lost', updated_at = :now
            WHERE project_id = :project_id
        """), {'now': datetime.utcnow(), 'project_id': row['project_id']})
        _insert_priced_interaction(session, client_id, 'Lost', '[Lost] Rejected from priced renewals')
        session.commit()
        return jsonify({'success': True, 'message': 'Renewal moved to lost'}), 200
    except Exception as e:
        session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/cleansing', methods=['GET'])
@token_required
@tenant_from_jwt
def get_cleansing():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        records = []

        # 1. CRM Leads
        lead_rows = (
            session.query(
                Opportunity_Details,
                Stage_Master.stage_name.label('cleansing_reason'),
                Supplier_Master.supplier_company_name.label('supplier_name'),
                Employee_Master.employee_name.label('assigned_to_name'),
                func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name')
            )
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
            .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .filter(Opportunity_Details.tenant_id == tenant_id)
            .filter(Stage_Master.stage_name.in_(['Invalid Number', 'Incorrect Supplier']))
            .order_by(Opportunity_Details.created_at.desc())
            .all()
        )

        for row in lead_rows:
            od = row[0]
            records.append({
                'id': od.opportunity_id,
                'client_id': od.opportunity_id,
                'display_id': od.tenant_lead_id,
                'display_order': od.tenant_lead_id,
                'business_name': row.business_name or 'Unknown',
                'contact_person': od.contact_person,
                'phone': od.tel_number,
                'mobile_no': od.mobile_no,
                'mpan_mpr': od.mpan_mpr,
                'mpan_top': od.mpan_mpr,
                'supplier_id': od.supplier_id,
                'supplier_name': row.supplier_name,
                'annual_usage': od.annual_usage,
                'start_date': _iso(od.start_date),
                'end_date': _iso(od.end_date),
                'cleansing_reason': row.cleansing_reason,
                'flagged_at': _iso(od.created_at),
                'notes': od.notes,
                'assigned_to_id': od.opportunity_owner_employee_id,
                'assigned_to_name': row.assigned_to_name,
                'source': 'lead',
            })

        # 2. Energy Clients (already using SessionLocal in your code)
        try:
            from backend.models import Energy_Contract_Master, Project_Details

            client_rows = (
                session.query(Client_Master, Energy_Contract_Master, Supplier_Master)
                .outerjoin(Project_Details, Client_Master.client_id == Project_Details.client_id)
                .outerjoin(Energy_Contract_Master, Project_Details.project_id == Energy_Contract_Master.project_id)
                .outerjoin(Supplier_Master, Energy_Contract_Master.supplier_id == Supplier_Master.supplier_id)
                .filter(
                    Client_Master.tenant_id == tenant_id,
                    Client_Master.is_deleted == True,
                    Client_Master.deleted_reason.in_(['Invalid Number', 'Incorrect Supplier']),
                )
                .all()
            )

            for client, contract, supplier in client_rows:
                records.append({
                    'id': client.client_id,
                    'client_id': client.client_id,
                    'display_id': getattr(client, 'display_id', None),
                    'display_order': getattr(client, 'display_order', None),
                    'business_name': getattr(client, 'client_company_name', None) or 'Unknown',
                    'contact_person': getattr(client, 'client_contact_name', None),
                    'phone': getattr(client, 'client_phone', None),
                    'mobile_no': getattr(client, 'mobile_no', None),
                    'mpan_mpr': getattr(contract, 'mpan_mpr', None) if contract else None,
                    'mpan_top': getattr(contract, 'mpan_top', None) if contract else None,
                    'supplier_id': contract.supplier_id if contract else None,
                    'supplier_name': supplier.supplier_company_name if supplier else None,
                    'annual_usage': getattr(contract, 'annual_usage', None) if contract else None,
                    'start_date': contract.contract_start_date.isoformat() if contract and contract.contract_start_date else None,
                    'end_date': contract.contract_end_date.isoformat() if contract and contract.contract_end_date else None,
                    'cleansing_reason': client.deleted_reason,
                    'flagged_at': client.deleted_at.isoformat() if client.deleted_at else None,
                    'notes': getattr(client, 'deleted_notes', None),
                    'assigned_to_id': getattr(client, 'assigned_to_id', None),
                    'assigned_to_name': None,
                    'source': 'energy_client',
                })
        except Exception as ec_err:
            import logging
            logging.getLogger(__name__).warning('Could not load energy clients for cleansing: %s', ec_err)

        records.sort(key=lambda x: x.get('flagged_at') or '', reverse=True)

        return jsonify({'records': records, 'total': len(records)}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/<int:opportunity_id>/callback', methods=['POST', 'OPTIONS'])
@token_required
@tenant_from_jwt
def leads_callback(opportunity_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session = SessionLocal()

    try:
        tenant_id = g.tenant_id
        data = request.get_json()

        status = data.get('status')
        callback_date = data.get('callback_date')
        notes = data.get('notes', '')

        current_app.logger.info(
            f'📥 Leads callback: opportunity_id={opportunity_id}, status={status}, '
            f'callback_date={callback_date}, tenant_id={tenant_id}'
        )

        if not status:
            return jsonify({'error': 'Status is required'}), 400

        # ✅ Resolve lead — URL param may be opportunity_id OR tenant_lead_id
        lead = session.query(Opportunity_Details).filter(
            Opportunity_Details.tenant_id == tenant_id,
            Opportunity_Details.opportunity_id == opportunity_id
        ).first()

        if not lead:
            lead = session.query(Opportunity_Details).filter(
                Opportunity_Details.tenant_id == tenant_id,
                Opportunity_Details.tenant_lead_id == opportunity_id
            ).first()

        if not lead:
            current_app.logger.error(
                f'❌ Lead not found: opportunity_id param={opportunity_id}, tenant={tenant_id}'
            )
            return jsonify({'error': 'Lead not found'}), 404

        real_id = lead.opportunity_id
        lead_client_id = lead.client_id

        current_app.logger.info(
            f'✅ Resolved lead: real_opportunity_id={real_id}, client_id={lead_client_id}'
        )

        # ✅ Capture old stage name for history
        old_stage = None
        if lead.stage_id:
            old_stage_row = session.query(Stage_Master).filter_by(
                stage_id=lead.stage_id
            ).first()
            old_stage = old_stage_row.stage_name if old_stage_row else None

        # ✅ Always resolve stage_id from DB — never trust frontend
        stage = session.query(Stage_Master).filter(
            Stage_Master.stage_name == status
        ).first()

        if not stage:
            return jsonify({'error': f'Stage not found: {status}'}), 400

        stage_id = stage.stage_id

        # ✅ If no client_id, create one
        if not lead_client_id:
            current_app.logger.warning(f'Lead {real_id} has no client_id, creating one...')
            new_client = Client_Master(
                tenant_id=tenant_id,  # ✅ keep original type, don't force int
                assigned_employee_id=lead.opportunity_owner_employee_id,
                client_company_name=lead.business_name or lead.opportunity_title or '[IMPORTED LEADS]',
                client_contact_name=lead.contact_person or '',
                client_phone=str(lead.tel_number).replace('.0', '') if lead.tel_number else '',
                client_mobile=lead.mobile_no or '',
                client_email=lead.email or '',
                default_currency_id=1,
                created_at=datetime.utcnow()
            )
            session.add(new_client)
            session.flush()
            lead_client_id = new_client.client_id
            lead.client_id = lead_client_id
            session.flush()
            current_app.logger.info(f'✅ Created client_id={lead_client_id} for lead {real_id}')

        # ✅ Update stage on lead
        lead.stage_id = stage_id
        session.flush()
        current_app.logger.info(f'Updated lead {real_id} stage: {old_stage} → {status}')

        # ✅ Build history note with status transition
        transition = f"Status: {old_stage or 'None'} → {status}"
        formatted_notes = f"[{status}] {transition}" + (f" | {notes}" if notes else "")

        reminder_date = None
        if callback_date:
            try:
                reminder_date = datetime.strptime(callback_date, '%Y-%m-%d').date()
            except ValueError:
                current_app.logger.warning(f'Invalid callback_date format: {callback_date}')

        # ✅ Write history BEFORE any early returns
        interaction = Client_Interactions(
            client_id=lead_client_id,
            contact_date=datetime.utcnow().date(),
            contact_method=1,
            reminder_date=reminder_date,
            notes=formatted_notes,
            next_steps=status,
            created_at=datetime.utcnow()
        )
        session.add(interaction)
        session.flush()
        current_app.logger.info(
            f'✅ History logged: interaction_id={interaction.interaction_id}'
        )

        CLEANSING_STATUSES = {'Invalid Number', 'Incorrect Supplier'}
        RECYCLE_BIN_STATUSES = {'Lost', 'Lost COT', 'Meter De-energised', 'Complaint', 'Duplicate'}

        if status in CLEANSING_STATUSES:
            client = session.query(Client_Master).filter(
                Client_Master.client_id == lead_client_id,
                Client_Master.tenant_id == tenant_id
            ).first()
            if client:
                client.is_deleted = True
                client.deleted_at = datetime.utcnow()
                client.deleted_reason = status
                if hasattr(client, 'is_cleansing'):
                    client.is_cleansing = True
                session.flush()

            session.commit()
            return jsonify({
                'success': True,
                'message': f'Moved to Cleansing ({status})',
                'moved_to_cleansing': True,
                'interaction_id': interaction.interaction_id,
            }), 200

        if status in RECYCLE_BIN_STATUSES:
            client = session.query(Client_Master).filter(
                Client_Master.client_id == lead_client_id,
                Client_Master.tenant_id == tenant_id
            ).first()
            if client:
                client.is_deleted = True
                client.deleted_at = datetime.utcnow()
                client.deleted_reason = status
                if hasattr(client, 'is_cleansing'):
                    client.is_cleansing = False
                session.flush()

            session.commit()
            return jsonify({
                'success': True,
                'message': f'Moved to recycle bin ({status})',
                'moved_to_recycle_bin': True,
                'interaction_id': interaction.interaction_id,
            }), 200

        session.commit()
        current_app.logger.info(
            f'✅ Callback saved for lead {real_id}, status: {status}'
        )

        return jsonify({
            'success': True,
            'message': 'Callback saved successfully',
            'status': status,
            'stage_id': stage_id,
            'callback_date': callback_date,
            'interaction_id': interaction.interaction_id,
        }), 200

    except Exception as e:
        session.rollback()
        current_app.logger.exception(f"❌ Error in leads_callback: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/allocated', methods=['GET'])
@token_required
@tenant_from_jwt
def get_allocated_leads():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        current_user = request.current_user
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1

        employee_id = getattr(current_user, 'employee_id', None)
        role_name = getattr(current_user, 'role', None)
        admin_user = is_crm_leads_admin_role(role_name)

        import logging
        logging.getLogger(__name__).warning(
            '🔍 get_allocated_leads: employee_id=%s is_admin=%s tenant=%s service=%s',
            employee_id, admin_user, tenant_id, service_param
        )

        if not employee_id:
            return jsonify([]), 200

        query = (
            session.query(
                Opportunity_Details,
                Stage_Master.stage_name,
                Employee_Master.employee_name.label('assigned_to_name'),
                func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name'),
                Supplier_Master.supplier_company_name.label('supplier_name')
            )
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
            .outerjoin(Client_Master, Opportunity_Details.client_id == Client_Master.client_id)
            .filter(
                (Opportunity_Details.tenant_id == tenant_id) |
                ((Opportunity_Details.client_id.isnot(None)) & (Client_Master.tenant_id == tenant_id))
            )
            .filter(Opportunity_Details.service_id == service_id)
            .filter(Opportunity_Details.is_allocated == True)
        )

        if not admin_user:
            query = query.filter(Opportunity_Details.opportunity_owner_employee_id == employee_id)

        query = query.order_by(Opportunity_Details.created_at.desc())
        rows = query.all()

        results = []
        for row in rows:
            od = row[0]
            result_dict = {k: _serial(getattr(od, k)) for k in od.__table__.columns.keys()}
            result_dict.update({
                'stage_name': row.stage_name,
                'assigned_to_name': row.assigned_to_name,
                'business_name': row.business_name,
                'supplier_name': row.supplier_name,
            })
            results.append(result_dict)

        logging.getLogger(__name__).warning(
            '✅ get_allocated_leads returning %d leads for employee_id=%s',
            len(results), employee_id
        )

        return jsonify(results), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()

@crm_bp.route('/leads/archives', methods=['GET'])
@token_required
@tenant_from_jwt
def get_archived_leads():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        current_user = request.current_user
        service_param = request.args.get('service', 'utilities')
        service_id = 2 if service_param.strip().lower() == 'water' else 1

        from backend.crm.utils.role_helpers import is_admin_user
        is_admin = is_admin_user(current_user)
        employee_id = getattr(current_user, 'employee_id', None)

        import logging
        logging.getLogger(__name__).warning(
            '🔍 get_archived_leads: employee_id=%s is_admin=%s tenant=%s service=%s',
            employee_id, is_admin, tenant_id, service_param
        )

        query = (
            session.query(
                Opportunity_Details,
                Stage_Master.stage_name,
                Employee_Master.employee_name.label('assigned_to_name'),
                func.coalesce(Opportunity_Details.business_name, Opportunity_Details.opportunity_title).label('business_name'),
                Supplier_Master.supplier_company_name.label('supplier_name')
            )
            .outerjoin(Stage_Master, Opportunity_Details.stage_id == Stage_Master.stage_id)
            .outerjoin(Employee_Master, Opportunity_Details.opportunity_owner_employee_id == Employee_Master.employee_id)
            .outerjoin(Supplier_Master, Opportunity_Details.supplier_id == Supplier_Master.supplier_id)
            .filter(Opportunity_Details.tenant_id == tenant_id)
            .filter(Opportunity_Details.service_id == service_id)
            .filter(Opportunity_Details.is_archived == True)
        )

        if not is_admin and employee_id:
            query = query.filter(Opportunity_Details.opportunity_owner_employee_id == employee_id)

        query = query.order_by(Opportunity_Details.created_at.desc())
        rows = query.all()

        results = []
        for row in rows:
            od = row[0]
            result_dict = {k: _serial(getattr(od, k)) for k in od.__table__.columns.keys()}
            result_dict.update({
                'stage_name': row.stage_name,
                'assigned_to_name': row.assigned_to_name,
                'business_name': row.business_name,
                'supplier_name': row.supplier_name,
            })
            results.append(result_dict)

        logging.getLogger(__name__).warning(
            '✅ get_archived_leads returning %d archived leads',
            len(results)
        )

        return jsonify(results), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()