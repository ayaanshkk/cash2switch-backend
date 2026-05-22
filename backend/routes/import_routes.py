"""
Bulk Import Route for Energy Customers and Leads
Handles Excel/CSV uploads and bulk insertion into database.
Designed for 50k-250k records via background thread + job polling.
"""

from flask import Blueprint, request, jsonify, current_app, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np
import os
import time
import uuid
import threading
from datetime import datetime, timedelta
from sqlalchemy import text
from flask_jwt_extended import jwt_required, get_jwt_identity
import logging
import psycopg2
import psycopg2.extras
from psycopg2.extras import execute_values

from ..models import (
    Client_Master, Project_Details, Energy_Contract_Master,
    Supplier_Master, Employee_Master, Services_Master
)
from .auth_helpers import token_required
from ..db import SessionLocal
from .leads_import_handler import import_leads_handler, download_leads_template_handler
from .job_store import create_job, update_job, get_job, append_error, finish_job, purge_old_jobs

logger = logging.getLogger(__name__)
import_bp = Blueprint('import', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
ENERGY_BATCH_SIZE = 500   # rows per INSERT round-trip
LEADS_BATCH_SIZE  = 2000  # Opportunity_Details is simpler, can go larger


# ---------------------------------------------------------------------------
# Raw psycopg2 connection — bypasses SQLAlchemy entirely for bulk writes
# ---------------------------------------------------------------------------

def _get_raw_connection():
    """
    Open a raw psycopg2 connection independent of SQLAlchemy.
    Handles Supabase URLs which contain options=--search_path=... 
    that psycopg2 cannot parse directly.
    """
    from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
    import re

    db_url = os.environ.get('DATABASE_URL', '')

    # Strip SQLAlchemy driver prefixes
    for prefix in ('postgresql+psycopg2://', 'postgresql+pg8000://',
                   'postgres+psycopg2://', 'postgres+pg8000://'):
        if db_url.startswith(prefix):
            db_url = 'postgresql://' + db_url[len(prefix):]
            break

    if not db_url:
        raise RuntimeError('DATABASE_URL environment variable not set')

    # ── Parse out the options param which Supabase adds ───────────────────────
    # Supabase adds: ?options=--search_path%3Dpublic  (encoded =)
    # psycopg2 chokes on the nested = in the query string
    parsed = urlparse(db_url)
    search_path = None

    if parsed.query:
        # Decode query string manually
        query_params = parse_qs(parsed.query, keep_blank_values=True)
        options_val  = query_params.pop('options', [None])[0]

        if options_val:
            # Extract --search_path=xxx from the options value
            m = re.search(r'--search_path[= ]([^\s&]+)', options_val)
            if m:
                search_path = m.group(1)

        # Rebuild URL without the options param
        new_query = urlencode(
            {k: v[0] for k, v in query_params.items()},
            safe=''
        )
        parsed    = parsed._replace(query=new_query)
        db_url    = urlunparse(parsed)

    # Remove trailing ? if query is now empty
    if db_url.endswith('?'):
        db_url = db_url[:-1]

    conn = psycopg2.connect(db_url)
    conn.autocommit = False

    # Apply search_path if extracted from options
    if search_path:
        cur = conn.cursor()
        # search_path may be comma-separated; quote each schema
        schemas = ', '.join(
            f'"{s.strip()}"' if not s.strip().startswith('"') else s.strip()
            for s in search_path.split(',')
        )
        cur.execute(f'SET search_path TO {schemas}')
        cur.close()
        conn.commit()

    return conn


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_tenant_id_from_user(user):
    if hasattr(user, 'tenant_id') and user.tenant_id is not None:
        return user.tenant_id
    session = SessionLocal()
    try:
        employee = session.query(Employee_Master).filter_by(
            employee_id=user.employee_id
        ).first()
        return employee.tenant_id if employee else None
    finally:
        session.close()


def parse_date(date_value):
    if date_value is None or date_value == '':
        return None
    try:
        if pd.isna(date_value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(date_value, datetime):
        return date_value.date()
    date_str = str(date_value).strip()
    if not date_str or date_str.lower() == 'nan':
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
                '%d %b %Y', '%d %B %Y', '%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value):
    if value is None or value == '':
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    try:
        cleaned = str(value).replace(',', '').replace('£', '').strip()
        return float(cleaned) if cleaned and cleaned != 'nan' else None
    except (ValueError, AttributeError):
        return None


def safe_str(value):
    if value is None or value == '':
        return ''
    try:
        if pd.isna(value):
            return ''
    except (TypeError, ValueError):
        pass
    s = str(value).strip()
    if s.endswith('.0') and s[:-2].replace('.', '', 1).isdigit():
        s = s[:-2]
    return s


def _count_rows(tmp_path: str, file_ext: str) -> int:
    try:
        if file_ext == 'csv':
            with open(tmp_path, 'r', encoding='utf-8-sig', errors='replace') as f:
                return max(0, sum(1 for _ in f) - 1)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            count = max(0, (wb.active.max_row or 1) - 1)
            wb.close()
            return count
    except Exception:
        return 0


def _build_column_map() -> dict:
    return {
        'client_name':                ['client name', 'business name', 'company name', 'name'],
        'trading_name':               ['trading name', 'trade name', 'business', 'company'],
        'main_contact':               ['main contact', 'contact person', 'contact', 'contact name'],
        'position':                   ['position', 'role', 'title', 'job title'],
        'tel_no':                     ['tel no', 'tel no.', 'phone', 'telephone', 'tel', 'phone number', 'contact number'],
        'mobile_no':                  ['mobile no', 'mobile no.', 'mobile', 'cell', 'mobile number'],
        'email':                      ['email', 'e-mail', 'email address'],
        'site_name':                  ['site name', 'site'],
        'month_sold':                 ['month sold', 'sale month'],
        'house_name':                 ['house name'],
        'house_number':               ['house number', 'house no', 'house no.'],
        'door_number':                ['door number', 'door no'],
        'address_line_1':             ['address line 1', 'address 1', 'street', 'address'],
        'address_line_2':             ['address line 2', 'address 2'],
        'address_line_3':             ['address line 3', 'address 3'],
        'town':                       ['town', 'city'],
        'county':                     ['county', 'region'],
        'postcode':                   ['postcode', 'post code', 'zip', 'postal code'],
        'home_door_number':           ['home door number', 'home door no'],
        'home_street':                ['home street'],
        'home_postcode':              ['home post code', 'home postcode'],
        'mpan_top':                   ['mpan top', 'mpan core', 'mpan'],
        'mpan_bottom':                ['mpan bottom', 'mpan llf'],
        'data_source':                ['data source'],
        'old_supplier':               ['old supplier'],
        'supplier':                   ['supplier', 'supplier name', 'new supplier'],
        'payment_type':               ['payment type'],
        'net_notch':                  ['net notch'],
        'term_sold':                  ['term sold', 'in contract', 'contract length', 'term'],
        'agent_sold':                 ['agent sold', 'agent'],
        'start_date':                 ['start date', 'contract start', 'start'],
        'contract_end':               ['contract end', 'end date', 'expiry', 'contract expiry', 'renewal date'],
        'stand_charge':               ['stand charge', 'standing charge', 'standing charge p/day'],
        'rate_1':                     ['rate 1', 'unit rate', 'rate', 'day rate', 'p/kwh'],
        'rate_2':                     ['rate 2', 'night rate'],
        'rate_3':                     ['rate 3', 'evening rate'],
        'aggregator':                 ['aggregator'],
        'annual_usage':               ['annual usage', 'usage', 'kwh', 'annual kwh', 'consumption'],
        'comms_paid':                 ['comms paid', 'commission', 'comm paid'],
        'trading_type':               ['trading type'],
        'company_number':             ['company number', 'co number', 'companies house', 'reg number'],
        'date_of_birth':              ['date of birth', 'dob'],
        'charity_ltd_company_number': ['charity/ltd company number', 'charity number', 'charity ltd company number'],
        'bank_name':                  ['bank name', 'bank'],
        'ac_number':                  ['ac number', 'account number', 'bank account', 'account no'],
        'sort_code':                  ['sort code', 'sortcode'],
        'partner_details':            ['partner details', 'partner'],
        'partner_dob':                ['partner date of birth', 'partner dob'],
        'credit_score':               ['credit score'],
        'password':                   ['password'],
    }


# ---------------------------------------------------------------------------
# File reading — fast path using vectorised pandas
# ---------------------------------------------------------------------------

def _read_file(tmp_path: str, file_ext: str) -> pd.DataFrame:
    """
    Read file into DataFrame. Auto-detects header row.
    Returns normalised DataFrame with string columns.
    """
    def _unnamed_ratio(cols):
        return sum(
            1 for c in cols
            if str(c).startswith('Unnamed:') or str(c).strip() == ''
        ) / max(len(cols), 1)

    def _read_raw(path, ext, header_row):
        if ext == 'csv':
            return pd.read_csv(
                path, encoding='utf-8-sig', dtype=str,
                header=header_row, low_memory=False
            )
        try:
            return pd.read_excel(
                path, engine='openpyxl', dtype=str, header=header_row
            )
        except Exception:
            return pd.read_excel(
                path, engine='xlrd', dtype=str, header=header_row
            )

    df = _read_raw(tmp_path, file_ext, 0)
    if _unnamed_ratio(df.columns) >= 0.5:
        for skip in range(1, 10):
            candidate = _read_raw(tmp_path, file_ext, skip)
            if _unnamed_ratio(candidate.columns) < 0.3:
                df = candidate
                break

    df = df.dropna(how='all').reset_index(drop=True)

    # Replace NaN/None with empty string for string cols — vectorised
    df = df.where(pd.notnull(df), None)
    return df


def _build_field_map(df: pd.DataFrame) -> dict:
    """Map internal field names to actual DataFrame column names."""
    import re
    orig_cols = list(df.columns)
    norm_map = {}
    for orig in orig_cols:
        normed = re.sub(r'\s+', ' ', str(orig).strip().lower().replace('_', ' '))
        norm_map[normed] = orig

    field_to_col = {}
    for field, aliases in _build_column_map().items():
        for alias in aliases:
            if alias in norm_map:
                field_to_col[field] = norm_map[alias]
                break
    return field_to_col


# ---------------------------------------------------------------------------
# Supplier resolution — uses raw_conn, no ORM
# ---------------------------------------------------------------------------

def _load_suppliers(raw_conn) -> dict:
    """Load all suppliers into a lowercase dict."""
    cur = raw_conn.cursor()
    cur.execute("""
        SELECT supplier_id, LOWER(TRIM(supplier_company_name))
        FROM "StreemLyne_MT"."Supplier_Master"
    """)
    result = {row[1]: row[0] for row in cur.fetchall()}
    cur.close()
    return result


def _resolve_supplier(name: str, suppliers_dict: dict, raw_conn) -> int | None:
    if not name:
        return None
    key = name.lower().strip()
    if key in suppliers_dict:
        return suppliers_dict[key]
    # Fuzzy match
    for k, v in suppliers_dict.items():
        if key in k or k in key:
            suppliers_dict[key] = v
            return v
    # Create new
    cur = raw_conn.cursor()
    try:
        cur.execute("""
            INSERT INTO "StreemLyne_MT"."Supplier_Master"
            (supplier_company_name, supplier_contact_name, supplier_provisions, created_at)
            VALUES (%s, 'Auto-imported', 3, %s)
            RETURNING supplier_id
        """, (name, datetime.utcnow()))
        new_id = cur.fetchone()[0]
        raw_conn.commit()
        suppliers_dict[key] = new_id
        return new_id
    except Exception as e:
        raw_conn.rollback()
        print(f"Supplier create failed '{name}': {e}")
        return None
    finally:
        cur.close()


# ---------------------------------------------------------------------------
# MPAN preload — raw_conn only
# ---------------------------------------------------------------------------

def _load_existing_mpans(raw_conn, tenant_id) -> dict:
    """
    Load existing MPANs for this tenant.
    Returns dict: mpan_lower -> { end_date, is_draft, assigned_employee_id }
    """
    cur = raw_conn.cursor()
    cur.execute("""
        SELECT ecm.mpan_number, ecm.contract_end_date,
               cm.is_draft, pd.assigned_employee_id
        FROM "StreemLyne_MT"."Energy_Contract_Master" ecm
        JOIN "StreemLyne_MT"."Project_Details" pd
            ON ecm.project_id = pd.project_id
        JOIN "StreemLyne_MT"."Client_Master" cm
            ON pd.client_id = cm.client_id
        WHERE cm.is_deleted = FALSE
          AND cm.tenant_id = %s
          AND ecm.mpan_number IS NOT NULL
          AND ecm.mpan_number != ''
    """, (str(tenant_id),))
    result = {}
    for mpan_num, end_dt, is_dr, assigned in cur.fetchall():
        if mpan_num:
            result[mpan_num.strip().lower()] = {
                'end_date': end_dt,
                'is_draft': is_dr,
                'assigned_employee_id': assigned,
            }
    cur.close()
    return result


def _load_existing_lead_mpans(raw_conn, tenant_id) -> dict:
    """Load existing lead MPANs for this tenant only."""
    cur = raw_conn.cursor()
    cur.execute("""
        SELECT od.mpan_mpr
        FROM "StreemLyne_MT"."Opportunity_Details" od
        WHERE od.mpan_mpr IS NOT NULL 
          AND od.mpan_mpr != ''
          AND od.tenant_id = %s
    """, (str(tenant_id),))
    result = {}
    for (mpan,) in cur.fetchall():
        if mpan:
            result[mpan.strip().lower()] = True
    cur.close()
    return result


def _get_default_stage(raw_conn) -> int:
    cur = raw_conn.cursor()
    cur.execute("""
        SELECT stage_id FROM "StreemLyne_MT"."Stage_Master"
        ORDER BY stage_id LIMIT 1
    """)
    row = cur.fetchone()
    cur.close()
    return row[0] if row else 1


# ---------------------------------------------------------------------------
# Batch flush — energy customers (renewals)
# ---------------------------------------------------------------------------

def _flush_energy_batch(
    raw_conn, batch, tenant_id, employee_id,
    opportunity_owner_id, is_draft_import, import_service_id,
    existing_mpans,
):
    if not batch:
        return 0, 0

    now = datetime.utcnow()
    cur = raw_conn.cursor()

    try:
        # ── 1. Client_Master ──────────────────────────────────────────────────
        client_tuples = [
            (
                str(tenant_id), opportunity_owner_id,
                r['business_name'] or '', r['contact_person'] or '',
                r['address'] or '', r['postcode'] or '',
                r['tel_no'] or '', r['mobile_no'], r['email'] or '',
                r['position'], r['company_number'], r['date_of_birth'],
                r['charity_ltd'], r['partner_details'],
                r['bank_name'], r['account_number'], r['sort_code'],
                r['home_door'], r['home_street'],
                r['partner_dob'], r['credit_score'],
                is_draft_import, now,
                '',     # client_website
                False,  # is_deleted
                False,  # is_archived
                1,      # default_currency_id
                False,  # is_allocated
            )
            for r in batch
        ]

        client_ids = execute_values(
            cur,
            """
            INSERT INTO "StreemLyne_MT"."Client_Master"
            (tenant_id, assigned_employee_id, client_company_name, client_contact_name,
             address, post_code, client_phone, client_mobile, client_email,
             position, company_number, date_of_birth, charity_ltd_company_number,
             partner_details, bank_name, account_number, sort_code,
             home_door_number, home_street, partner_dob, credit_score,
             is_draft, created_at, client_website,
             is_deleted, is_archived, default_currency_id, is_allocated)
            VALUES %s
            RETURNING client_id
            """,
            client_tuples,
            fetch=True,
        )
        client_ids = [row[0] for row in client_ids]

        if len(client_ids) != len(batch):
            raw_conn.rollback()
            cur.close()
            return 0, len(batch)

        # ── 2. Project_Details ────────────────────────────────────────────────
        project_tuples = [
            (
                cid,
                r['business_name'] or 'Renewal Contract',
                'Imported renewal contract',
                r['site_address'] or '',
                r['annual_usage'],
                employee_id,
                opportunity_owner_id,
                r['contract_start'],
                r['contract_end'],
                r['site_name'],
                r['month_sold'],
                r['house_name'],
                r['house_number'],
                r['door_number'],
                r['town'],
                r['county'],
                None,   # status
                now,
                now,
            )
            for cid, r in zip(client_ids, batch)
        ]

        project_ids = execute_values(
            cur,
            """
            INSERT INTO "StreemLyne_MT"."Project_Details"
            (client_id, project_title, project_description, address, "Misc_Col2",
             employee_id, assigned_employee_id, start_date, end_date,
             site_name, month_sold, house_name, house_number, door_number,
             town, county, status, created_at, updated_at)
            VALUES %s
            RETURNING project_id
            """,
            project_tuples,
            fetch=True,
        )
        project_ids = [row[0] for row in project_ids]

        # ── 3. Energy_Contract_Master ─────────────────────────────────────────
        contract_tuples = [
            (
                pid, employee_id,
                r['supplier_id'], r['old_supplier_id'],
                r['contract_start'], r['contract_end'],
                import_service_id,
                r['rate_1'] or 0.0,
                1,      # currency_id
                now, now,
                r['mpan_top'] or '', r['mpan_bottom'] or '',
                r['net_notch'], r['rate_2'], r['rate_3'],
                r['comms_paid'], r['stand_charge'],
                r['aggregator'], r['rate_1'],
                r['payment_type'], r['term_sold'],
                '',     # terms_of_sale
            )
            for pid, r in zip(project_ids, batch)
        ]

        execute_values(
            cur,
            """
            INSERT INTO "StreemLyne_MT"."Energy_Contract_Master"
            (project_id, employee_id, supplier_id, old_supplier_id,
             contract_start_date, contract_end_date, service_id,
             unit_rate, currency_id, created_at, updated_at,
             mpan_number, mpan_bottom, net_notch, rate_2, rate_3,
             comms_paid, standing_charge, aggregator, rate_1,
             payment_type, term_sold, terms_of_sale)
            VALUES %s
            """,
            contract_tuples,
            fetch=False,
        )

        raw_conn.commit()
        cur.close()

        # Update intra-file dedup map
        for r in batch:
            if r['mpan_top']:
                existing_mpans[r['mpan_top'].strip().lower()] = {
                    'end_date': r['end_date'],
                    'is_draft': is_draft_import,
                    'assigned_employee_id': opportunity_owner_id,
                }

        return len(batch), 0

    except Exception as e:
        raw_conn.rollback()
        try:
            cur.close()
        except Exception:
            pass
        print(f"Energy batch failed: {str(e).split(chr(10))[0][:200]}")
        return 0, len(batch)


# ---------------------------------------------------------------------------
# Background worker — energy customers
# ---------------------------------------------------------------------------

def _run_energy_import(
    job_id: str,
    tmp_path: str,
    file_ext: str,
    tenant_id,
    employee_id: int,
    is_draft_import: bool,
    opportunity_owner_id,
    assigned_employee_name,
    import_service_id: int,
):
    sql_logger = logging.getLogger('sqlalchemy.engine')
    original_level = sql_logger.level
    sql_logger.setLevel(logging.WARNING)

    raw_conn = None

    try:
        # ── 1. Read file ──────────────────────────────────────────────────────
        try:
            df = _read_file(tmp_path, file_ext)
        except Exception as e:
            finish_job(job_id, 'failed')
            append_error(job_id, f'Failed to read file: {str(e)[:200]}')
            return
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        field_map = _build_field_map(df)
        total_rows = len(df)
        update_job(job_id, total=total_rows)
        print(f"[job:{job_id}] Energy import: {total_rows} rows, "
              f"mapped {len(field_map)} fields, "
              f"draft={is_draft_import}, service={import_service_id}")

        if total_rows == 0:
            finish_job(job_id, 'done')
            update_job(job_id, processed=0, successful=0, duplicates=0)
            return

        # ── 2. Open raw connection + preload lookups ───────────────────────────
        raw_conn = _get_raw_connection()
        suppliers_dict  = _load_suppliers(raw_conn)
        existing_mpans  = _load_existing_mpans(raw_conn, tenant_id)

        print(f"[job:{job_id}] Loaded {len(suppliers_dict)} suppliers, "
              f"{len(existing_mpans)} existing MPANs")

        def gcol(field):
            col = field_map.get(field, '')
            return col if col and col in df.columns else None

        # ── 3. Vectorised parse — much faster than iterrows() ─────────────────
        # Pull all needed columns as numpy arrays once, then zip
        def col_vals(field):
            c = gcol(field)
            return df[c].tolist() if c else [None] * total_rows

        client_names   = col_vals('client_name')
        trading_names  = col_vals('trading_name')
        main_contacts  = col_vals('main_contact')
        positions      = col_vals('position')
        tel_nos        = col_vals('tel_no')
        mobile_nos     = col_vals('mobile_no')
        emails         = col_vals('email')
        site_names     = col_vals('site_name')
        addr1s         = col_vals('address_line_1')
        addr2s         = col_vals('address_line_2')
        addr3s         = col_vals('address_line_3')
        towns          = col_vals('town')
        counties       = col_vals('county')
        postcodes      = col_vals('postcode')
        mpan_tops      = col_vals('mpan_top')
        mpan_bottoms   = col_vals('mpan_bottom')
        suppliers      = col_vals('supplier')
        old_suppliers  = col_vals('old_supplier')
        payment_types  = col_vals('payment_type')
        annual_usages  = col_vals('annual_usage')
        start_dates    = col_vals('start_date')
        end_dates      = col_vals('contract_end')
        stand_charges  = col_vals('stand_charge')
        rates_1        = col_vals('rate_1')
        rates_2        = col_vals('rate_2')
        rates_3        = col_vals('rate_3')
        net_notches    = col_vals('net_notch')
        comms_paids    = col_vals('comms_paid')
        terms_sold     = col_vals('term_sold')
        aggregators    = col_vals('aggregator')
        company_nos    = col_vals('company_number')
        dobs           = col_vals('date_of_birth')
        charity_nos    = col_vals('charity_ltd_company_number')
        month_solds    = col_vals('month_sold')
        house_names    = col_vals('house_name')
        house_numbers  = col_vals('house_number')
        door_numbers   = col_vals('door_number')
        partner_dets   = col_vals('partner_details')
        bank_names     = col_vals('bank_name')
        ac_numbers     = col_vals('ac_number')
        sort_codes     = col_vals('sort_code')
        home_doors     = col_vals('home_door_number')
        home_streets   = col_vals('home_street')
        partner_dobs   = col_vals('partner_dob')
        credit_scores  = col_vals('credit_score')

        # ── 4. Row loop ───────────────────────────────────────────────────────
        success_count   = 0
        error_count     = 0
        duplicate_count = 0
        pending_batch   = []
        start_time      = time.time()

        for i in range(total_rows):
            try:
                client_name    = safe_str(client_names[i])
                trading_name   = safe_str(trading_names[i])
                main_contact   = safe_str(main_contacts[i])
                position       = safe_str(positions[i])
                tel_no         = safe_str(tel_nos[i])
                mobile_no      = safe_str(mobile_nos[i])
                email          = safe_str(emails[i])
                site_name      = safe_str(site_names[i])
                addr1          = safe_str(addr1s[i])
                addr2          = safe_str(addr2s[i])
                addr3          = safe_str(addr3s[i])
                town           = safe_str(towns[i])
                county         = safe_str(counties[i])
                postcode       = safe_str(postcodes[i])
                mpan_top       = safe_str(mpan_tops[i])
                mpan_bottom    = safe_str(mpan_bottoms[i])
                supplier_name  = safe_str(suppliers[i])
                old_sup_name   = safe_str(old_suppliers[i])
                payment_type   = safe_str(payment_types[i])
                annual_usage   = parse_number(annual_usages[i])
                start_date     = parse_date(start_dates[i])
                end_date       = parse_date(end_dates[i])
                stand_charge   = parse_number(stand_charges[i])
                rate_1         = parse_number(rates_1[i])
                rate_2         = parse_number(rates_2[i])
                rate_3         = parse_number(rates_3[i])
                net_notch      = parse_number(net_notches[i])
                comms_paid     = parse_number(comms_paids[i])
                term_sold      = parse_number(terms_sold[i])
                aggregator     = safe_str(aggregators[i])
                company_number = safe_str(company_nos[i])
                dob            = parse_date(dobs[i])
                charity_no     = safe_str(charity_nos[i])
                month_sold     = safe_str(month_solds[i])
                house_name     = safe_str(house_names[i])
                house_number   = safe_str(house_numbers[i])
                door_number    = safe_str(door_numbers[i])
                partner_det    = safe_str(partner_dets[i])
                bank_name      = safe_str(bank_names[i])
                ac_number      = safe_str(ac_numbers[i])
                sort_code      = safe_str(sort_codes[i])
                home_door      = safe_str(home_doors[i])
                home_street    = safe_str(home_streets[i])
                partner_dob    = parse_date(partner_dobs[i])
                credit_score   = parse_number(credit_scores[i])

                address_parts = [p for p in [addr1, addr2, addr3, town, county]
                                 if p and p.lower() != 'nan']
                address       = ', '.join(address_parts)
                site_address  = site_name or address
                business_name  = trading_name or client_name
                contact_person = main_contact or client_name

                if not business_name and not tel_no and not email and not mpan_top and not contact_person:
                    continue

                # ── MPAN duplicate check ──────────────────────────────────────
                if mpan_top:
                    existing = existing_mpans.get(mpan_top.strip().lower())
                    if existing:
                        is_assigned_non_draft = (
                            existing.get('assigned_employee_id') is not None
                            and not existing.get('is_draft')
                        )
                        existing_end = existing.get('end_date')
                        same_or_older = (
                            existing_end and end_date and end_date <= existing_end
                        )
                        if is_assigned_non_draft and (same_or_older or not end_date):
                            duplicate_count += 1
                            continue

                contract_start = start_date or datetime.utcnow().date()
                contract_end   = end_date or (contract_start + timedelta(days=365))

                pending_batch.append({
                    'business_name':   business_name or '',
                    'contact_person':  contact_person or '',
                    'address':         address or '',
                    'postcode':        postcode or '',
                    'tel_no':          tel_no or '',
                    'mobile_no':       mobile_no or None,
                    'email':           email or '',
                    'position':        position or None,
                    'company_number':  company_number or None,
                    'date_of_birth':   dob,
                    'charity_ltd':     charity_no or None,
                    'partner_details': partner_det or None,
                    'bank_name':       bank_name or None,
                    'account_number':  ac_number or None,
                    'sort_code':       sort_code or None,
                    'home_door':       home_door or None,
                    'home_street':     home_street or None,
                    'partner_dob':     partner_dob,
                    'credit_score':    credit_score,
                    'site_address':    site_address or address or '',
                    'annual_usage':    int(annual_usage) if annual_usage else None,
                    'site_name':       site_name or None,
                    'month_sold':      month_sold or None,
                    'house_name':      house_name or None,
                    'house_number':    house_number or None,
                    'door_number':     door_number or None,
                    'town':            town or None,
                    'county':          county or None,
                    'supplier_id':     _resolve_supplier(supplier_name, suppliers_dict, raw_conn),
                    'old_supplier_id': _resolve_supplier(old_sup_name, suppliers_dict, raw_conn) if old_sup_name else None,
                    'contract_start':  contract_start,
                    'contract_end':    contract_end,
                    'mpan_top':        mpan_top or '',
                    'mpan_bottom':     mpan_bottom or '',
                    'rate_1':          rate_1,
                    'rate_2':          rate_2,
                    'rate_3':          rate_3,
                    'net_notch':       net_notch,
                    'comms_paid':      comms_paid,
                    'stand_charge':    stand_charge,
                    'aggregator':      aggregator or None,
                    'payment_type':    payment_type or None,
                    'term_sold':       term_sold,
                    'end_date':        end_date,
                })

            except Exception as row_err:
                error_count += 1
                append_error(job_id, f"Row {i + 2}: {str(row_err).split(chr(10))[0][:120]}")
                continue

            if len(pending_batch) >= ENERGY_BATCH_SIZE:
                ins, err = _flush_energy_batch(
                    raw_conn, pending_batch, tenant_id, employee_id,
                    opportunity_owner_id, is_draft_import,
                    import_service_id, existing_mpans,
                )
                success_count += ins
                error_count   += err
                pending_batch  = []

                elapsed = time.time() - start_time
                rate    = success_count / elapsed if elapsed > 0 else 1
                eta_s   = (total_rows - i) / rate if rate > 0 else 0
                update_job(
                    job_id,
                    processed=i + 1,
                    successful=success_count,
                    duplicates=duplicate_count,
                )
                print(f"[job:{job_id}] {i+1}/{total_rows} | "
                      f"{success_count} inserted | "
                      f"{rate:.0f} rec/s | "
                      f"ETA {eta_s/60:.1f} min")

        # Final batch
        if pending_batch:
            ins, err = _flush_energy_batch(
                raw_conn, pending_batch, tenant_id, employee_id,
                opportunity_owner_id, is_draft_import,
                import_service_id, existing_mpans,
            )
            success_count += ins
            error_count   += err

        elapsed = time.time() - start_time
        rate = success_count / elapsed if elapsed > 0 else 0
        print(
            f"[job:{job_id}] DONE — {success_count} ok, "
            f"{duplicate_count} dup, {error_count} err "
            f"in {elapsed:.1f}s ({rate:.0f} rec/s)"
        )
        finish_job(job_id, 'done')
        update_job(
            job_id,
            processed=total_rows,
            successful=success_count,
            duplicates=duplicate_count,
        )

    except Exception as fatal:
        import traceback; traceback.print_exc()
        if raw_conn:
            try: raw_conn.rollback()
            except Exception: pass
        finish_job(job_id, 'failed')
        append_error(job_id, f"Fatal: {str(fatal)[:200]}")
        print(f"[job:{job_id}] FATAL: {fatal}")

    finally:
        sql_logger.setLevel(original_level)
        if raw_conn:
            try: raw_conn.close()
            except Exception: pass


# ---------------------------------------------------------------------------
# Background worker — leads
# ---------------------------------------------------------------------------

def _run_leads_import(
    job_id: str,
    tmp_path: str,
    file_ext: str,
    tenant_id,
    employee_id: int,
    is_draft_import: bool,
    opportunity_owner_id,
    assigned_employee_name,
    import_service_id: int,
    default_stage_id: int,
):
    sql_logger = logging.getLogger('sqlalchemy.engine')
    original_level = sql_logger.level
    sql_logger.setLevel(logging.WARNING)

    raw_conn = None

    try:
        # ── 1. Read file ──────────────────────────────────────────────────────
        try:
            df = _read_file(tmp_path, file_ext)
        except Exception as e:
            finish_job(job_id, 'failed')
            append_error(job_id, f'Failed to read file: {str(e)[:200]}')
            return
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        field_map  = _build_field_map(df)
        total_rows = len(df)
        update_job(job_id, total=total_rows)
        print(f"[job:{job_id}] Leads: {total_rows} rows, {len(field_map)} fields mapped")

        if total_rows == 0:
            finish_job(job_id, 'done')
            update_job(job_id, processed=0, successful=0, duplicates=0)
            return

        # ── 2. Open raw connection + preload lookups FIRST ────────────────────
        # Must happen before vectorised extraction so suppliers_dict exists
        raw_conn = _get_raw_connection()
        suppliers_dict      = _load_suppliers(raw_conn)
        existing_lead_mpans = _load_existing_lead_mpans(raw_conn, tenant_id)

        # ── Pre-generate tenant_opportunity_id ────────────────────────────────
        cur = raw_conn.cursor()
        cur.execute("""
            SELECT COALESCE(MAX(tenant_opportunity_id), 0)
            FROM "StreemLyne_MT"."Opportunity_Details"
            WHERE tenant_id = %s
        """, (str(tenant_id),))
        tenant_opp_counter = cur.fetchone()[0]
        cur.close()
        print(f"[job:{job_id}] tenant_opportunity_id starting from {tenant_opp_counter + 1}")

        # ── 3. Vectorised column extraction ───────────────────────────────────
        def col(field) -> pd.Series:
            c = field_map.get(field)
            if c and c in df.columns:
                return df[c].fillna('')
            return pd.Series([''] * total_rows, index=df.index)

        def vstr(series):
            s = series.astype(str).str.strip()
            mask = s.str.endswith('.0') & s.str[:-2].str.match(r'^\d+$')
            s = s.where(~mask, s.str[:-2])
            return s.where(~s.isin(['nan', 'None', 'NaT']), '')

        def vdate(series):
            return [parse_date(v) for v in series.tolist()]

        def vnum(series):
            s = series.astype(str).str.replace(',', '', regex=False)
            s = s.str.replace('£', '', regex=False).str.strip()
            s = s.where(s.str.match(r'^-?\d+(\.\d+)?$'), other=None)
            return pd.to_numeric(s, errors='coerce')

        l_client   = vstr(col('client_name')).tolist()
        l_trading  = vstr(col('trading_name')).tolist()
        l_contact  = vstr(col('main_contact')).tolist()
        l_tel      = vstr(col('tel_no')).tolist()
        l_mobile   = vstr(col('mobile_no')).tolist()
        l_email    = vstr(col('email')).tolist()
        l_mpan_top = vstr(col('mpan_top')).tolist()
        l_mpan_bot = vstr(col('mpan_bottom')).tolist()
        l_supplier = vstr(col('supplier')).tolist()
        l_postcode = vstr(col('postcode')).tolist()
        l_payment  = vstr(col('payment_type')).tolist()

        l_usage    = vnum(col('annual_usage')).tolist()
        l_sc       = vnum(col('stand_charge')).tolist()
        l_r1       = vnum(col('rate_1')).tolist()
        l_r2       = vnum(col('rate_2')).tolist()
        l_r3       = vnum(col('rate_3')).tolist()
        l_nn       = vnum(col('net_notch')).tolist()

        d_start    = vdate(col('start_date'))
        d_end      = vdate(col('contract_end'))

        print(f"[job:{job_id}] Vectorised extraction complete")

        # Pre-resolve all unique suppliers in one pass — zero per-row DB hits
        unique_sups = set(l_supplier) - {'', 'nan', 'None'}
        for sup in unique_sups:
            _resolve_supplier(sup, suppliers_dict, raw_conn)
        print(f"[job:{job_id}] Resolved {len(unique_sups)} unique suppliers")

        # ── 4. Row loop ───────────────────────────────────────────────────────
        def ns(v): return v if v else None

        def nn_val(v):
            try:
                return float(v) if v is not None and not (
                    isinstance(v, float) and pd.isna(v)
                ) else None
            except (TypeError, ValueError):
                return None

        success_count   = 0
        error_count     = 0
        duplicate_count = 0
        pending_tuples  = []   # ← pre-built tuples, not dicts
        start_time      = time.time()
        now             = datetime.utcnow()

        for i in range(total_rows):
            try:
                client_name = l_client[i]
                trading     = l_trading[i]
                contact     = l_contact[i]
                tel         = l_tel[i]
                mobile      = l_mobile[i]
                email       = l_email[i]
                mpan        = l_mpan_top[i]
                mpan_b      = l_mpan_bot[i]
                business    = trading or client_name
                person      = contact or client_name
                sup_name    = l_supplier[i]
                postcode    = l_postcode[i]
                payment     = l_payment[i]
                start_d     = d_start[i]
                end_d       = d_end[i]

                # Skip empty rows
                if not business and not tel and not mobile and not email and not mpan and not person:
                    continue

                # In-memory MPAN duplicate check
                if mpan:
                    mpan_key = mpan.strip().lower()
                    if mpan_key in existing_lead_mpans:
                        duplicate_count += 1
                        continue

                # Supplier lookup — already pre-resolved, pure dict lookup
                supplier_id = None
                if sup_name:
                    sup_key = sup_name.lower().strip()
                    supplier_id = suppliers_dict.get(sup_key)
                    if not supplier_id:
                        for k, v in suppliers_dict.items():
                            if sup_key in k or k in sup_key:
                                supplier_id = v
                                break

                tenant_opp_counter += 1

                pending_tuples.append((
                    str(tenant_id),
                    None,
                    business or '',
                    'Imported lead',
                    now,
                    opportunity_owner_id,
                    default_stage_id,
                    0, 1, now,
                    ns(business), ns(person),
                    ns(tel) or ns(mobile),
                    ns(mobile), ns(email),
                    ns(mpan), ns(mpan_b),
                    start_d, end_d,
                    import_service_id, supplier_id,
                    nn_val(l_usage[i]),
                    nn_val(l_sc[i]),
                    nn_val(l_r1[i]), nn_val(l_r2[i]), nn_val(l_r3[i]),
                    nn_val(l_nn[i]),
                    ns(payment), ns(postcode),
                    is_draft_import,
                    tenant_opp_counter,
                ))

                # Register MPAN immediately for intra-file dedup
                if mpan:
                    existing_lead_mpans[mpan.strip().lower()] = True

            except Exception as row_err:
                error_count += 1
                if error_count <= 20:
                    append_error(
                        job_id,
                        f"Row {i + 2}: {str(row_err).split(chr(10))[0][:120]}"
                    )
                continue

            # ── Flush batch ───────────────────────────────────────────────────
            if len(pending_tuples) >= LEADS_BATCH_SIZE:
                ins, skipped = _flush_leads_tuples(raw_conn, pending_tuples)
                success_count   += ins
                duplicate_count += skipped
                pending_tuples   = []

                elapsed = time.time() - start_time
                rate    = success_count / elapsed if elapsed > 0 else 1
                eta_s   = (total_rows - i) / rate if rate > 0 else 0
                update_job(
                    job_id,
                    processed=i + 1,
                    successful=success_count,
                    duplicates=duplicate_count,
                )
                print(f"[job:{job_id}] {i+1}/{total_rows} | "
                      f"{success_count} ok | {rate:.0f} rec/s | "
                      f"ETA {eta_s/60:.1f}m")

        # Final batch
        if pending_tuples:
            ins, skipped = _flush_leads_tuples(raw_conn, pending_tuples)
            success_count   += ins
            duplicate_count += skipped

        elapsed = time.time() - start_time
        rate = success_count / elapsed if elapsed > 0 else 0
        print(f"[job:{job_id}] LEADS DONE — {success_count} ok, "
              f"{duplicate_count} dup, {error_count} err "
              f"in {elapsed:.1f}s ({rate:.0f} rec/s)")

        finish_job(job_id, 'done')
        update_job(
            job_id,
            processed=total_rows,
            successful=success_count,
            duplicates=duplicate_count,
        )

    except Exception as fatal:
        import traceback; traceback.print_exc()
        if raw_conn:
            try: raw_conn.rollback()
            except Exception: pass
        finish_job(job_id, 'failed')
        append_error(job_id, f"Fatal: {str(fatal)[:200]}")

    finally:
        sql_logger.setLevel(original_level)
        if raw_conn:
            try: raw_conn.close()
            except Exception: pass

def _flush_leads_tuples(raw_conn, tuples):
    if not tuples:
        return 0, 0

    try:
        # ── Must commit any open transaction before changing autocommit ────────
        raw_conn.commit()

        # ── DDL: disable trigger (needs autocommit=True) ──────────────────────
        raw_conn.autocommit = True
        cur = raw_conn.cursor()
        cur.execute("""
            ALTER TABLE "StreemLyne_MT"."Opportunity_Details"
            DISABLE TRIGGER trigger_set_tenant_opportunity_display_id
        """)
        cur.close()
        raw_conn.autocommit = False

        # ── Bulk insert ───────────────────────────────────────────────────────
        cur = raw_conn.cursor()
        execute_values(
            cur,
            """
            INSERT INTO "StreemLyne_MT"."Opportunity_Details"
            (tenant_id, client_id, opportunity_title, opportunity_description,
             opportunity_date, opportunity_owner_employee_id, stage_id,
             opportunity_value, currency_id, created_at,
             business_name, contact_person, tel_number, mobile_no, email,
             mpan_mpr, mpan_bottom, start_date, end_date, service_id,
             supplier_id, annual_usage, stand_charge, rate_1, rate_2,
             rate_3, net_notch, payment_type, postcode, is_draft,
             tenant_opportunity_id)
            VALUES %s
            """,
            tuples,
            fetch=False,
            page_size=500,
        )
        raw_conn.commit()
        cur.close()

        # ── DDL: re-enable trigger ────────────────────────────────────────────
        raw_conn.commit()  # ensure clean state before autocommit switch
        raw_conn.autocommit = True
        cur = raw_conn.cursor()
        cur.execute("""
            ALTER TABLE "StreemLyne_MT"."Opportunity_Details"
            ENABLE TRIGGER trigger_set_tenant_opportunity_display_id
        """)
        cur.close()
        raw_conn.autocommit = False

        return len(tuples), 0

    except Exception as e:
        # Restore safe state
        try:
            raw_conn.autocommit = False
        except Exception:
            pass
        try:
            raw_conn.rollback()
        except Exception:
            pass
        # Always re-enable trigger
        try:
            raw_conn.commit()
            raw_conn.autocommit = True
            cur2 = raw_conn.cursor()
            cur2.execute("""
                ALTER TABLE "StreemLyne_MT"."Opportunity_Details"
                ENABLE TRIGGER trigger_set_tenant_opportunity_display_id
            """)
            cur2.close()
            raw_conn.autocommit = False
        except Exception:
            pass
        print(f"Leads flush failed: {str(e).split(chr(10))[0][:200]}")
        return 0, len(tuples)

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@import_bp.route('/energy-customers', methods=['POST', 'OPTIONS'])
@token_required
def import_energy_customers():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Please upload .xlsx, .xls, or .csv'}), 400

    tenant_id = get_tenant_id_from_user(request.current_user)
    if not tenant_id:
        return jsonify({'error': 'Tenant not found for user'}), 400

    employee_id          = request.current_user.employee_id
    is_draft_import      = str(request.form.get('is_draft', '')).strip().lower() in {'1', 'true', 'yes', 'on'}
    assigned_employee_id = request.form.get('assigned_employee_id', type=int)
    opportunity_owner_id = None if is_draft_import else (assigned_employee_id or employee_id)

    assigned_employee_name = None
    if assigned_employee_id and not is_draft_import:
        session = SessionLocal()
        try:
            ae = session.query(Employee_Master).filter_by(
                employee_id=assigned_employee_id,
                tenant_id=tenant_id,
            ).first()
            if not ae:
                return jsonify({'error': f'Invalid employee ID: {assigned_employee_id}'}), 400
            assigned_employee_name = ae.employee_name
        finally:
            session.close()

    service_param     = request.form.get('service') or request.args.get('service', 'utilities')
    service_id_map    = {'utilities': 1, 'electricity': 1, 'water': 2, 'gas': 3}
    import_service_id = service_id_map.get(service_param.strip().lower(), 1)

    filename  = secure_filename(file.filename)
    file_ext  = filename.rsplit('.', 1)[1].lower()
    tmp_path  = f'/tmp/import_{uuid.uuid4().hex}.{file_ext}'
    file.save(tmp_path)

    total_rows = _count_rows(tmp_path, file_ext)
    purge_old_jobs(max_age_hours=24)

    job_id = uuid.uuid4().hex
    create_job(job_id, total_rows, tenant_id=tenant_id)

    print(f"[job:{job_id}] Energy import thread starting — "
          f"{total_rows} rows, tenant={tenant_id}, service={import_service_id}")

    thread = threading.Thread(
        target=_run_energy_import,
        args=(
            job_id, tmp_path, file_ext,
            tenant_id, employee_id,
            is_draft_import, opportunity_owner_id, assigned_employee_name,
            import_service_id,
        ),
        daemon=True,
        name=f'import-energy-{job_id[:8]}',
    )
    thread.start()

    return jsonify({
        'job_id':     job_id,
        'status':     'running',
        'total_rows': total_rows,
        'message':    f'Import started. Poll /import/status/{job_id} for progress.',
    }), 202


@import_bp.route('/leads', methods=['POST', 'OPTIONS'])
@token_required
def import_leads():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    filename = secure_filename(file.filename)
    if not ('.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls', 'csv'}):
        return jsonify({'error': 'Invalid file type. Please upload .xlsx, .xls, or .csv'}), 400

    tenant_id = get_tenant_id_from_user(request.current_user)
    if not tenant_id:
        return jsonify({'error': 'Tenant not found for user'}), 400

    employee_id          = request.current_user.employee_id
    is_draft_import      = str(request.form.get('is_draft', '')).strip().lower() in {'1', 'true', 'yes', 'on'}
    assigned_employee_id = request.form.get('assigned_employee_id', type=int)
    opportunity_owner_id = None if is_draft_import else (assigned_employee_id or employee_id)

    assigned_employee_name = None
    if assigned_employee_id and not is_draft_import:
        session = SessionLocal()
        try:
            ae = session.query(Employee_Master).filter_by(
                employee_id=assigned_employee_id,
                tenant_id=tenant_id,
            ).first()
            if not ae:
                return jsonify({'error': f'Invalid employee ID: {assigned_employee_id}'}), 400
            assigned_employee_name = ae.employee_name
        finally:
            session.close()

    service_param     = request.args.get('service', 'electricity')
    service_id_map    = {'electricity': 1, 'utilities': 1, 'water': 2, 'gas': 3}
    import_service_id = service_id_map.get(service_param.strip().lower(), 1)

    # Resolve default stage before leaving request context
    session = SessionLocal()
    try:
        from ..models import Stage_Master
        default_stage    = session.query(Stage_Master).order_by(Stage_Master.stage_id).first()
        default_stage_id = default_stage.stage_id if default_stage else 1
    finally:
        session.close()

    file_ext = filename.rsplit('.', 1)[1].lower()
    tmp_path = f'/tmp/import_{uuid.uuid4().hex}.{file_ext}'
    file.save(tmp_path)

    total_rows = _count_rows(tmp_path, file_ext)
    purge_old_jobs(max_age_hours=24)

    job_id = uuid.uuid4().hex
    create_job(job_id, total_rows, tenant_id=tenant_id)

    print(f"[job:{job_id}] Leads import thread starting — "
          f"{total_rows} rows, tenant={tenant_id}")

    thread = threading.Thread(
        target=_run_leads_import,
        args=(
            job_id, tmp_path, file_ext,
            tenant_id, employee_id,
            is_draft_import, opportunity_owner_id, assigned_employee_name,
            import_service_id, default_stage_id,
        ),
        daemon=True,
        name=f'import-leads-{job_id[:8]}',
    )
    thread.start()

    return jsonify({
        'job_id':     job_id,
        'status':     'running',
        'total_rows': total_rows,
        'message':    f'Import started. Poll /import/status/{job_id} for progress.',
    }), 202


@import_bp.route('/status/<job_id>', methods=['GET'])
@token_required
def import_status(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404

    total     = max(job.get('total') or 1, 1)
    processed = job.get('processed', 0)

    return jsonify({
        'job_id':       job_id,
        'status':       job.get('status', 'running'),
        'total':        total,
        'processed':    processed,
        'successful':   job.get('successful', 0),
        'duplicates':   job.get('duplicates', 0),
        'progress_pct': round(processed / total * 100, 1),
        'errors':       job.get('errors', [])[-20:],
        'started_at':   job.get('started_at'),
        'finished_at':  job.get('finished_at'),
    }), 200


# ---------------------------------------------------------------------------
# Template downloads
# ---------------------------------------------------------------------------

@import_bp.route('/template', methods=['GET'])
@token_required
def download_template():
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Renewals Import Template"

        headers = [
            "Client Name", "Trading Name", "Main Contact", "Position",
            "Tel No", "Mobile No", "Email", "Site Name", "Month Sold", "",
            "Address Line 1", "Address Line 2", "Address Line 3",
            "Town", "County", "Postcode", "Mpan Top", "Mpan Bottom",
            "", "", "Data Source", "Welcome Call", "Payment Type",
            "Supplier", "Net Notch", "In Contract", "Agent Sold",
            "Start Date", "Contract End", "Stand Charge",
            "Rate 1", "Rate 2", "Rate 3", "", "", "Aggregator",
            "Annual Usage", "Comms Paid", "Company Number",
            "Date of Birth", "Bank Name", "Ac Number",
            "Sort Code", "Charity/Ltd Company Number", "Partner Details",
        ]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if header:
                cell.fill = header_fill
                cell.font = header_font

        example = [
            "ABC Limited", "ABC Trading", "John Smith", "Director",
            "07700900000", "07700900001", "john@abc.com", "Main Site", "Jan-24", "",
            "123 Main St", "Unit 5", "Industrial Estate", "London",
            "Greater London", "SW1A 1AA", "1100012314490", "04031N12", "", "",
            "Renewals", "Yes", "DD", "British Gas", "0.1",
            "1 Year", "Sales Team", "01/01/2024", "31/12/2024", "45.13",
            "35.00", "26.46", "", "", "",
            "Online", "25000", "7.92", "12345678", "",
            "Barclays", "12345678", "20-00-00", "", "",
        ]
        for col, value in enumerate(example, 1):
            ws.cell(row=2, column=col, value=value)

        for col in ws.columns:
            max_length = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=0
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='cash2switch_renewals_template.xlsx',
        )
    except Exception as e:
        current_app.logger.exception(f"Template download failed: {e}")
        return jsonify({'error': 'Failed to generate template'}), 500


@import_bp.route('/leads/template', methods=['GET'])
@token_required
def download_leads_template():
    return download_leads_template_handler()


# ---------------------------------------------------------------------------
# Sequence reset
# ---------------------------------------------------------------------------

@import_bp.route('/energy-clients/reset-sequence', methods=['POST'])
@jwt_required()
def reset_energy_client_sequence():
    current_user = get_jwt_identity()
    tenant_id    = current_user.get('tenant_id')

    session = SessionLocal()
    try:
        session.execute(text("""
            SELECT setval(
                pg_get_serial_sequence('"StreemLyne_MT"."Client_Master"', 'client_id'),
                COALESCE((
                    SELECT MAX(client_id) FROM "StreemLyne_MT"."Client_Master"
                    WHERE tenant_id = :tid
                ), 0), true
            )
        """), {'tid': tenant_id})
        session.execute(text("""
            SELECT setval(
                pg_get_serial_sequence('"StreemLyne_MT"."Project_Details"', 'project_id'),
                COALESCE((SELECT MAX(project_id) FROM "StreemLyne_MT"."Project_Details"), 0), true
            )
        """))
        session.execute(text("""
            SELECT setval(
                pg_get_serial_sequence(
                    '"StreemLyne_MT"."Energy_Contract_Master"',
                    'energy_contract_master_id'
                ),
                COALESCE((
                    SELECT MAX(energy_contract_master_id)
                    FROM "StreemLyne_MT"."Energy_Contract_Master"
                ), 0), true
            )
        """))
        session.commit()
        return jsonify({'message': 'All sequences reset successfully', 'success': True}), 200
    except Exception as e:
        session.rollback()
        logger.error(f"Error resetting sequences: {str(e)}")
        return jsonify({'error': 'Failed to reset sequences'}), 500
    finally:
        session.close()